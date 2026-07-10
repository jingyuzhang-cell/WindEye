#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
监管违规穿透式查询系统 v4
================================================================
解决 v3 的两大核心问题：

【问题1：社区匹配过宽（全部社区都被选中）】
根因：
  - 2-gram 扫描命中率极高（"证"、"监"等高频字无处不在）
  - 扩展关键词（"证券法"、"行政处罚"）在所有社区中普遍存在
修复：
  ① 去掉 2-gram 扫描，改用精确词汇分级匹配（核心词 > 扩展词）
  ② 对每个社区计算关联分（词频加权 + 关键词命中数），取 Top-K 社区
  ③ 用 DeepSeek 对候选社区摘要做语义相关性打分（0-10分），只保留高分社区
  ④ 废除"兜底取最大社区"策略

【问题2：路径无实质意义（2跳通用连接）】
根因：
  - 锚点选取仅按度数排名，"上市公司"度数最高但最不具体
  - BFS 找最短路径，2跳通用路径优先被返回
  - 没有路径质量评估机制
修复：
  ① LLM 锚点精选：将候选节点名称列表发给 DeepSeek，让它选出最与事件相关的节点
  ② 路径最小长度约束：合规/违规路径至少 3 跳（保证有中间语义节点）
  ③ 路径质量评分：评估路径中间节点与事件关键词的语义覆盖度
  ④ 通用节点黑名单：过滤"上市公司"、"本所"、"违反本办法"等作为端点
================================================================
"""

import json
import os
import re
import time
import warnings
from collections import defaultdict, deque
from typing import Dict, List, Set, Tuple, Optional, Any

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import networkx as nx
import numpy as np
import pandas as pd
import requests

warnings.filterwarnings('ignore')

plt.rcParams['font.sans-serif'] = ['SimHei', 'DejaVu Sans', 'Arial Unicode MS', 'sans-serif']
plt.rcParams['axes.unicode_minus'] = False

# ==================== 配置 ====================
API_BASE   = "https://api.deepseek.com/v1"
API_KEY    = "sk-0a57f72b50854ace9d134a5eb697c4dc"
DATA_DIR   = "data"
OUTPUT_DIR = "v4_output"

# ---- 社区匹配参数 ----
COMMUNITY_CANDIDATE_TOPK  = 8     # 词汇评分后保留的候选社区数
COMMUNITY_SCORE_THRESHOLD = 5     # LLM打分阈值（0-10），低于此分数的社区被过滤
# True=跳过LLM精筛（快速但精度低），False=使用LLM精筛（推荐）
FAST_MODE = False

# ---- 节点锚点参数 ----
ANCHOR_CANDIDATE_SIZE = 20   # 送给LLM精选前的候选池大小
ANCHOR_TOPK           = 8    # 最终保留的锚点数

# ---- 路径搜索参数 ----
BFS_MAX_DEPTH       = 10
COMPLIANCE_MIN_HOPS = 3      # 合规路径最少跳数（过滤2跳通用连接）
VIOLATION_MIN_HOPS  = 3      # 违规路径最少跳数
MAX_PATHS_PER_QUERY = 5

# ---- 通用节点黑名单（不应作为路径端点）----
GENERIC_NODE_BLACKLIST = {
    '上市公司', '发行人', '本所', '本办法', '违反本办法',
    '公告', '报告', '审议通过', '相关规定', '有关规定',
}

# ---- 节点类型分层 ----
LAYER_TYPES = {
    'responsibility': {'PartyWithResponsibility', 'AdvantageHolder', 'Actor'},
    'violation':      {'Action', 'Means'},
    'regulatory':     {'RegulatoryAuthority'},
    'legal':          {'Section', 'Chapter', 'Title', 'Law'},
    'other':          {'Restriction', 'Event', 'Penalty', 'Obligation'},
}
TYPE_TO_LAYER: Dict[str, str] = {}
for _layer, _types in LAYER_TYPES.items():
    for _t in _types:
        TYPE_TO_LAYER[_t] = _layer

PERSPECTIVE_COLOR = {
    'responsibility': '#E53935',
    'violation':      '#FF8F00',
    'regulatory':     '#1E88E5',
    'legal':          '#43A047',
    'other':          '#8E24AA',
}
PERSPECTIVE_ZH = {
    'responsibility': '责任方',
    'violation':      '违规行为',
    'regulatory':     '监管机构',
    'legal':          '法规条款',
    'other':          '其他',
}
PATH_COLOR = {
    'compliance':        '#1565C0',
    'violation':         '#B71C1C',
    'violation_partial': '#E65100',
}


# ==================== 工具 ====================

def remove_markdown(text: str) -> str:
    if not text:
        return text
    text = re.sub(r'```[\w]*\n?', '', text)
    text = re.sub(r'```', '', text)
    text = re.sub(r'\*\*(.+?)\*\*', r'\1', text)
    text = re.sub(r'__(.+?)__', r'\1', text)
    text = re.sub(r'\*(.+?)\*', r'\1', text)
    text = re.sub(r'^#{1,6}\s+', '', text, flags=re.MULTILINE)
    text = re.sub(r'\[([^\]]+)\]\([^\)]+\)', r'\1', text)
    text = re.sub(r'`([^`]+)`', r'\1', text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def call_deepseek(prompt: str, max_tokens: int = 1500,
                  temperature: float = 0.1, retries: int = 3) -> str:
    headers = {"Authorization": f"Bearer {API_KEY}",
               "Content-Type": "application/json"}
    payload = {
        "model": "deepseek-chat",
        "messages": [{"role": "user", "content": prompt}],
        "max_tokens": max_tokens,
        "temperature": temperature,
    }
    for attempt in range(retries):
        try:
            resp = requests.post(f"{API_BASE}/chat/completions",
                                 headers=headers, json=payload, timeout=90)
            if resp.status_code == 200:
                return remove_markdown(
                    resp.json()['choices'][0]['message']['content'])
            print(f"    API 状态码: {resp.status_code}")
        except Exception as e:
            print(f"    API 调用异常 (尝试{attempt+1}): {e}")
        time.sleep(4)
    raise RuntimeError("DeepSeek API 调用失败")


def parse_json_safe(text: str) -> Any:
    """安全解析 JSON，优先解析 {} 再解析 []"""
    for start_char, end_char in [('{', '}'), ('[', ']')]:
        s = text.find(start_char)
        e = text.rfind(end_char) + 1
        if s >= 0 and e > s:
            try:
                return json.loads(text[s:e])
            except Exception:
                pass
    return None


# ==================== 数据加载 ====================

class DataLoader:
    def __init__(self, data_dir: str = DATA_DIR):
        self.data_dir      = data_dir
        self.vis:          Dict[str, dict]                  = {}
        self.reports:      Dict[str, pd.DataFrame]          = {}
        self.hierarchy:    pd.DataFrame                     = pd.DataFrame()
        self.kg_node_map:  Dict[str, dict]                  = {}
        self.kg_edges:     List[dict]                       = []
        self.G:            nx.DiGraph                       = nx.DiGraph()
        self.G_ud:         nx.Graph                         = nx.Graph()
        self.comm_nodes:   Dict[str, Dict[int, List[str]]]  = {}
        self.node_comm:    Dict[str, Dict[str, int]]        = {}
        self.comm_summary: Dict[str, Dict[int, str]]        = {}

    def load(self):
        d = self.data_dir
        print("正在加载数据...")
        for p in ['responsibility', 'violation', 'regulatory']:
            with open(f"{d}/{p}_visualization_data.json", 'r', encoding='utf-8') as f:
                self.vis[p] = json.load(f)

        for p, fname in [
            ('responsibility', '责任方社区报告.xlsx'),
            ('violation',      '违规行为社区报告.xlsx'),
            ('regulatory',     '监管机构社区报告.xlsx'),
        ]:
            self.reports[p] = pd.read_excel(f"{d}/{fname}")

        self.hierarchy = pd.read_excel(f"{d}/community_hierarchy_v3_fixed.xlsx")

        with open(f"{d}/merged_regulatory_unified.txt", 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                data = json.loads(line)
                if data['type'] == 'node':
                    self.kg_node_map[data['id']] = data
                elif data['type'] == 'relationship':
                    self.kg_edges.append(data)

        self._build_graph()
        self._build_community_index()
        self._build_summary_index()
        print(f"加载完成：{len(self.kg_node_map)} 节点  "
              f"{self.G.number_of_edges()} 有向边")

    def _build_graph(self):
        for node in self.kg_node_map.values():
            ntype = node['labels'][0] if node['labels'] else 'Unknown'
            name  = node['properties'].get('name', '')
            self.G.add_node(node['id'], name=name, node_type=ntype)
        for rel in self.kg_edges:
            sid = rel['start']['id']
            eid = rel['end']['id']
            if sid in self.G and eid in self.G:
                self.G.add_edge(sid, eid, rel_type=rel['label'])
        self.G_ud = self.G.to_undirected()

    def _build_community_index(self):
        for p in ['responsibility', 'violation', 'regulatory']:
            self.comm_nodes[p] = defaultdict(list)
            self.node_comm[p]  = {}
            for vnode in self.vis[p]['nodes']:
                nid = vnode['id']
                cid = int(vnode['community'])
                self.comm_nodes[p][cid].append(nid)
                self.node_comm[p][nid] = cid

    def _build_summary_index(self):
        for p in ['responsibility', 'violation', 'regulatory']:
            self.comm_summary[p] = {}
            df = self.reports[p]
            for _, row in df.iterrows():
                cid     = int(row['community'])
                title   = str(row.get('title', ''))
                summary = str(row.get('summary', ''))
                kw      = str(row.get('key_words', ''))
                self.comm_summary[p][cid] = f"{title}。{summary}。关键词：{kw}"

    def node_name(self, nid: str) -> str:
        n = self.kg_node_map.get(nid, {})
        return n.get('properties', {}).get('name', nid[:16]) if n else nid[:16]

    def node_type(self, nid: str) -> str:
        n = self.kg_node_map.get(nid, {})
        return (n.get('labels', ['?'])[0] if n else '?')

    def node_layer(self, nid: str) -> str:
        return TYPE_TO_LAYER.get(self.node_type(nid), 'other')

    def community_all_node_names(self, perspective: str, cid: int) -> List[str]:
        return [self.node_name(nid)
                for nid in self.comm_nodes[perspective].get(cid, [])]


# ==================== 精准实体提取 ====================

class EntityExtractor:
    def __init__(self, loader: DataLoader):
        self.loader = loader

    def extract(self, event_desc: str) -> Dict[str, List[str]]:
        print("\n[步骤1] 提取结构化实体关键词...")
        prompt = f"""你是资本市场法规合规专家。请从下列事件中提取关键实体，只给出知识图谱中会出现的规范法律术语（不要给出人名/公司名等具体实体，要给出其法律分类名称）。

事件描述：
{event_desc}

返回严格 JSON（不含其他内容）：
{{
  "责任方":   ["法律意义上的责任主体类型术语1", "类型术语2"],
  "违规行为": ["违规行为类型或手段的规范术语1", "术语2"],
  "监管机构": ["有权监管此类事件的机构名称1", "名称2"],
  "核心法律概念": ["本事件涉及的核心法律概念1", "概念2", "概念3"]
}}

注意：
- 责任方举例："内幕信息知情人"、"持股5%%以上股东"、"董事、监事、高级管理人员"、"实际控制人"
- 违规行为举例："内幕交易"、"短线交易"、"操纵证券市场"、"虚假陈述"、"信息披露违规"
- 监管机构举例："中国证监会"、"证券交易所"、"派出机构"
- 核心法律概念举例："内幕信息"、"重大事项"、"公开前"、"利用"""
        try:
            raw    = call_deepseek(prompt, max_tokens=500)
            result = parse_json_safe(raw) or {}
        except Exception as ex:
            print(f"  提取失败: {ex}")
            result = {}

        all_kw: List[str] = []
        for key in ["责任方", "违规行为", "监管机构", "核心法律概念"]:
            all_kw.extend(result.get(key, []))
        result['_all'] = list(dict.fromkeys(all_kw))

        print(f"  责任方词:     {result.get('责任方', [])}")
        print(f"  违规行为词:   {result.get('违规行为', [])}")
        print(f"  监管机构词:   {result.get('监管机构', [])}")
        print(f"  核心法律概念: {result.get('核心法律概念', [])}")
        return result


# ==================== 精准社区匹配（词汇评分 + LLM精筛）====================

class PrecisionCommunityMatcher:
    """
    两阶段精准社区匹配：
    阶段1：词汇分级打分（高区分度词优先）→ Top-K 候选
    阶段2：LLM 语义相关性打分（0-10）→ 过滤低分社区
    """

    # 通用高频词黑名单（出现在所有社区，不提供区分度）
    HIGH_FREQ_BLACKLIST = {
        '证券', '公司', '规定', '法律', '法规', '本办法', '本规定',
        '依法', '依据', '应当', '不得', '行为', '处理', '相关',
        '违反', '违规', '合规', '管理', '监督', '机构',
    }

    def __init__(self, loader: DataLoader):
        self.loader = loader

    def _score_community(self, perspective: str, cid: int,
                         core_kws: List[str], general_kws: List[str]) -> float:
        summary    = self.loader.comm_summary[perspective].get(cid, '')
        node_names = self.loader.community_all_node_names(perspective, cid)
        full_text  = summary + ' ' + ' '.join(node_names)
        score = 0.0
        # 核心关键词：完整匹配得3分，3字以上子串匹配得0.8分
        for kw in core_kws:
            if len(kw) < 2 or kw in self.HIGH_FREQ_BLACKLIST:
                continue
            if kw in full_text:
                score += 3.0
            else:
                for sub_len in range(len(kw), 2, -1):
                    for i in range(len(kw) - sub_len + 1):
                        sub = kw[i:i+sub_len]
                        if sub not in self.HIGH_FREQ_BLACKLIST and sub in full_text:
                            score += 0.8
                            break
                    else:
                        continue
                    break
        # 普通关键词：完整匹配得1分
        for kw in general_kws:
            if len(kw) < 2 or kw in self.HIGH_FREQ_BLACKLIST:
                continue
            if kw in full_text:
                score += 1.0
        return score

    def match_candidates(self, entities: Dict[str, List[str]],
                         topk: int = COMMUNITY_CANDIDATE_TOPK
                         ) -> Dict[str, List[Tuple[int, float]]]:
        kw_map = {
            'responsibility': {
                'core':    entities.get('责任方', []),
                'general': entities.get('核心法律概念', []),
            },
            'violation': {
                'core':    entities.get('违规行为', []),
                'general': entities.get('核心法律概念', []),
            },
            'regulatory': {
                'core':    entities.get('监管机构', []),
                'general': entities.get('违规行为', []),
            },
        }
        candidates: Dict[str, List[Tuple[int, float]]] = {}
        for p in ['responsibility', 'violation', 'regulatory']:
            scored = []
            for cid in self.loader.comm_nodes[p].keys():
                sc = self._score_community(
                    p, cid,
                    kw_map[p]['core'],
                    kw_map[p]['general'],
                )
                if sc > 0:
                    scored.append((cid, sc))
            scored.sort(key=lambda x: x[1], reverse=True)
            candidates[p] = scored[:topk]
        return candidates

    def llm_filter(self, candidates: Dict[str, List[Tuple[int, float]]],
                   event_desc: str,
                   threshold: float = COMMUNITY_SCORE_THRESHOLD
                   ) -> Dict[str, Set[int]]:
        matched: Dict[str, Set[int]] = {
            'responsibility': set(),
            'violation':      set(),
            'regulatory':     set(),
        }
        if FAST_MODE:
            for p, scored in candidates.items():
                matched[p] = {cid for cid, _ in scored}
            return matched

        pname = {'responsibility': '责任方',
                 'violation':      '违规行为',
                 'regulatory':     '监管机构'}
        for p, scored in candidates.items():
            if not scored:
                continue
            comm_info_lines = []
            for cid, pre_score in scored:
                summary    = self.loader.comm_summary[p].get(cid, '')[:150]
                node_names = self.loader.community_all_node_names(p, cid)[:8]
                comm_info_lines.append(
                    f"  社区#{cid}（预评分{pre_score:.1f}）：{summary}\n"
                    f"    代表节点：{', '.join(node_names)}")

            prompt = f"""你是资本市场法规专家。请评估以下每个{pname[p]}社区与事件的相关性（0-10分，10分最相关）。

【事件描述】
{event_desc}

【待评分的{pname[p]}社区】
{chr(10).join(comm_info_lines)}

评分标准：
- 9-10分：社区内容直接描述了事件涉及的{pname[p]}类型/行为/机构
- 7-8分：社区内容与事件存在明显关联
- 5-6分：社区内容与事件有一定关联
- 0-4分：社区内容与事件关联性弱

返回严格 JSON（社区编号->分数），不含其他内容：
{{"0": 7, "3": 9, "5": 2}}"""
            try:
                raw    = call_deepseek(prompt, max_tokens=300, temperature=0.1)
                scores = parse_json_safe(raw) or {}
                for cid_str, score in scores.items():
                    try:
                        cid   = int(cid_str)
                        score = float(score)
                        if score >= threshold:
                            matched[p].add(cid)
                    except (ValueError, TypeError):
                        pass
                print(f"  {pname[p]} LLM精筛分数: {scores}")
            except Exception as ex:
                print(f"  {pname[p]} LLM精筛失败: {ex}，使用词汇候选")
                matched[p] = {cid for cid, _ in scored}
        return matched

    def match(self, entities: Dict[str, List[str]],
              event_desc: str) -> Dict[str, Set[int]]:
        print("\n[步骤2] 精准社区匹配（词汇评分 + LLM精筛）...")
        candidates = self.match_candidates(entities)

        pname = {'responsibility': '责任方',
                 'violation':      '违规行为',
                 'regulatory':     '监管机构'}
        for p, scored in candidates.items():
            print(f"  {pname[p]} 词汇候选: "
                  f"{[(cid, round(sc,1)) for cid, sc in scored]}")

        matched = self.llm_filter(candidates, event_desc)

        print(f"  精筛后责任方社区:   {sorted(matched['responsibility'])}")
        print(f"  精筛后违规行为社区: {sorted(matched['violation'])}")
        print(f"  精筛后监管机构社区: {sorted(matched['regulatory'])}")

        matched = self._complement_missing(matched, candidates)
        return matched

    def _complement_missing(self,
                            matched:    Dict[str, Set[int]],
                            candidates: Dict[str, List[Tuple[int, float]]]
                            ) -> Dict[str, Set[int]]:
        """只在某层完全为空时才补全，且只取词汇分最高的1个"""
        h = self.loader.hierarchy

        def top_candidate(p: str) -> Optional[int]:
            lst = candidates.get(p, [])
            return lst[0][0] if lst else None

        for p in ['responsibility', 'violation', 'regulatory']:
            if matched[p]:
                continue
            # 先尝试层级关系推导
            other_ps = [op for op in ['responsibility', 'violation', 'regulatory']
                        if op != p and matched[op]]
            found: Set[int] = set()
            for op in other_ps:
                rows = h[(h['source_perspective'] == op) &
                         (h['source_community_id'].isin(matched[op])) &
                         (h['target_perspective'] == p)]
                found.update(rows['target_community_id'].tolist())
                rows = h[(h['target_perspective'] == op) &
                         (h['target_community_id'].isin(matched[op])) &
                         (h['source_perspective'] == p)]
                found.update(rows['source_community_id'].tolist())
            if found:
                matched[p] = found
                print(f"  → 层级推导补全 {p}: {sorted(found)}")
            else:
                top = top_candidate(p)
                if top is not None:
                    matched[p] = {top}
                    print(f"  → 词汇兜底补全 {p}: 社区#{top}")
        return matched


# ==================== LLM 精选锚点节点 ====================

class LLMAnchorSelector:
    def __init__(self, loader: DataLoader):
        self.loader = loader

    def select(self, entities: Dict[str, List[str]],
               matched_communities: Dict[str, Set[int]],
               event_desc: str) -> Dict[str, List[str]]:
        print("\n[步骤3] LLM精选节点锚点...")
        anchors: Dict[str, List[str]] = {
            'responsibility': [],
            'violation':      [],
            'regulatory':     [],
        }
        for p in ['responsibility', 'violation', 'regulatory']:
            anchors[p] = self._select_for(p, matched_communities[p],
                                          entities, event_desc)
            names = [self.loader.node_name(n) for n in anchors[p][:6]]
            print(f"  {p} 锚点({len(anchors[p])}个): {names}")

        anchors = self._expand_empty(anchors)
        return anchors

    def _select_for(self, perspective: str, community_ids: Set[int],
                    entities: Dict[str, List[str]],
                    event_desc: str) -> List[str]:
        target_types = LAYER_TYPES[perspective]
        # 收集候选（去掉黑名单节点）
        candidates: List[Tuple[str, str, float]] = []
        for cid in community_ids:
            for nid in self.loader.comm_nodes[perspective].get(cid, []):
                if self.loader.node_type(nid) not in target_types:
                    continue
                name = self.loader.node_name(nid)
                if name in GENERIC_NODE_BLACKLIST:
                    continue
                candidates.append((nid, name, self.loader.G.degree(nid)))

        if not candidates:
            return []

        candidates.sort(key=lambda x: x[2], reverse=True)
        top = candidates[:ANCHOR_CANDIDATE_SIZE]

        if FAST_MODE or len(top) <= 3:
            return [nid for nid, _, _ in top[:ANCHOR_TOPK]]

        pname = {'responsibility': '责任方',
                 'violation':      '违规行为',
                 'regulatory':     '监管机构'}
        candidate_lines = [f"  [{i}] {name}（频次:{int(deg)}）"
                           for i, (nid, name, deg) in enumerate(top)]
        prompt = f"""你是资本市场法规专家。下列是知识图谱中的{pname[perspective]}类型节点，请选出与事件最相关的节点编号（最多选{ANCHOR_TOPK}个）。

【事件描述】
{event_desc}

【候选节点列表】
{chr(10).join(candidate_lines)}

选择标准：节点名称与事件描述的责任主体/违规行为/监管场景直接相关。
请优先选择具体的、专指性强的节点，避免选"上市公司"、"发行人"等过于通用的节点（除非它确实是唯一选项）。

返回严格 JSON，只含选中的编号列表，不含其他内容：
[0, 3, 7]"""
        try:
            raw     = call_deepseek(prompt, max_tokens=200, temperature=0.1)
            indices = parse_json_safe(raw)
            if isinstance(indices, list):
                selected = [top[idx][0] for idx in indices
                            if isinstance(idx, int) and 0 <= idx < len(top)]
                if selected:
                    return selected[:ANCHOR_TOPK]
        except Exception as ex:
            print(f"    LLM锚点选取失败({perspective}): {ex}")

        return [nid for nid, _, _ in top[:ANCHOR_TOPK]]

    def _expand_empty(self, anchors: Dict[str, List[str]]) -> Dict[str, List[str]]:
        G = self.loader.G_ud
        for p in ['responsibility', 'violation', 'regulatory']:
            if anchors[p]:
                continue
            target_types = LAYER_TYPES[p]
            found: List[Tuple[str, float]] = []
            for op in ['responsibility', 'violation', 'regulatory']:
                if op == p:
                    continue
                for nid in anchors[op]:
                    for nbr in G.neighbors(nid):
                        if self.loader.node_type(nbr) in target_types:
                            name = self.loader.node_name(nbr)
                            if name not in GENERIC_NODE_BLACKLIST:
                                found.append((nbr, self.loader.G.degree(nbr)))
            found.sort(key=lambda x: x[1], reverse=True)
            anchors[p] = [nid for nid, _ in found[:ANCHOR_TOPK]]
            if anchors[p]:
                names = [self.loader.node_name(n) for n in anchors[p][:4]]
                print(f"  → {p} 邻居扩展锚点: {names}")
        return anchors


# ==================== 高质量路径搜索 ====================

class QualityPathFinder:
    """
    高质量路径搜索器

    改进：
    1. 最小跳数约束（过滤无意义2跳通用连接）
    2. 路径质量评分（中间节点语义覆盖度）
    3. 通用端点过滤
    4. 多轮搜索（有向严格→有向松弛→无向兜底）
    """

    BRIDGE_TYPES = LAYER_TYPES['legal'] | LAYER_TYPES['other']

    def __init__(self, loader: DataLoader, event_keywords: List[str]):
        self.loader  = loader
        self.G       = loader.G
        self.G_ud    = loader.G_ud
        self.kw_set  = {kw.lower() for kw in event_keywords if len(kw) >= 2}

    def _path_score(self, node_ids: List[str]) -> float:
        if len(node_ids) < 2:
            return 0.0
        score = 0.0
        # 端点黑名单惩罚
        for ep in [node_ids[0], node_ids[-1]]:
            if self.loader.node_name(ep) in GENERIC_NODE_BLACKLIST:
                score -= 3.0
        # 中间节点语义覆盖
        for nid in node_ids[1:-1]:
            name = self.loader.node_name(nid).lower()
            for kw in self.kw_set:
                if kw in name or name in kw:
                    score += 2.0
                    break
        # 层次覆盖奖励
        layers = {self.loader.node_layer(nid) for nid in node_ids}
        score += len(layers) * 1.5
        # 路径长度奖励
        score += min(len(node_ids) - 2, 4) * 0.5
        return score

    def _bfs(self, source_ids: List[str], target_ids: List[str],
             min_hops: int = 2, max_depth: int = BFS_MAX_DEPTH,
             use_undirected: bool = False) -> List[List[str]]:
        if not source_ids or not target_ids:
            return []
        G       = self.G_ud if use_undirected else self.G
        tgt_set = set(target_ids)
        found:  List[List[str]] = []
        seen:   Set[Tuple]      = set()

        valid_sources = [s for s in source_ids
                         if self.loader.node_name(s) not in GENERIC_NODE_BLACKLIST]
        if not valid_sources:
            valid_sources = source_ids

        for src in valid_sources[:ANCHOR_TOPK]:
            if src not in G:
                continue
            queue = deque([(src, [src], {src})])
            while queue:
                cur, path, visited = queue.popleft()
                if len(path) > max_depth + 1:
                    continue
                hops = len(path) - 1
                if cur in tgt_set and hops >= min_hops:
                    if (self.loader.node_name(cur) not in GENERIC_NODE_BLACKLIST
                            or len(found) == 0):
                        key = tuple(path)
                        if key not in seen:
                            seen.add(key)
                            found.append(path)
                    continue
                if hops >= max_depth:
                    continue
                nbrs = (list(G.neighbors(cur)) if use_undirected
                        else list(G.successors(cur)))
                for nxt in nbrs:
                    if nxt in visited:
                        continue
                    # 剪枝：相同类型节点不连续出现（桥接节点除外）
                    if len(path) >= 3:
                        last_type = self.loader.node_type(path[-1])
                        nxt_type  = self.loader.node_type(nxt)
                        if (nxt_type == last_type
                                and nxt_type not in self.BRIDGE_TYPES
                                and nxt not in tgt_set):
                            continue
                    queue.append((nxt, path + [nxt], visited | {nxt}))
        return found

    def _select_best(self, paths: List[List[str]], n: int) -> List[List[str]]:
        scored = [(p, self._path_score(p)) for p in paths]
        scored.sort(key=lambda x: x[1], reverse=True)
        seen:   Set[Tuple]   = set()
        result: List[List[str]] = []
        for p, _ in scored:
            key = tuple(p)
            if key not in seen:
                seen.add(key)
                result.append(p)
            if len(result) >= n:
                break
        return result

    def find_compliance_paths(self, resp_anchors: List[str],
                               reg_anchors: List[str]) -> List[dict]:
        paths = self._bfs(resp_anchors, reg_anchors,
                          min_hops=COMPLIANCE_MIN_HOPS, max_depth=7)
        if not paths:
            paths = self._bfs(resp_anchors, reg_anchors,
                              min_hops=COMPLIANCE_MIN_HOPS,
                              max_depth=BFS_MAX_DEPTH, use_undirected=True)
        if not paths:
            # 放宽最小跳数（兜底）
            paths = self._bfs(resp_anchors, reg_anchors,
                              min_hops=2, max_depth=BFS_MAX_DEPTH,
                              use_undirected=True)
        best = self._select_best(paths, MAX_PATHS_PER_QUERY)
        return [self._to_record(p, 'compliance') for p in best]

    def find_violation_paths(self, resp_anchors: List[str],
                              viol_anchors: List[str],
                              reg_anchors: List[str]) -> List[dict]:
        viol_set = set(viol_anchors)
        # 段1: 责任方 → 违规行为
        seg1 = self._bfs(resp_anchors, viol_anchors, min_hops=1, max_depth=5)
        if not seg1:
            seg1 = self._bfs(resp_anchors, viol_anchors, min_hops=1,
                             max_depth=7, use_undirected=True)
        # 段2: 违规行为 → 监管机构
        seg2 = self._bfs(viol_anchors, reg_anchors, min_hops=1, max_depth=5)
        if not seg2:
            seg2 = self._bfs(viol_anchors, reg_anchors, min_hops=1,
                             max_depth=7, use_undirected=True)
        # 拼接
        results: List[List[str]] = []
        seen:    Set[Tuple]      = set()
        for p1 in seg1:
            for p2 in seg2:
                if p1[-1] == p2[0]:
                    combined = p1 + p2[1:]
                    if len(combined) - 1 >= VIOLATION_MIN_HOPS:
                        key = tuple(combined)
                        if key not in seen:
                            seen.add(key)
                            results.append(combined)
        # 直接端到端搜索（路径中必须含违规节点）
        if not results:
            direct = self._bfs(resp_anchors, reg_anchors,
                               min_hops=VIOLATION_MIN_HOPS,
                               max_depth=BFS_MAX_DEPTH, use_undirected=True)
            for p in direct:
                if any(n in viol_set for n in p[1:-1]):
                    key = tuple(p)
                    if key not in seen:
                        seen.add(key)
                        results.append(p)
        # 保底：返回段1（责任方→违规行为半段）
        if not results and seg1:
            best_seg1 = self._select_best(seg1, 2)
            return [self._to_record(p, 'violation_partial') for p in best_seg1]

        best = self._select_best(results, MAX_PATHS_PER_QUERY)
        return [self._to_record(p, 'violation') for p in best]

    def _to_record(self, node_ids: List[str], ptype: str) -> dict:
        edges = []
        for i in range(len(node_ids) - 1):
            src, tgt = node_ids[i], node_ids[i+1]
            data = self.G.get_edge_data(src, tgt)
            if data is None:
                data = self.G.get_edge_data(tgt, src)
                rt   = (data or {}).get('rel_type', '关联')
                edges.append((src, tgt, f"{rt}(↔)"))
            else:
                edges.append((src, tgt, data.get('rel_type', '关联')))
        return {
            'nodes': node_ids,
            'edges': edges,
            'type':  ptype,
            'score': round(self._path_score(node_ids), 2),
        }


# ==================== 可视化 ====================

def visualize_community_network(loader: DataLoader,
                                matched: Dict[str, Set[int]],
                                comm_paths: List[Tuple],
                                out_file: str):
    G = nx.DiGraph()
    colors, sizes, labels = [], [], {}
    for p, comms in matched.items():
        for cid in comms:
            nid = f"{p}_{cid}"
            G.add_node(nid)
            colors.append(PERSPECTIVE_COLOR.get(p, '#999'))
            sizes.append(3200)
            labels[nid] = f"{PERSPECTIVE_ZH.get(p,p)}\n社区#{cid}"
    for sp, si, tp, ti, rt, strong in comm_paths:
        sn, tn = f"{sp}_{si}", f"{tp}_{ti}"
        if sn in G and tn in G:
            G.add_edge(sn, tn, width=3 if strong else 1)

    fig, ax = plt.subplots(figsize=(16, 10))
    if len(G) == 0:
        ax.text(0.5, 0.5, '未找到匹配社区', ha='center', va='center', fontsize=16)
    else:
        pos    = nx.spring_layout(G, k=2.5, iterations=80, seed=42)
        nx.draw_networkx_nodes(G, pos, node_color=colors,
                               node_size=sizes, alpha=0.88, ax=ax)
        es     = list(G.edges())
        widths = [G[u][v].get('width', 1) for u, v in es]
        nx.draw_networkx_edges(G, pos, width=widths, alpha=0.6,
                               arrows=True, arrowsize=22, ax=ax)
        nx.draw_networkx_labels(G, pos, labels, font_size=9, ax=ax)
        legend_handles = [
            mpatches.Patch(color=PERSPECTIVE_COLOR[p],
                           label=PERSPECTIVE_ZH[p])
            for p in ['responsibility', 'violation', 'regulatory']
        ]
        ax.legend(handles=legend_handles, loc='upper left', fontsize=10)
    ax.set_title("社区级责任递进路径（v4）", fontsize=14, pad=12)
    ax.axis('off')
    plt.tight_layout()
    plt.savefig(out_file, dpi=180, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f"  ✓ 社区级: {out_file}")


def visualize_node_paths(loader: DataLoader,
                         node_paths: Dict[str, List],
                         out_file: str):
    comp_paths = node_paths.get('compliance', [])
    viol_paths = (node_paths.get('violation', []) +
                  node_paths.get('violation_partial', []))
    n_rows = max(len(comp_paths), len(viol_paths), 1)

    fig, axes = plt.subplots(n_rows, 2,
                              figsize=(26, max(5 * n_rows, 5)))
    if n_rows == 1:
        axes = axes.reshape(1, 2)
    fig.suptitle("节点级穿透路径可视化（v4）\n蓝=合规路径  红=违规路径",
                 fontsize=15, fontweight='bold', y=1.01)

    def get_color(nid: str) -> str:
        return PERSPECTIVE_COLOR.get(loader.node_layer(nid), '#9E9E9E')

    def draw(ax, rec: Optional[dict], title: str, arrow_color: str):
        ax.axis('off')
        ax.set_title(title, fontsize=10, color=arrow_color,
                     pad=6, fontweight='bold')
        if rec is None:
            ax.text(0.5, 0.5, '（无路径）', ha='center', va='center',
                    fontsize=10, color='#BDBDBD')
            return
        nodes = rec['nodes']
        edges = rec['edges']
        n     = len(nodes)
        if n == 0:
            return
        xs     = np.linspace(0.06, 0.94, n)
        y_node = 0.58
        y_lbl  = 0.20
        for i, (s, t, rt) in enumerate(edges):
            ax.annotate('', xy=(xs[i+1]-0.025, y_node),
                        xytext=(xs[i]+0.025, y_node),
                        arrowprops=dict(arrowstyle='->',
                                        color=arrow_color,
                                        lw=2.2, mutation_scale=18))
            mid_x = (xs[i] + xs[i+1]) / 2
            ax.text(mid_x, y_node+0.15, rt,
                    ha='center', va='center', fontsize=7.5,
                    color=arrow_color,
                    bbox=dict(boxstyle='round,pad=0.25',
                              facecolor='white',
                              edgecolor=arrow_color,
                              alpha=0.9, linewidth=1.2))
        for i, nid in enumerate(nodes):
            x     = xs[i]
            color = get_color(nid)
            layer = loader.node_layer(nid)
            circle = plt.Circle((x, y_node), 0.048,
                                  color=color, zorder=5, alpha=0.92)
            ax.add_patch(circle)
            ax.text(x, y_node+0.005,
                    PERSPECTIVE_ZH.get(layer, '?')[:3],
                    ha='center', va='center', fontsize=6.5,
                    color='white', fontweight='bold', zorder=6)
            name   = loader.node_name(nid)
            name_d = name[:14] + '..' if len(name) > 14 else name
            ax.text(x, y_lbl,
                    f"{name_d}\n({loader.node_type(nid)})",
                    ha='center', va='top', fontsize=7.2,
                    bbox=dict(boxstyle='round,pad=0.3',
                              facecolor='#FAFAFA',
                              edgecolor='#CCCCCC', alpha=0.95))
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)

    ptype_zh = {'violation': '违规', 'violation_partial': '违规(半段)'}
    for row_i in range(n_rows):
        cr = comp_paths[row_i] if row_i < len(comp_paths) else None
        vr = viol_paths[row_i] if row_i < len(viol_paths) else None
        sc = f"  质量分:{cr['score']}" if cr else ""
        sv = f"  质量分:{vr['score']}" if vr else ""
        ct = (f"合规路径#{row_i+1}（{len(cr['nodes'])}节点）{sc}"
              if cr else f"合规路径#{row_i+1}")
        vt_pfx = ptype_zh.get((vr or {}).get('type', ''), '违规')
        vt = (f"{vt_pfx}路径#{row_i+1}（{len(vr['nodes'])}节点）{sv}"
              if vr else f"违规路径#{row_i+1}")
        draw(axes[row_i][0], cr, ct, PATH_COLOR['compliance'])
        draw(axes[row_i][1], vr, vt,
             PATH_COLOR.get((vr or {}).get('type', 'violation'),
                            PATH_COLOR['violation']))

    legend_handles = [
        mpatches.Patch(color=PERSPECTIVE_COLOR[p],
                       label=f"{PERSPECTIVE_ZH[p]}节点")
        for p in ['responsibility', 'violation', 'regulatory', 'legal', 'other']
    ]
    fig.legend(handles=legend_handles, loc='lower center', ncol=5,
               fontsize=9, bbox_to_anchor=(0.5, -0.03))
    plt.tight_layout()
    plt.savefig(out_file, dpi=180, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f"  ✓ 节点级: {out_file}")


# ==================== 社区级路径 ====================

def build_community_paths(matched: Dict[str, Set[int]],
                          hierarchy: pd.DataFrame) -> List[Tuple]:
    paths = []
    for _, row in hierarchy.iterrows():
        sp     = row['source_perspective']
        si     = int(row['source_community_id'])
        tp     = row['target_perspective']
        ti     = int(row['target_community_id'])
        rt     = row['relation_type']
        strong = bool(row['is_strong_link'])
        if (si in matched.get(sp, set()) and
                ti in matched.get(tp, set())):
            paths.append((sp, si, tp, ti, rt, strong))
    return paths


# ==================== Prompt + 文本构建 ====================

def build_prompt(event_desc: str, community_text: str,
                 triples_text: str, comm_paths: List[Tuple],
                 node_paths: Dict[str, List],
                 loader: DataLoader) -> str:
    pzh = {'responsibility': '责任方',
           'violation':      '违规行为',
           'regulatory':     '监管机构'}
    comm_lines = [
        f"  {pzh.get(sp,sp)}社区#{si} --[{rt},{'强' if s else '弱'}]--> "
        f"{pzh.get(tp,tp)}社区#{ti}"
        for sp, si, tp, ti, rt, s in comm_paths
    ]
    node_lines = []
    for ptype, plabel in [('compliance',        '合规'),
                          ('violation',         '违规'),
                          ('violation_partial', '违规(部分)')]:
        for i, rec in enumerate(node_paths.get(ptype, [])[:2], 1):
            step = ' → '.join(loader.node_name(n) for n in rec['nodes'])
            node_lines.append(
                f"  {plabel}路径{i}（质量分{rec['score']}）: {step}")
    return f"""你是资本市场法规合规领域资深专家，请对以下事件进行五维穿透式分析。

【事件描述】
{event_desc}

━━━━ 一、社区级责任链路（宏观）━━━━
{chr(10).join(comm_lines[:20]) if comm_lines else '（未找到社区级路径）'}

━━━━ 二、节点级穿透路径（微观）━━━━
{chr(10).join(node_lines) if node_lines else '（未找到节点级路径）'}

━━━━ 三、知识图谱原始三元组（精细法律依据）━━━━
{triples_text}

━━━━ 四、相关社区法规报告 ━━━━
{community_text}

━━━━ 分析要求（五个维度，纯文本不使用 Markdown）━━━━

1. 责任主体认定：具体指出责任主体类型，引用三元组中的关系为依据
2. 违规行为定性：判断违规类型，援引相关法条
3. 监管机构职责与处罚依据：说明哪些机构可介入，可触发哪些处罚
4. 合规义务梳理：基于合规路径列出应履行的合规义务
5. 风险提示与建议：具体可操作的合规建议及监管趋势提示"""


def get_community_text(loader: DataLoader,
                       matched: Dict[str, Set[int]]) -> str:
    parts = []
    for p, pname in [('responsibility', '责任方'),
                     ('violation',      '违规行为'),
                     ('regulatory',     '监管机构')]:
        comms = sorted(matched.get(p, set()))
        if not comms:
            continue
        parts.append(f"\n{'='*50}\n【{pname}社区报告】\n{'='*50}")
        for cid in comms[:4]:
            df   = loader.reports[p]
            rows = df[df['community'] == cid]
            if rows.empty:
                continue
            row = rows.iloc[0]
            parts.append(f"\n社区#{cid}：{row.get('title','')}")
            parts.append(f"摘要：{str(row.get('summary',''))[:250]}")
            try:
                findings = json.loads(row.get('findings', '[]'))
                for fi, f_item in enumerate(findings[:2], 1):
                    parts.append(
                        f"  发现{fi}：{f_item.get('summary','')} — "
                        f"{f_item.get('explanation','')[:100]}")
            except Exception:
                pass
    return '\n'.join(parts)


def get_triples_text(loader: DataLoader,
                     node_paths: Dict[str, List]) -> str:
    seen:  Set[Tuple] = set()
    lines: List[str]  = []
    ptype_zh = {'compliance':        '合规',
                'violation':         '违规',
                'violation_partial': '违规(部分)'}
    for ptype in ['compliance', 'violation', 'violation_partial']:
        recs = node_paths.get(ptype, [])
        if not recs:
            continue
        for rec in recs:
            lines.append(f"\n[{ptype_zh.get(ptype,ptype)}路径  "
                         f"质量分:{rec['score']}]")
            for src, tgt, rt in rec['edges']:
                key = (src, tgt, rt)
                if key in seen:
                    continue
                seen.add(key)
                lines.append(
                    f"  {loader.node_name(src)}（{loader.node_type(src)}）"
                    f" -[{rt}]-> "
                    f"{loader.node_name(tgt)}（{loader.node_type(tgt)}）")
    return '\n'.join(lines) if lines else '（未找到三元组）'


# ==================== 主系统 ====================

class RegulatoryQuerySystemV4:

    def __init__(self, data_dir: str = DATA_DIR):
        self.loader = DataLoader(data_dir)
        self.loader.load()

    def query(self, event_desc: str, output_dir: str = OUTPUT_DIR) -> dict:
        os.makedirs(output_dir, exist_ok=True)
        print("\n" + "=" * 70)
        print("  监管违规穿透式查询系统 v4")
        print("  精准社区匹配 + LLM锚点精选 + 质量路径过滤")
        print("=" * 70)

        # Step 1
        entities = EntityExtractor(self.loader).extract(event_desc)

        # Step 2
        matched = PrecisionCommunityMatcher(self.loader).match(
            entities, event_desc)

        # Step 3
        anchors = LLMAnchorSelector(self.loader).select(
            entities, matched, event_desc)

        # Step 4
        print("\n[步骤4] 高质量节点路径搜索...")
        finder     = QualityPathFinder(self.loader, entities.get('_all', []))
        comp_paths = finder.find_compliance_paths(
            anchors['responsibility'], anchors['regulatory'])
        viol_paths = finder.find_violation_paths(
            anchors['responsibility'], anchors['violation'],
            anchors['regulatory'])
        node_paths = {'compliance': comp_paths, 'violation': viol_paths}
        print(f"  合规路径: {len(comp_paths)} 条，"
              f"违规路径: {len(viol_paths)} 条")
        for ptype, paths in node_paths.items():
            for i, rec in enumerate(paths, 1):
                steps = ' → '.join(self.loader.node_name(n)
                                   for n in rec['nodes'])
                print(f"  {ptype}#{i}（{len(rec['nodes'])}跳 "
                      f"质量分{rec['score']}）: {steps[:120]}")

        # Step 5
        comm_paths = build_community_paths(matched, self.loader.hierarchy)
        print(f"\n[步骤5] 社区级路径: {len(comm_paths)} 条")

        # Step 6
        print("\n[步骤6] 生成可视化...")
        comm_vis = os.path.join(output_dir, "event_network.png")
        node_vis = os.path.join(output_dir, "node_level_path.png")
        visualize_community_network(
            self.loader, matched, comm_paths, comm_vis)
        visualize_node_paths(self.loader, node_paths, node_vis)

        # Step 7
        print("\n[步骤7] 调用 DeepSeek 生成综合分析...")
        community_text = get_community_text(self.loader, matched)
        triples_text   = get_triples_text(self.loader, node_paths)
        prompt         = build_prompt(
            event_desc, community_text, triples_text,
            comm_paths, node_paths, self.loader)
        try:
            final_answer = call_deepseek(
                prompt, max_tokens=4000, temperature=0.3)
        except Exception as e:
            final_answer = f"LLM 调用失败：{e}"
            print(f"  错误: {e}")

        # Step 8
        print("\n[步骤8] 保存输出...")
        self._save(output_dir, event_desc, entities, matched, anchors,
                   comm_paths, node_paths, triples_text, community_text,
                   final_answer, prompt)

        print("\n" + "=" * 70)
        print("  查询完成！")
        print("=" * 70)
        return dict(entities=entities, matched=matched, anchors=anchors,
                    comm_paths=comm_paths, node_paths=node_paths,
                    final_answer=final_answer)

    def _save(self, output_dir, event_desc, entities, matched, anchors,
              comm_paths, node_paths, triples_text, community_text,
              final_answer, prompt):
        ld = self.loader

        def wp(fname, content):
            with open(os.path.join(output_dir, fname), 'w',
                      encoding='utf-8') as f:
                f.write(content)

        lines = ["=" * 65, "节点级穿透路径报告（v4）", "=" * 65,
                 f"\n事件：{event_desc}\n"]
        for ptype, pname in [('compliance',        '合规路径'),
                              ('violation',         '违规路径'),
                              ('violation_partial', '违规路径(半段)')]:
            recs = node_paths.get(ptype, [])
            if not recs:
                continue
            lines.append(f"\n【{pname}】")
            for i, rec in enumerate(recs, 1):
                lines.append(f"  路径#{i}（{len(rec['nodes'])}跳  "
                              f"质量分:{rec['score']}）：")
                for src, tgt, rt in rec['edges']:
                    lines.append(
                        f"    {ld.node_name(src)}（{ld.node_type(src)}）"
                        f" -[{rt}]-> "
                        f"{ld.node_name(tgt)}（{ld.node_type(tgt)}）")
        wp("node_paths_report.txt", '\n'.join(lines))
        wp("community_reports.txt", community_text)
        wp("node_triples.txt", triples_text)
        wp("debug_prompt.txt", prompt)
        wp("final_analysis.txt",
           "=" * 65 + "\n监管违规综合分析报告（v4）\n" + "=" * 65 +
           f"\n\n【事件】\n{event_desc}\n\n【分析结果】\n{final_answer}\n")

        self._save_excel(output_dir, event_desc, entities, matched, anchors,
                         comm_paths, node_paths, final_answer)

        for fname in ['event_network.png', 'node_level_path.png',
                      'node_paths_report.txt', 'node_triples.txt',
                      'final_analysis.txt', 'analysis_summary_v4.xlsx']:
            if os.path.exists(os.path.join(output_dir, fname)):
                print(f"    ✓ {fname}")

    def _save_excel(self, output_dir, event_desc, entities, matched, anchors,
                    comm_paths, node_paths, final_answer):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb  = Workbook()
        ld  = self.loader
        pzh = {'responsibility': '责任方',
               'violation':      '违规行为',
               'regulatory':     '监管机构'}

        def hdr(ws, colors):
            for cell, color in zip(ws[1], colors):
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=color)

        # Sheet1: 概述
        ws1 = wb.active
        ws1.title = "事件概述"
        ws1.append(["监管违规穿透分析报告 v4"])
        ws1['A1'].font = Font(size=13, bold=True)
        ws1.append([])
        ws1.append(["事件描述", event_desc])
        ws1.append([])
        for k, label in [("责任方",     "责任方"),
                          ("违规行为",   "违规行为"),
                          ("监管机构",   "监管机构"),
                          ("核心法律概念", "核心法律概念")]:
            ws1.append([label, str(entities.get(k, []))])
        ws1.append([])
        ws1.append(["LLM分析摘要", final_answer[:2000]])
        ws1.column_dimensions['A'].width = 18
        ws1.column_dimensions['B'].width = 100

        # Sheet2: 精筛社区
        ws2 = wb.create_sheet("精筛社区")
        ws2.append(['视角', '社区ID', '社区标题', '摘要', '关键词'])
        hdr(ws2, ['1565C0'] * 5)
        for p in ['responsibility', 'violation', 'regulatory']:
            df = ld.reports[p]
            for cid in sorted(matched.get(p, set())):
                rows = df[df['community'] == cid]
                if rows.empty:
                    continue
                row = rows.iloc[0]
                ws2.append([pzh.get(p, p), cid,
                             row.get('title', ''),
                             str(row.get('summary', ''))[:300],
                             row.get('key_words', '')])
        for col, w in zip('ABCDE', [12, 8, 40, 80, 40]):
            ws2.column_dimensions[col].width = w

        # Sheet3: 节点锚点
        ws3 = wb.create_sheet("节点锚点")
        ws3.append(['视角', '节点ID', '名称', '类型', '层', '度数'])
        hdr(ws3, ['1B5E20'] * 6)
        for p in ['responsibility', 'violation', 'regulatory']:
            for nid in anchors.get(p, []):
                ws3.append([pzh.get(p, p), nid,
                             ld.node_name(nid),
                             ld.node_type(nid),
                             PERSPECTIVE_ZH.get(ld.node_layer(nid), '?'),
                             ld.G.degree(nid)])
        for col, w in zip('ABCDEF', [12, 28, 22, 18, 12, 8]):
            ws3.column_dimensions[col].width = w

        # Sheet4: 节点路径
        ws4 = wb.create_sheet("节点级路径")
        ws4.append(['路径类型', '路径#', '质量分', '步骤#',
                    '源节点名称', '源类型', '关系',
                    '目标节点名称', '目标类型'])
        hdr(ws4, ['B71C1C'] * 9)
        ptype_zh = {'compliance':        '合规',
                    'violation':         '违规',
                    'violation_partial': '违规(半段)'}
        for ptype in ['compliance', 'violation', 'violation_partial']:
            for pno, rec in enumerate(node_paths.get(ptype, []), 1):
                for step_i, (src, tgt, rt) in enumerate(rec['edges'], 1):
                    ws4.append([
                        ptype_zh.get(ptype, ptype), pno,
                        rec['score'], step_i,
                        ld.node_name(src), ld.node_type(src),
                        rt,
                        ld.node_name(tgt), ld.node_type(tgt),
                    ])
        for col, w in zip('ABCDEFGHI', [14, 6, 8, 6, 24, 20, 14, 24, 20]):
            ws4.column_dimensions[col].width = w

        # Sheet5: 社区级路径
        ws5 = wb.create_sheet("社区级路径")
        ws5.append(['序号', '源视角', '源社区', '目标视角',
                    '目标社区', '关系类型', '强度'])
        hdr(ws5, ['4A148C'] * 7)
        for i, (sp, si, tp, ti, rt, strong) in enumerate(comm_paths, 1):
            ws5.append([i, pzh.get(sp, sp), si,
                        pzh.get(tp, tp), ti, rt,
                        "强" if strong else "弱"])
        for col, w in zip('ABCDEFG', [6, 14, 10, 14, 10, 20, 6]):
            ws5.column_dimensions[col].width = w

        excel_path = os.path.join(output_dir, "analysis_summary_v4.xlsx")
        wb.save(excel_path)


# ==================== 入口 ====================

def main():
    print("=" * 70)
    print("  监管违规穿透式查询系统 v4")
    print("=" * 70)
    system = RegulatoryQuerySystemV4(data_dir=DATA_DIR)
    print("\n请输入事件描述（输入完成后按 Enter，再输入 END 结束）：")
    lines = []
    while True:
        line = input()
        if line.strip().upper() == 'END':
            break
        lines.append(line)
    event_desc = '\n'.join(lines).strip()
    if not event_desc:
        print("错误：事件描述不能为空！")
        return
    results = system.query(event_desc, output_dir=OUTPUT_DIR)
    print("\n" + "=" * 70)
    print("最终分析结果：")
    print("=" * 70)
    print(results['final_answer'])


if __name__ == "__main__":
    main()
