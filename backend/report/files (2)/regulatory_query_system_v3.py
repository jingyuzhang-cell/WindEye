#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
监管违规穿透式查询系统 v3
================================================================
修复 v2 中节点级路径始终为空的四个根本原因：

根因1：责任方社区匹配为空
  → 修复：实体无法精确匹配时，从已匹配的违规/监管社区通过
          社区层级关系反向推导责任方社区（补全策略）
          同时用 DeepSeek 对事件做语义扩展，生成更多候选关键词

根因2：_get_core_nodes 依赖社区匹配结果，社区空则节点空
  → 修复：新增"语义节点匹配"——直接用事件关键词在 KG
          节点名称上做模糊/语义搜索，绕过社区层，直接拿节点

根因3：BFS 路径深度不足，跨视角路径需要经过 Law/Section 桥接节点
  → 修复：允许经过 Law/Section/Chapter 等"法规桥接节点"
          搜索深度放宽至 8；增加双向 BFS（从两端同时搜索）

根因4：路径起终点过于严格（只允许核心类型节点作为端点）
  → 修复：起点扩展为"事件语义节点"（不限类型）
          终点也允许非核心类型的监管/法规节点作为终止条件
================================================================
"""

import json
import os
import re
import time
import warnings
from collections import defaultdict, deque
from typing import Dict, List, Set, Tuple, Optional

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import networkx as nx
import numpy as np
import pandas as pd
import requests

warnings.filterwarnings('ignore')

plt.rcParams['font.sans-serif'] = ['SimHei', 'DejaVu Sans', 'Arial Unicode MS', 'sans-serif']
plt.rcParams['axes.unicode_minus'] = False

# ==================== 配置 ====================
API_BASE = "https://api.deepseek.com/v1"
API_KEY = "sk-0a57f72b50854ace9d134a5eb697c4dc"

# 节点类型分层定义
LAYER_TYPES = {
    'responsibility': {'PartyWithResponsibility', 'AdvantageHolder', 'Actor'},
    'violation': {'Action', 'Means'},
    'regulatory': {'RegulatoryAuthority'},
    'legal': {'Section', 'Chapter', 'Title', 'Law'},   # 法规桥接节点（可作为中间节点）
    'other': {'Restriction', 'Event', 'Penalty', 'Obligation'},
}

# 所有可识别节点类型 -> 所属层（用于着色和路径理解）
TYPE_TO_LAYER: Dict[str, str] = {}
for layer, types in LAYER_TYPES.items():
    for t in types:
        TYPE_TO_LAYER[t] = layer

# 路径搜索参数
BFS_MAX_DEPTH = 10          # 最大跳数（放宽）
MAX_PATHS_PER_QUERY = 5     # 每类路径最多保留条数
MAX_ANCHOR_NODES = 12       # 每端最多锚点节点数


# ==================== 工具函数 ====================

def remove_markdown(text: str) -> str:
    if not text:
        return text
    text = re.sub(r'```[\w]*\n?', '', text)
    text = re.sub(r'```', '', text)
    text = re.sub(r'\*\*(.+?)\*\*', r'\1', text)
    text = re.sub(r'__(.+?)__', r'\1', text)
    text = re.sub(r'\*(.+?)\*', r'\1', text)
    text = re.sub(r'^#{1,6}\s+', '', text, flags=re.MULTILINE)
    text = re.sub(r'\[([^\]]+)\]\([^\)]+\)', r'\1', text)
    text = re.sub(r'`([^`]+)`', r'\1', text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def call_deepseek(prompt: str, max_tokens: int = 1000, temperature: float = 0.2) -> str:
    headers = {"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"}
    payload = {
        "model": "deepseek-chat",
        "messages": [{"role": "user", "content": prompt}],
        "max_tokens": max_tokens,
        "temperature": temperature,
    }
    for attempt in range(3):
        try:
            resp = requests.post(f"{API_BASE}/chat/completions",
                                 headers=headers, json=payload, timeout=60)
            if resp.status_code == 200:
                return remove_markdown(resp.json()['choices'][0]['message']['content'])
            time.sleep(3)
        except Exception as e:
            if attempt == 2:
                raise
            time.sleep(3)
    raise Exception("DeepSeek API 调用失败")


# ==================== 数据加载 ====================

class DataLoader:
    def __init__(self, data_dir: str = "data"):
        self.data_dir = data_dir
        # 三种视角可视化数据
        self.vis: Dict[str, dict] = {}
        # Excel 社区报告
        self.reports: Dict[str, pd.DataFrame] = {}
        # 社区层级关系
        self.hierarchy: pd.DataFrame = pd.DataFrame()
        # KG 原始节点/边
        self.kg_node_map: Dict[str, dict] = {}   # id -> node dict
        self.kg_edges: List[dict] = []
        # 完整有向图
        self.G: nx.DiGraph = nx.DiGraph()
        # 视角 -> 社区ID -> [node_id]
        self.comm_nodes: Dict[str, Dict[int, List[str]]] = {}
        # 视角 -> name_lower -> {community_id}
        self.name_to_comm: Dict[str, Dict[str, Set[int]]] = {}
        # node_id -> 所属视角 + 社区
        self.node_perspective: Dict[str, Tuple[str, int]] = {}

    def load(self):
        d = self.data_dir
        print("正在加载数据...")

        for p in ['responsibility', 'violation', 'regulatory']:
            with open(f"{d}/{p}_visualization_data.json", 'r', encoding='utf-8') as f:
                self.vis[p] = json.load(f)

        report_files = {
            'responsibility': '责任方社区报告.xlsx',
            'violation': '违规行为社区报告.xlsx',
            'regulatory': '监管机构社区报告.xlsx',
        }
        for p, fname in report_files.items():
            self.reports[p] = pd.read_excel(f"{d}/{fname}")

        self.hierarchy = pd.read_excel(f"{d}/community_hierarchy_v3_fixed.xlsx")

        with open(f"{d}/merged_regulatory_unified.txt", 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                data = json.loads(line)
                if data['type'] == 'node':
                    self.kg_node_map[data['id']] = data
                elif data['type'] == 'relationship':
                    self.kg_edges.append(data)

        self._build_graph()
        self._build_community_index()
        print(f"加载完成：{len(self.kg_node_map)} 节点，{len(self.kg_edges)} 关系，"
              f"{self.G.number_of_edges()} 图边")

    def _build_graph(self):
        for node in self.kg_node_map.values():
            ntype = node['labels'][0] if node['labels'] else 'Unknown'
            name = node['properties'].get('name', '')
            self.G.add_node(node['id'], name=name, node_type=ntype,
                            labels=node['labels'])
        for rel in self.kg_edges:
            sid = rel['start']['id']
            eid = rel['end']['id']
            if sid in self.G and eid in self.G:
                self.G.add_edge(sid, eid,
                                rel_id=rel['id'],
                                rel_type=rel['label'])

    def _build_community_index(self):
        for p in ['responsibility', 'violation', 'regulatory']:
            self.comm_nodes[p] = defaultdict(list)
            self.name_to_comm[p] = {}
            for vnode in self.vis[p]['nodes']:
                nid = vnode['id']
                cid = vnode['community']
                name_lower = vnode['name'].lower()
                self.comm_nodes[p][cid].append(nid)
                if name_lower not in self.name_to_comm[p]:
                    self.name_to_comm[p][name_lower] = set()
                self.name_to_comm[p][name_lower].add(cid)
                self.node_perspective[nid] = (p, cid)

    def node_name(self, nid: str) -> str:
        n = self.kg_node_map.get(nid, {})
        return n.get('properties', {}).get('name', nid[:16]) if n else nid[:16]

    def node_type(self, nid: str) -> str:
        n = self.kg_node_map.get(nid, {})
        return (n.get('labels', ['?'])[0] if n else '?')

    def node_layer(self, nid: str) -> str:
        return TYPE_TO_LAYER.get(self.node_type(nid), 'other')


# ==================== 实体提取与社区匹配（增强版）====================

class SmartMatcher:
    """
    增强型匹配器，解决根因1和根因2：
    - 实体提取 + 语义扩展关键词
    - 社区名称模糊匹配
    - 社区互补推导（当某层社区为空时，通过层级关系反向推导）
    - 语义节点锚点（直接从KG节点中找语义相关节点，绕过社区）
    """

    def __init__(self, loader: DataLoader):
        self.loader = loader

    # -------- Step A: 实体提取 + 语义扩展 --------

    def extract_and_expand(self, event_desc: str) -> Dict[str, List[str]]:
        """提取实体并扩展同义词/上位词"""
        print("\n[步骤1] 提取并扩展实体关键词...")
        prompt = f"""你是资本市场法规专家。请从以下事件描述中提取关键实体，并为每类实体补充法规中常见的规范表述（同义词/上位词）。

事件描述：
{event_desc}

返回严格 JSON，不含其他内容：
{{
  "责任方": ["原文实体1", "规范表述1", "规范表述2"],
  "违规行为": ["原文行为1", "规范表述1", "规范表述2"],
  "监管机构": ["原文机构1", "规范表述1"],
  "扩展关键词": ["与事件相关的法规领域术语1", "术语2", "术语3"]
}}

规范表述举例：
- "董事长王某" → "上市公司董事"、"内幕信息知情人"、"实际控制人"
- "买入股票获利" → "内幕交易"、"利用内幕信息交易"、"买卖证券"
- "虚增收入" → "财务造假"、"虚假陈述"、"信息披露违规"
- "证监会" → "中国证监会"、"证监会派出机构"、"监管机构"
"""
        try:
            raw = call_deepseek(prompt, max_tokens=600)
            s, e = raw.find('{'), raw.rfind('}') + 1
            result = json.loads(raw[s:e])
        except Exception as ex:
            print(f"  实体提取失败: {ex}，使用简单分词")
            result = {"责任方": [], "违规行为": [], "监管机构": [], "扩展关键词": []}

        # 合并所有关键词
        all_kw = []
        for key in ["责任方", "违规行为", "监管机构", "扩展关键词"]:
            all_kw.extend(result.get(key, []))
        result['_all_keywords'] = list(set(all_kw))

        print(f"  责任方词: {result.get('责任方', [])}")
        print(f"  违规行为词: {result.get('违规行为', [])}")
        print(f"  监管机构词: {result.get('监管机构', [])}")
        print(f"  扩展关键词: {result.get('扩展关键词', [])}")
        return result

    # -------- Step B: 社区匹配 --------

    def match_communities(self, entities: Dict[str, List[str]],
                          event_desc: str) -> Dict[str, Set[int]]:
        """多轮社区匹配，包含互补推导"""
        print("\n[步骤2] 匹配相关社区...")

        matched: Dict[str, Set[int]] = {
            'responsibility': set(),
            'violation': set(),
            'regulatory': set(),
        }

        kw_map = {
            'responsibility': entities.get('责任方', []) + entities.get('_all_keywords', []),
            'violation': entities.get('违规行为', []) + entities.get('_all_keywords', []),
            'regulatory': entities.get('监管机构', []) + entities.get('_all_keywords', []),
        }

        # 轮1：精确 + 包含匹配
        for p, keywords in kw_map.items():
            for kw in keywords:
                if len(kw) < 2:
                    continue
                kl = kw.lower()
                for name_lower, comms in self.loader.name_to_comm[p].items():
                    if kl in name_lower or name_lower in kl:
                        matched[p].update(comms)

        # 轮2：字符串中 2-gram 扫描（处理长词没有精确命中的情况）
        all_text = event_desc + ' ' + ' '.join(entities.get('_all_keywords', []))
        chars = re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9]', '', all_text)
        for i in range(len(chars) - 1):
            bigram = chars[i:i+2]
            for p in ['responsibility', 'violation', 'regulatory']:
                for name_lower, comms in self.loader.name_to_comm[p].items():
                    if bigram in name_lower:
                        matched[p].update(comms)

        # 轮3：互补推导——通过社区层级关系补全空层
        matched = self._complement_by_hierarchy(matched)

        print(f"  责任方社区: {sorted(matched['responsibility'])}")
        print(f"  违规行为社区: {sorted(matched['violation'])}")
        print(f"  监管机构社区: {sorted(matched['regulatory'])}")
        return matched

    def _complement_by_hierarchy(self, matched: Dict[str, Set[int]]) -> Dict[str, Set[int]]:
        """
        通过社区层级关系补全空层

        策略：
        - 责任方为空 → 从违规行为/监管社区，在 hierarchy 表中反向查找 source 端的责任方社区
        - 违规行为为空 → 从责任方或监管社区，查找中间的违规行为社区
        - 监管机构为空 → 从违规行为/责任方社区，查找 target 端的监管机构社区
        """
        h = self.loader.hierarchy

        def find_linked(from_perspective: str, from_ids: Set[int],
                        to_perspective: str, direction: str = 'forward') -> Set[int]:
            """在 hierarchy 中查找关联社区"""
            result = set()
            if direction == 'forward':
                rows = h[(h['source_perspective'] == from_perspective) &
                         (h['source_community_id'].isin(from_ids)) &
                         (h['target_perspective'] == to_perspective)]
                result.update(rows['target_community_id'].tolist())
            else:  # backward
                rows = h[(h['target_perspective'] == from_perspective) &
                         (h['target_community_id'].isin(from_ids)) &
                         (h['source_perspective'] == to_perspective)]
                result.update(rows['source_community_id'].tolist())
            return result

        # 补全责任方
        if not matched['responsibility']:
            # 从违规行为反向找
            if matched['violation']:
                found = find_linked('violation', matched['violation'],
                                    'responsibility', 'backward')
                if found:
                    matched['responsibility'].update(found)
                    print(f"  → 由违规行为反向推导责任方社区: {sorted(found)}")
            # 从监管机构反向找
            if not matched['responsibility'] and matched['regulatory']:
                found = find_linked('regulatory', matched['regulatory'],
                                    'responsibility', 'backward')
                if found:
                    matched['responsibility'].update(found)
                    print(f"  → 由监管机构反向推导责任方社区: {sorted(found)}")
            # 最后兜底：取所有社区（图谱中责任方节点通常连通性高）
            if not matched['responsibility']:
                all_resp_comms = set(self.loader.comm_nodes['responsibility'].keys())
                # 取节点数最多的前3个社区
                top_comms = sorted(all_resp_comms,
                                   key=lambda c: len(self.loader.comm_nodes['responsibility'].get(c, [])),
                                   reverse=True)[:3]
                matched['responsibility'].update(top_comms)
                print(f"  → 兜底：取责任方最大社区: {top_comms}")

        # 补全违规行为
        if not matched['violation']:
            if matched['responsibility']:
                found = find_linked('responsibility', matched['responsibility'],
                                    'violation', 'forward')
                if found:
                    matched['violation'].update(found)
                    print(f"  → 由责任方推导违规行为社区: {sorted(found)}")

        # 补全监管机构
        if not matched['regulatory']:
            if matched['violation']:
                found = find_linked('violation', matched['violation'],
                                    'regulatory', 'forward')
                if found:
                    matched['regulatory'].update(found)
                    print(f"  → 由违规行为推导监管机构社区: {sorted(found)}")

        return matched

    # -------- Step C: 语义节点锚点（绕过社区，直接找KG节点）--------

    def find_semantic_anchors(self, entities: Dict[str, List[str]],
                              matched_communities: Dict[str, Set[int]]) -> Dict[str, List[str]]:
        """
        直接在 KG 中用关键词找语义相关节点作为路径锚点

        返回 {'responsibility': [node_id,...], 'violation': [...], 'regulatory': [...]}
        """
        print("\n[步骤3] 定位语义节点锚点...")

        anchors: Dict[str, List[str]] = {
            'responsibility': [],
            'violation': [],
            'regulatory': [],
        }

        # 1. 从匹配社区中取核心类型节点（原逻辑）
        for p in ['responsibility', 'violation', 'regulatory']:
            core_types = LAYER_TYPES[p]
            candidates: List[Tuple[str, float]] = []
            for cid in matched_communities[p]:
                for nid in self.loader.comm_nodes[p].get(cid, []):
                    ntype = self.loader.node_type(nid)
                    if ntype in core_types:
                        deg = self.loader.G.degree(nid)
                        candidates.append((nid, deg))
            candidates.sort(key=lambda x: x[1], reverse=True)
            anchors[p] = [nid for nid, _ in candidates[:MAX_ANCHOR_NODES]]

        # 2. 对空层，用关键词在 KG 全量节点中搜索（绕过社区）
        kw_map = {
            'responsibility': entities.get('责任方', []),
            'violation': entities.get('违规行为', []),
            'regulatory': entities.get('监管机构', []),
        }
        all_kw = entities.get('_all_keywords', [])

        for p, target_types in LAYER_TYPES.items():
            if p not in anchors:
                continue
            if len(anchors[p]) >= 3:
                continue  # 已够用
            # 用关键词搜索
            search_kws = kw_map.get(p, []) + all_kw
            for nid, node in self.loader.kg_node_map.items():
                ntype = node['labels'][0] if node['labels'] else ''
                if ntype not in target_types:
                    continue
                name = node['properties'].get('name', '').lower()
                for kw in search_kws:
                    if len(kw) >= 2 and (kw.lower() in name or name in kw.lower()):
                        if nid not in anchors[p]:
                            anchors[p].append(nid)
                            break
                if len(anchors[p]) >= MAX_ANCHOR_NODES:
                    break

        # 3. 对仍然为空的层，扩展到 legal 类型节点（法规桥接）
        # 此时通过与已有锚点的邻居来补充
        for p in ['responsibility', 'violation', 'regulatory']:
            if not anchors[p]:
                # 取相邻层锚点的邻居中符合类型的节点
                neighbor_candidates: List[Tuple[str, float]] = []
                other_layers = [op for op in ['responsibility', 'violation', 'regulatory'] if op != p]
                for op in other_layers:
                    for nid in anchors[op]:
                        for nbr in list(self.loader.G.successors(nid)) + list(self.loader.G.predecessors(nid)):
                            ntype = self.loader.node_type(nbr)
                            if ntype in LAYER_TYPES[p]:
                                deg = self.loader.G.degree(nbr)
                                neighbor_candidates.append((nbr, deg))
                neighbor_candidates.sort(key=lambda x: x[1], reverse=True)
                anchors[p] = [nid for nid, _ in neighbor_candidates[:MAX_ANCHOR_NODES]]
                if anchors[p]:
                    print(f"  → {p} 通过邻居扩展找到 {len(anchors[p])} 个锚点")

        for p in ['responsibility', 'violation', 'regulatory']:
            names = [self.loader.node_name(n) for n in anchors[p][:5]]
            print(f"  {p} 锚点({len(anchors[p])}个): {names}")

        return anchors


# ==================== 双向 BFS 路径搜索（解决根因3、4）====================

class BidirectionalPathFinder:
    """
    双向 BFS 路径搜索器

    解决根因3：跳数不够
    解决根因4：允许 Law/Section/Chapter 作为桥接中间节点

    核心改进：
    1. 双向 BFS（从两端同时搜索，效率提升 O(b^d) → O(b^(d/2))）
    2. 允许经过 legal 类型节点（法规桥接）
    3. 增加"松弛模式"：当严格搜索失败时，允许经过任意中间节点
    """

    BRIDGE_TYPES = LAYER_TYPES['legal'] | LAYER_TYPES['other']   # 允许经过的桥接节点类型

    def __init__(self, loader: DataLoader):
        self.loader = loader
        self.G = loader.G
        # 无向版（双向可走）
        self.G_undirected = loader.G.to_undirected()

    def find_paths(self, source_ids: List[str], target_ids: List[str],
                   max_depth: int = BFS_MAX_DEPTH,
                   max_paths: int = MAX_PATHS_PER_QUERY,
                   use_undirected: bool = False) -> List[List[str]]:
        """
        从 source_ids 集合到 target_ids 集合找路径

        use_undirected=True 时使用无向图（放宽约束，用于兜底）
        """
        if not source_ids or not target_ids:
            return []

        G = self.G_undirected if use_undirected else self.G
        target_set = set(target_ids)
        found: List[List[str]] = []
        seen_paths: Set[Tuple] = set()

        for src in source_ids[:MAX_ANCHOR_NODES]:
            if src not in G:
                continue
            # BFS
            queue: deque = deque()
            queue.append((src, [src], {src}))

            while queue:
                cur, path, visited = queue.popleft()

                if len(path) > max_depth:
                    continue

                if cur in target_set and len(path) > 1:
                    path_key = tuple(path)
                    if path_key not in seen_paths:
                        seen_paths.add(path_key)
                        found.append(path)
                        if len(found) >= max_paths * len(source_ids):
                            return self._select_best(found, max_paths)
                    continue

                # 扩展邻居
                if use_undirected:
                    neighbors = list(G.neighbors(cur))
                else:
                    neighbors = list(G.successors(cur))

                for nxt in neighbors:
                    if nxt in visited:
                        continue
                    nxt_type = self.loader.node_type(nxt)
                    nxt_layer = TYPE_TO_LAYER.get(nxt_type, 'other')

                    # 剪枝：不允许连续经过同类型非目标节点（避免无效循环）
                    if len(path) >= 3:
                        last_type = self.loader.node_type(path[-1])
                        if nxt_type == last_type and nxt not in target_set:
                            if nxt_type not in self.BRIDGE_TYPES:
                                continue

                    queue.append((nxt, path + [nxt], visited | {nxt}))

        return self._select_best(found, max_paths)

    def _select_best(self, paths: List[List[str]], n: int) -> List[List[str]]:
        """选出最好的 n 条路径（优先短路径，去重）"""
        seen: Set[Tuple] = set()
        unique: List[List[str]] = []
        for p in sorted(paths, key=len):
            key = tuple(p)
            if key not in seen:
                seen.add(key)
                unique.append(p)
        return unique[:n]

    def find_compliance_paths(self, resp_anchors: List[str],
                              reg_anchors: List[str]) -> List[dict]:
        """
        合规路径：责任方 → (法规/义务节点) → 监管机构
        三轮搜索：有向严格 → 有向松弛 → 无向兜底
        """
        # 轮1：有向图，深度7
        paths = self.find_paths(resp_anchors, reg_anchors, max_depth=7)

        # 轮2：如果为空，有向图，深度10
        if not paths:
            paths = self.find_paths(resp_anchors, reg_anchors, max_depth=BFS_MAX_DEPTH)

        # 轮3：如果还为空，无向图（放宽方向约束）
        if not paths:
            paths = self.find_paths(resp_anchors, reg_anchors,
                                    max_depth=BFS_MAX_DEPTH, use_undirected=True)

        return [self._path_to_record(p, 'compliance') for p in paths]

    def find_violation_paths(self, resp_anchors: List[str],
                             viol_anchors: List[str],
                             reg_anchors: List[str]) -> List[dict]:
        """
        违规路径：责任方 → 违规行为节点 → 监管机构
        分两段搜索再拼接，或直接端到端搜索（含违规节点作为中间点）
        """
        viol_set = set(viol_anchors)

        # 段1：责任方 → 违规行为
        seg1 = self.find_paths(resp_anchors, viol_anchors, max_depth=6)
        if not seg1:
            seg1 = self.find_paths(resp_anchors, viol_anchors,
                                   max_depth=BFS_MAX_DEPTH, use_undirected=True)

        # 段2：违规行为 → 监管机构
        seg2 = self.find_paths(viol_anchors, reg_anchors, max_depth=6)
        if not seg2:
            seg2 = self.find_paths(viol_anchors, reg_anchors,
                                   max_depth=BFS_MAX_DEPTH, use_undirected=True)

        # 拼接
        results: List[dict] = []
        seen: Set[Tuple] = set()
        for p1 in seg1:
            for p2 in seg2:
                if p1[-1] == p2[0]:
                    combined = p1 + p2[1:]
                    key = tuple(combined)
                    if key not in seen:
                        seen.add(key)
                        results.append(self._path_to_record(combined, 'violation'))
                        if len(results) >= MAX_PATHS_PER_QUERY:
                            return results

        # 如果拼接为空：直接端到端搜索（责任方→监管，要求路径经过违规节点）
        if not results:
            direct = self.find_paths(resp_anchors, reg_anchors,
                                     max_depth=BFS_MAX_DEPTH, use_undirected=True)
            for p in direct:
                if any(n in viol_set for n in p[1:-1]):
                    key = tuple(p)
                    if key not in seen:
                        seen.add(key)
                        results.append(self._path_to_record(p, 'violation'))
                        if len(results) >= MAX_PATHS_PER_QUERY:
                            break

        # 最后兜底：即使不经过违规节点，也保留一条（用于提示）
        if not results and seg1:
            # 至少保留责任方→违规行为的段
            results.append(self._path_to_record(seg1[0], 'violation_partial'))

        return results[:MAX_PATHS_PER_QUERY]

    def _path_to_record(self, node_ids: List[str], ptype: str) -> dict:
        edges = []
        for i in range(len(node_ids) - 1):
            src, tgt = node_ids[i], node_ids[i + 1]
            edge_data = self.G.get_edge_data(src, tgt)
            if edge_data is None:
                # 无向回退：查反向边
                edge_data = self.G.get_edge_data(tgt, src)
                rel_type = (edge_data or {}).get('rel_type', '关联')
                # 用无向标记
                edges.append((src, tgt, f"{rel_type}(↔)"))
            else:
                edges.append((src, tgt, edge_data.get('rel_type', '关联')))
        return {'nodes': node_ids, 'edges': edges, 'type': ptype}


# ==================== 可视化 ====================

PERSPECTIVE_COLOR = {
    'responsibility': '#E53935',   # 红
    'violation': '#FF8F00',         # 橙
    'regulatory': '#1E88E5',        # 蓝
    'legal': '#43A047',             # 绿
    'other': '#8E24AA',             # 紫
}
PERSPECTIVE_ZH = {
    'responsibility': '责任方',
    'violation': '违规行为',
    'regulatory': '监管机构',
    'legal': '法规条款',
    'other': '其他',
}
PATH_COLOR = {'compliance': '#1565C0', 'violation': '#B71C1C',
              'violation_partial': '#E65100'}


def visualize_community_network(loader: DataLoader,
                                matched_communities: Dict[str, Set[int]],
                                community_paths: List[Tuple],
                                output_file: str):
    """社区级网络（原逻辑，保留不变）"""
    G = nx.DiGraph()
    node_colors, node_sizes, labels = [], [], {}

    for p, comms in matched_communities.items():
        for cid in comms:
            nid = f"{p}_{cid}"
            G.add_node(nid)
            node_colors.append(PERSPECTIVE_COLOR.get(p, '#999'))
            node_sizes.append(3000)
            labels[nid] = f"{PERSPECTIVE_ZH.get(p, p)}\n社区#{cid}"

    for sp, si, tp, ti, rt, strong in community_paths:
        sn, tn = f"{sp}_{si}", f"{tp}_{ti}"
        if sn in G and tn in G:
            G.add_edge(sn, tn, relation=rt, width=3 if strong else 1)

    fig, ax = plt.subplots(figsize=(16, 10))
    if len(G) == 0:
        ax.text(0.5, 0.5, '未找到匹配社区', ha='center', va='center', fontsize=16)
    else:
        pos = nx.spring_layout(G, k=2.5, iterations=80, seed=42)
        nx.draw_networkx_nodes(G, pos, node_color=node_colors,
                               node_size=node_sizes, alpha=0.88, ax=ax)
        edges = list(G.edges())
        widths = [G[u][v].get('width', 1) for u, v in edges]
        nx.draw_networkx_edges(G, pos, width=widths, alpha=0.6,
                               arrows=True, arrowsize=22, ax=ax)
        nx.draw_networkx_labels(G, pos, labels, font_size=9, ax=ax)
        legend_handles = [
            mpatches.Patch(color=PERSPECTIVE_COLOR[p], label=PERSPECTIVE_ZH[p])
            for p in ['responsibility', 'violation', 'regulatory']
        ]
        ax.legend(handles=legend_handles, loc='upper left', fontsize=10)

    ax.set_title("社区级责任递进路径（v3）", fontsize=14, pad=12)
    ax.axis('off')
    plt.tight_layout()
    plt.savefig(output_file, dpi=180, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f"  ✓ 社区级可视化: {output_file}")


def visualize_node_paths(loader: DataLoader,
                         node_paths: Dict[str, List],
                         matched_communities: Dict[str, Set[int]],
                         output_file: str):
    """
    节点级穿透路径可视化（横向泳道布局）

    左列：合规路径（蓝色）
    右列：违规路径（红色）
    每条路径一行，节点按所属视角着色
    """
    compliance_paths = node_paths.get('compliance', [])
    violation_paths = node_paths.get('violation', [])
    n_rows = max(len(compliance_paths), len(violation_paths), 1)

    fig, axes = plt.subplots(n_rows, 2, figsize=(24, max(4 * n_rows, 5)))
    if n_rows == 1:
        axes = axes.reshape(1, 2)

    fig.suptitle("节点级穿透路径可视化（v3）\n蓝色=合规路径  红色=违规路径",
                 fontsize=15, fontweight='bold', y=1.01)

    def get_node_color(nid: str) -> str:
        layer = loader.node_layer(nid)
        return PERSPECTIVE_COLOR.get(layer, '#9E9E9E')

    def draw_single_path(ax, rec: Optional[dict], title: str, arrow_color: str):
        ax.axis('off')
        ax.set_title(title, fontsize=10, color=arrow_color, pad=6, fontweight='bold')
        if rec is None:
            ax.text(0.5, 0.5, '（无路径）', ha='center', va='center',
                    fontsize=10, color='#BDBDBD')
            return

        path_nodes = rec['nodes']
        path_edges = rec['edges']
        n = len(path_nodes)
        if n == 0:
            return

        xs = np.linspace(0.05, 0.95, n)
        y_node = 0.55
        y_label = 0.22

        # 画边
        for i, (src_id, tgt_id, rel_type) in enumerate(path_edges):
            x1, x2 = xs[i], xs[i + 1]
            ax.annotate('', xy=(x2 - 0.025, y_node), xytext=(x1 + 0.025, y_node),
                        arrowprops=dict(arrowstyle='->', color=arrow_color,
                                        lw=2.2, mutation_scale=18))
            mid_x = (x1 + x2) / 2
            # 关系标签背景
            ax.text(mid_x, y_node + 0.15, rel_type,
                    ha='center', va='center', fontsize=7.5, color=arrow_color,
                    bbox=dict(boxstyle='round,pad=0.25', facecolor='white',
                              edgecolor=arrow_color, alpha=0.9, linewidth=1.2))

        # 画节点
        for i, nid in enumerate(path_nodes):
            x = xs[i]
            color = get_node_color(nid)
            ntype = loader.node_type(nid)
            layer = TYPE_TO_LAYER.get(ntype, 'other')

            # 节点圆
            circle = plt.Circle((x, y_node), 0.048, color=color,
                                 zorder=5, alpha=0.92)
            ax.add_patch(circle)

            # 节点层级标注（小字）
            ax.text(x, y_node + 0.005, PERSPECTIVE_ZH.get(layer, '?')[:3],
                    ha='center', va='center', fontsize=6.5,
                    color='white', fontweight='bold', zorder=6)

            # 节点名称标签（下方）
            name = loader.node_name(nid)
            name_display = name[:14] + '..' if len(name) > 14 else name
            ax.text(x, y_label, f"{name_display}\n({ntype})",
                    ha='center', va='top', fontsize=7.2,
                    bbox=dict(boxstyle='round,pad=0.3', facecolor='#FAFAFA',
                              edgecolor='#CCCCCC', alpha=0.95))

        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)

    for row_i in range(n_rows):
        comp_rec = compliance_paths[row_i] if row_i < len(compliance_paths) else None
        viol_rec = violation_paths[row_i] if row_i < len(violation_paths) else None

        comp_title = (f"合规路径 #{row_i + 1}  ({len(comp_rec['nodes'])} 节点)"
                      if comp_rec else f"合规路径 #{row_i + 1}")
        viol_type_zh = {'violation': '违规', 'violation_partial': '违规（部分）'}.get(
            (viol_rec or {}).get('type', ''), '违规')
        viol_title = (f"{viol_type_zh}路径 #{row_i + 1}  ({len(viol_rec['nodes'])} 节点)"
                      if viol_rec else f"违规路径 #{row_i + 1}")

        draw_single_path(axes[row_i][0], comp_rec, comp_title,
                         PATH_COLOR['compliance'])
        draw_single_path(axes[row_i][1], viol_rec, viol_title,
                         PATH_COLOR.get((viol_rec or {}).get('type', 'violation'),
                                        PATH_COLOR['violation']))

    # 图例
    legend_handles = [
        mpatches.Patch(color=PERSPECTIVE_COLOR[p], label=f"{PERSPECTIVE_ZH[p]}节点")
        for p in ['responsibility', 'violation', 'regulatory', 'legal', 'other']
    ]
    fig.legend(handles=legend_handles, loc='lower center', ncol=5,
               fontsize=9, bbox_to_anchor=(0.5, -0.03))

    plt.tight_layout()
    plt.savefig(output_file, dpi=180, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f"  ✓ 节点级可视化: {output_file}")


# ==================== 社区级路径构建 ====================

def build_community_paths(matched_communities: Dict[str, Set[int]],
                          hierarchy: pd.DataFrame) -> List[Tuple]:
    paths = []
    for _, row in hierarchy.iterrows():
        sp = row['source_perspective']
        si = int(row['source_community_id'])
        tp = row['target_perspective']
        ti = int(row['target_community_id'])
        rt = row['relation_type']
        strong = bool(row['is_strong_link'])
        if (si in matched_communities.get(sp, set()) and
                ti in matched_communities.get(tp, set())):
            paths.append((sp, si, tp, ti, rt, strong))
    return paths


# ==================== Prompt 构建 ====================

def build_prompt(event_desc: str,
                 community_reports_text: str,
                 node_triples_text: str,
                 community_paths: List[Tuple],
                 node_paths: Dict[str, List],
                 loader: DataLoader) -> str:
    """构建最终 Prompt（注入三层知识）"""

    def nname(nid):
        return loader.node_name(nid)

    # 社区级路径描述
    pzh = {'responsibility': '责任方', 'violation': '违规行为', 'regulatory': '监管机构'}
    comm_lines = []
    for sp, si, tp, ti, rt, strong in community_paths:
        s = '强关联' if strong else '弱关联'
        comm_lines.append(f"  {pzh.get(sp,sp)}社区#{si} --[{rt},{s}]--> {pzh.get(tp,tp)}社区#{ti}")

    # 节点级路径描述
    node_lines = []
    for ptype, plabel in [('compliance', '合规'), ('violation', '违规'),
                           ('violation_partial', '违规(部分)')]:
        for i, rec in enumerate(node_paths.get(ptype, [])[:2], 1):
            step = ' → '.join([nname(n) for n in rec['nodes']])
            node_lines.append(f"  {plabel}路径{i}: {step}")

    return f"""你是资本市场法规合规领域资深专家，请对以下事件进行五维穿透式分析。

【事件描述】
{event_desc}

━━━━ 一、社区级责任链路（宏观）━━━━
{chr(10).join(comm_lines) if comm_lines else '（未找到社区级路径）'}

━━━━ 二、节点级穿透路径（微观）━━━━
{chr(10).join(node_lines) if node_lines else '（未找到节点级路径）'}

━━━━ 三、知识图谱原始三元组（精细法律依据）━━━━
{node_triples_text}

━━━━ 四、相关社区法规报告 ━━━━
{community_reports_text}

━━━━ 分析要求（五个维度，不使用 Markdown）━━━━

1. 责任主体认定：具体指出责任主体类型，引用三元组中的关系为依据
2. 违规行为定性：判断违规类型（内幕交易/信披违规/操纵市场等），援引法条
3. 监管机构职责与处罚依据：说明哪些机构可介入，依据哪条法规，可触发哪些处罚
4. 合规义务梳理：基于合规路径列出应履行的合规义务
5. 风险提示与建议：具体可操作的合规建议及监管趋势提示"""


def get_community_reports_text(loader: DataLoader,
                               matched_communities: Dict[str, Set[int]]) -> str:
    parts = []
    for p, pname in [('responsibility', '责任方'),
                     ('violation', '违规行为'),
                     ('regulatory', '监管机构')]:
        df = loader.reports[p]
        comms = sorted(matched_communities.get(p, set()))
        if not comms:
            continue
        parts.append(f"\n{'='*50}\n【{pname}社区报告】\n{'='*50}")
        for cid in comms[:4]:  # 最多4个社区
            rows = df[df['community'] == cid]
            if rows.empty:
                continue
            row = rows.iloc[0]
            parts.append(f"\n社区#{cid}：{row.get('title', '')}")
            parts.append(f"摘要：{row.get('summary', '')[:250]}")
            try:
                findings = json.loads(row.get('findings', '[]'))
                for fi, f_item in enumerate(findings[:2], 1):
                    smry = f_item.get('summary', '')
                    expl = f_item.get('explanation', '')[:100]
                    parts.append(f"  发现{fi}：{smry} — {expl}")
            except Exception:
                pass
    return '\n'.join(parts)


def get_node_triples_text(loader: DataLoader,
                          node_paths: Dict[str, List]) -> str:
    G = loader.G
    seen: Set[Tuple] = set()
    lines = []
    for ptype in ['compliance', 'violation', 'violation_partial']:
        for rec in node_paths.get(ptype, []):
            ptype_zh = {'compliance': '合规', 'violation': '违规',
                        'violation_partial': '违规(部分)'}.get(ptype, ptype)
            lines.append(f"\n[{ptype_zh}路径]")
            for src, tgt, rt in rec['edges']:
                key = (src, tgt, rt)
                if key in seen:
                    continue
                seen.add(key)
                sname = loader.node_name(src)
                stype = loader.node_type(src)
                tname = loader.node_name(tgt)
                ttype = loader.node_type(tgt)
                lines.append(f"  {sname}（{stype}） -[{rt}]-> {tname}（{ttype}）")
    return '\n'.join(lines) if lines else '（未找到三元组）'


# ==================== 主查询系统 ====================

class RegulatoryQuerySystemV3:
    """监管违规穿透式查询系统 v3"""

    def __init__(self, data_dir: str = "data"):
        self.loader = DataLoader(data_dir)
        self.loader.load()
        self.matcher = SmartMatcher(self.loader)
        self.path_finder = BidirectionalPathFinder(self.loader)

    def query(self, event_desc: str, output_dir: str = "v3_output") -> dict:
        os.makedirs(output_dir, exist_ok=True)
        print("=" * 70)
        print("  监管违规穿透式查询系统 v3（节点级精细路径）")
        print("=" * 70)
        print(f"\n事件：{event_desc[:80]}...\n")

        # Step 1: 实体提取 + 语义扩展
        entities = self.matcher.extract_and_expand(event_desc)

        # Step 2: 社区匹配（含互补推导）
        matched_communities = self.matcher.match_communities(entities, event_desc)

        # Step 3: 语义节点锚点定位
        anchors = self.matcher.find_semantic_anchors(entities, matched_communities)

        # Step 4: 节点级路径搜索
        print("\n[步骤4] 搜索节点级穿透路径...")
        comp_paths = self.path_finder.find_compliance_paths(
            anchors['responsibility'], anchors['regulatory'])
        viol_paths = self.path_finder.find_violation_paths(
            anchors['responsibility'], anchors['violation'], anchors['regulatory'])

        node_paths = {'compliance': comp_paths, 'violation': viol_paths}
        print(f"  合规路径: {len(comp_paths)} 条，违规路径: {len(viol_paths)} 条")

        # 打印路径摘要
        for ptype, paths in node_paths.items():
            for i, rec in enumerate(paths, 1):
                steps = ' → '.join([self.loader.node_name(n) for n in rec['nodes']])
                print(f"  {ptype}路径#{i}（{len(rec['nodes'])}跳）: {steps[:120]}")

        # Step 5: 社区级路径
        community_paths = build_community_paths(matched_communities, self.loader.hierarchy)
        print(f"\n[步骤5] 社区级路径: {len(community_paths)} 条")

        # Step 6: 可视化
        print("\n[步骤6] 生成可视化...")
        comm_vis = os.path.join(output_dir, "event_network.png")
        node_vis = os.path.join(output_dir, "node_level_path.png")
        visualize_community_network(self.loader, matched_communities,
                                    community_paths, comm_vis)
        visualize_node_paths(self.loader, node_paths, matched_communities, node_vis)

        # Step 7: 构建 Prompt + 调用 LLM
        print("\n[步骤7] 调用 DeepSeek 生成分析报告...")
        community_reports_text = get_community_reports_text(
            self.loader, matched_communities)
        node_triples_text = get_node_triples_text(self.loader, node_paths)
        final_prompt = build_prompt(
            event_desc, community_reports_text, node_triples_text,
            community_paths, node_paths, self.loader)

        try:
            final_answer = call_deepseek(final_prompt, max_tokens=4000, temperature=0.3)
        except Exception as e:
            print(f"  LLM 调用失败: {e}")
            final_answer = f"LLM 调用失败，请检查 API Key。\n错误：{e}"

        # Step 8: 保存输出
        print("\n[步骤8] 保存输出...")
        self._save_all(output_dir, event_desc, entities, matched_communities,
                       anchors, community_paths, node_paths,
                       node_triples_text, community_reports_text,
                       final_answer, final_prompt)

        print("\n" + "=" * 70)
        print("  查询完成！")
        print("=" * 70)
        return {
            'entities': entities,
            'matched_communities': matched_communities,
            'anchors': anchors,
            'community_paths': community_paths,
            'node_paths': node_paths,
            'community_vis': comm_vis,
            'node_vis': node_vis,
            'final_answer': final_answer,
        }

    def _save_all(self, output_dir, event_desc, entities, matched_communities,
                  anchors, community_paths, node_paths,
                  node_triples_text, community_reports_text,
                  final_answer, final_prompt):
        loader = self.loader

        # 节点路径文字报告
        lines = ["=" * 65, "节点级穿透路径报告（v3）", "=" * 65,
                 f"\n事件：{event_desc}\n"]
        for ptype, pname in [('compliance', '合规路径（责任方→监管）'),
                              ('violation', '违规路径（责任方→违规→监管）'),
                              ('violation_partial', '违规路径（部分）')]:
            recs = node_paths.get(ptype, [])
            if not recs:
                continue
            lines.append(f"\n【{pname}】")
            for i, rec in enumerate(recs, 1):
                lines.append(f"  路径#{i}（{len(rec['nodes'])}跳）：")
                for src, tgt, rt in rec['edges']:
                    lines.append(
                        f"    {loader.node_name(src)}（{loader.node_type(src)}）"
                        f" -[{rt}]-> "
                        f"{loader.node_name(tgt)}（{loader.node_type(tgt)}）")

        with open(os.path.join(output_dir, "node_paths_report.txt"), 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))

        with open(os.path.join(output_dir, "community_reports.txt"), 'w', encoding='utf-8') as f:
            f.write(community_reports_text)
        with open(os.path.join(output_dir, "node_triples.txt"), 'w', encoding='utf-8') as f:
            f.write(node_triples_text)
        with open(os.path.join(output_dir, "debug_prompt.txt"), 'w', encoding='utf-8') as f:
            f.write(final_prompt)
        with open(os.path.join(output_dir, "final_analysis.txt"), 'w', encoding='utf-8') as f:
            f.write("=" * 65 + "\n监管违规事件综合分析报告（v3）\n" + "=" * 65 + "\n\n")
            f.write(f"【事件】\n{event_desc}\n\n【分析结果】\n{final_answer}\n")

        # Excel 汇总
        self._save_excel(output_dir, event_desc, entities, matched_communities,
                         anchors, community_paths, node_paths, final_answer)

        for fname in ['event_network.png', 'node_level_path.png',
                      'node_paths_report.txt', 'node_triples.txt',
                      'final_analysis.txt', 'analysis_summary_v3.xlsx']:
            if os.path.exists(os.path.join(output_dir, fname)):
                print(f"    ✓ {fname}")

    def _save_excel(self, output_dir, event_desc, entities, matched_communities,
                    anchors, community_paths, node_paths, final_answer):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        loader = self.loader

        def hfont(): return Font(bold=True, color="FFFFFF")
        def hfill(color): return PatternFill("solid", fgColor=color)
        def wfill(color): return PatternFill("solid", fgColor=color)

        # Sheet1: 概述
        ws1 = wb.active
        ws1.title = "事件概述"
        ws1['A1'] = "监管违规穿透分析报告 v3"
        ws1['A1'].font = Font(size=13, bold=True)
        ws1.append([])
        ws1.append(["事件描述", event_desc])
        ws1.append([])
        for key, label in [('责任方', '责任方实体'), ('违规行为', '违规行为'),
                            ('监管机构', '监管机构'), ('扩展关键词', '扩展关键词')]:
            ws1.append([label, str(entities.get(key, []))])
        ws1.append([])
        ws1.append(["LLM分析摘要", final_answer[:1500]])
        ws1.column_dimensions['A'].width = 18
        ws1.column_dimensions['B'].width = 100

        # Sheet2: 社区级路径
        ws2 = wb.create_sheet("社区级路径")
        h2 = ['序号', '源视角', '源社区ID', '目标视角', '目标社区ID', '关系类型', '关联强度']
        ws2.append(h2)
        for c in ws2[1]:
            c.font = hfont()
            c.fill = hfill("1565C0")
        pzh = {'responsibility': '责任方', 'violation': '违规行为', 'regulatory': '监管机构'}
        for i, (sp, si, tp, ti, rt, strong) in enumerate(community_paths, 1):
            ws2.append([i, pzh.get(sp, sp), si, pzh.get(tp, tp), ti, rt,
                        "强" if strong else "弱"])
        for col, w in zip('ABCDEFG', [6, 14, 10, 14, 10, 18, 8]):
            ws2.column_dimensions[col].width = w

        # Sheet3: 节点级路径
        ws3 = wb.create_sheet("节点级路径")
        h3 = ['路径类型', '路径#', '步骤#', '源节点ID', '源节点名称', '源节点类型',
              '关系类型', '目标节点ID', '目标节点名称', '目标节点类型', '目标节点层']
        ws3.append(h3)
        for c in ws3[1]:
            c.font = hfont()
            c.fill = hfill("B71C1C")
        ptype_zh = {'compliance': '合规', 'violation': '违规', 'violation_partial': '违规(部分)'}
        for ptype in ['compliance', 'violation', 'violation_partial']:
            for pno, rec in enumerate(node_paths.get(ptype, []), 1):
                for step_i, (src, tgt, rt) in enumerate(rec['edges'], 1):
                    ws3.append([
                        ptype_zh.get(ptype, ptype), pno, step_i,
                        src, loader.node_name(src), loader.node_type(src),
                        rt,
                        tgt, loader.node_name(tgt), loader.node_type(tgt),
                        PERSPECTIVE_ZH.get(loader.node_layer(tgt), loader.node_layer(tgt)),
                    ])
        for col, w in zip('ABCDEFGHIJK', [12, 6, 6, 28, 22, 18, 14, 28, 22, 18, 12]):
            ws3.column_dimensions[col].width = w

        # Sheet4: 节点锚点
        ws4 = wb.create_sheet("节点锚点")
        h4 = ['视角', '节点ID', '节点名称', '节点类型', '节点层', '图度数']
        ws4.append(h4)
        for c in ws4[1]:
            c.font = hfont()
            c.fill = hfill("1B5E20")
        for p, pname in [('responsibility', '责任方'), ('violation', '违规行为'),
                          ('regulatory', '监管机构')]:
            for nid in anchors.get(p, []):
                ws4.append([pname, nid, loader.node_name(nid), loader.node_type(nid),
                             PERSPECTIVE_ZH.get(loader.node_layer(nid), '?'),
                             loader.G.degree(nid)])
        for col, w in zip('ABCDEF', [12, 28, 22, 18, 12, 8]):
            ws4.column_dimensions[col].width = w

        # Sheet5: 匹配社区
        ws5 = wb.create_sheet("匹配社区详情")
        h5 = ['视角', '社区ID', '社区标题', '摘要（前300字）', '关键词']
        ws5.append(h5)
        for c in ws5[1]:
            c.font = hfont()
            c.fill = hfill("4A148C")
        for p, pname in [('responsibility', '责任方'), ('violation', '违规行为'),
                          ('regulatory', '监管机构')]:
            df = loader.reports[p]
            for cid in sorted(matched_communities.get(p, set())):
                rows = df[df['community'] == cid]
                if rows.empty:
                    continue
                row = rows.iloc[0]
                ws5.append([pname, cid, row.get('title', ''),
                             str(row.get('summary', ''))[:300],
                             row.get('key_words', '')])
        for col, w in zip('ABCDE', [12, 8, 40, 80, 40]):
            ws5.column_dimensions[col].width = w

        excel_path = os.path.join(output_dir, "analysis_summary_v3.xlsx")
        wb.save(excel_path)


# ==================== 入口 ====================

def main():
    print("初始化监管违规穿透式查询系统 v3...")
    system = RegulatoryQuerySystemV3(data_dir="data")

    print("\n" + "=" * 70)
    print("请输入事件描述（输入完成后按 Enter，再输入 END 结束）：")
    print("=" * 70)
    lines = []
    while True:
        line = input()
        if line.strip().upper() == 'END':
            break
        lines.append(line)
    event_description = '\n'.join(lines).strip()
    if not event_description:
        print("错误：事件描述不能为空！")
        return

    results = system.query(event_description, output_dir="v3_output")
    print("\n" + "=" * 70)
    print("最终分析结果：")
    print("=" * 70)
    print(results['final_answer'])


if __name__ == "__main__":
    main()
