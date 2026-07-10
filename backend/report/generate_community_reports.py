#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
社区报告生成器
================================================================
功能：
1. 读取weighted_leiden_clustering.py生成的社区划分结果
2. 结合政策法规原文，为每种视角下的每个社区调用DeepSeek API生成社区报告
3. 输出包含完整字段的Excel文件，为RAG检索做准备

字段：id, community, level, title, parent, children, summary,
      key_words, findings, rank, rating_explanation, full_content_json
================================================================
"""

import os
import sys
import json
import uuid
import time
import hashlib
import pandas as pd
import traceback
from typing import Dict, List, Tuple, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests

# ==================== 配置 ====================

API_BASE = "https://api.deepseek.com/v1"
# API_KEY = os.environ.get("GRAPHRAG_API_KEY", "")
API_KEY="sk-0a57f72b50854ace9d134a5eb697c4dc"
MODEL_NAME = "deepseek-chat"  # DeepSeek模型名称

# 输入目录（leiden聚类结果输出目录）
CLUSTERING_RESULT_DIR = "weighted_leiden_results"
# 政策法规原文目录
POLICY_DOC_DIR = "policy_docs"
# 图谱数据文件
GRAPH_DATA_FILE = "data/merged_regulatory_unified.txt"
# 输出目录
OUTPUT_DIR = "community_reports"

# 视角配置
PERSPECTIVES = {
    "responsibility": "责任方",
    "regulatory": "监管机构",
    "violation": "违规行为",
}

# API调用配置
MAX_RETRIES = 3
RETRY_DELAY = 5  # 秒
REQUEST_INTERVAL = 1.5  # 请求间隔（秒），避免触发频率限制
MAX_TOKENS = 4096
TEMPERATURE = 1.0

# ==================== 数据加载模块 ====================


def load_graph_data(filepath: str) -> Tuple[Dict, Dict]:
    """
    加载知识图谱原始数据，返回节点字典和关系列表

    Returns:
        nodes_dict: {node_id: node_data}
        relationships_list: [rel_data, ...]
    """
    nodes_dict = {}
    relationships_list = []

    with open(filepath, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                data = json.loads(line)
                if data.get("type") == "node":
                    nodes_dict[data["id"]] = data
                elif data.get("type") == "relationship":
                    relationships_list.append(data)
            except json.JSONDecodeError:
                continue

    print(f"  ✓ 图谱数据: {len(nodes_dict)} 节点, {len(relationships_list)} 关系")
    return nodes_dict, relationships_list


def load_policy_documents(policy_dir: str) -> str:
    """
    加载政策法规原文目录下的所有txt文件，合并返回

    Returns:
        合并后的政策法规全文
    """
    all_text = []
    if not os.path.exists(policy_dir):
        print(f"  ⚠ 政策法规目录不存在: {policy_dir}，将仅使用图谱数据生成报告")
        return ""

    txt_files = sorted([f for f in os.listdir(policy_dir) if f.endswith(".txt")])
    for fname in txt_files:
        fpath = os.path.join(policy_dir, fname)
        with open(fpath, "r", encoding="utf-8") as f:
            content = f.read().strip()
        if content:
            all_text.append(f"=== {fname} ===\n{content}")
            print(f"  ✓ 已加载政策原文: {fname} ({len(content)} 字)")

    if not all_text:
        print(f"  ⚠ 政策法规目录为空: {policy_dir}")
        return ""

    return "\n\n".join(all_text)


def load_clustering_results(
    result_dir: str, perspective_key: str
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, Optional[dict]]:
    """
    加载某个视角的Leiden聚类结果

    Returns:
        df_nodes: 节点聚类结果
        df_rels: 关系聚类结果
        df_summary: 社区统计摘要
        vis_data: 可视化JSON数据 (可选)
    """
    prefix = perspective_key

    nodes_path = os.path.join(result_dir, f"{prefix}_1_节点聚类结果.csv")
    rels_path = os.path.join(result_dir, f"{prefix}_2_关系聚类结果.csv")
    summary_path = os.path.join(result_dir, f"{prefix}_3_社区统计摘要.csv")
    vis_path = os.path.join(result_dir, f"{prefix}_visualization_data.json")

    df_nodes = pd.read_csv(nodes_path, encoding="utf-8-sig")
    df_rels = pd.read_csv(rels_path, encoding="utf-8-sig")
    df_summary = pd.read_csv(summary_path, encoding="utf-8-sig")

    vis_data = None
    if os.path.exists(vis_path):
        with open(vis_path, "r", encoding="utf-8") as f:
            vis_data = json.load(f)

    # 过滤掉全局统计行
    df_summary = df_summary[df_summary["社区ID"] != "全局统计"].copy()
    df_summary["社区ID"] = df_summary["社区ID"].astype(int)

    print(
        f"  ✓ {PERSPECTIVES[perspective_key]}聚类结果: "
        f"{len(df_nodes)} 节点, {len(df_rels)} 关系, "
        f"{len(df_summary)} 个社区"
    )
    return df_nodes, df_rels, df_summary, vis_data


# ==================== 社区数据提取模块 ====================


def extract_community_data(
    community_id: int,
    df_nodes: pd.DataFrame,
    df_rels: pd.DataFrame,
    df_summary: pd.DataFrame,
    nodes_dict: Dict,
) -> Dict:
    """
    提取单个社区的详细数据，用于构建LLM prompt

    Returns:
        包含社区节点、关系、统计信息的字典
    """
    # 社区统计信息
    summary_row = df_summary[df_summary["社区ID"] == community_id]
    summary_info = summary_row.iloc[0].to_dict() if len(summary_row) > 0 else {}

    # 社区内节点
    community_nodes = df_nodes[df_nodes["所属社区"] == community_id].copy()

    # 社区内部关系（起止节点都在社区内）
    community_node_ids = set(community_nodes["节点ID"].tolist())
    internal_rels = df_rels[
        (df_rels["起始节点ID"].isin(community_node_ids))
        & (df_rels["终止节点ID"].isin(community_node_ids))
    ].copy()

    # 跨社区关系
    cross_rels = df_rels[
        (
            (df_rels["起始节点ID"].isin(community_node_ids))
            | (df_rels["终止节点ID"].isin(community_node_ids))
        )
        & (df_rels["是否跨社区"] == "是")
    ].copy()

    # 按类型分组节点
    nodes_by_type = {}
    for _, row in community_nodes.iterrows():
        ntype = row["节点类型"]
        if ntype not in nodes_by_type:
            nodes_by_type[ntype] = []
        node_info = {
            "id": row["节点ID"],
            "name": row["节点名称"],
            "type": ntype,
            "is_core": row["是否核心类型"] == "是",
            "degree": int(row["度数"]),
            "weighted_degree": float(row["加权度数"]),
            "in_degree": int(row["入度"]),
            "out_degree": int(row["出度"]),
        }
        nodes_by_type[ntype].append(node_info)

    # 关系列表
    relations = []
    for _, row in internal_rels.iterrows():
        relations.append(
            {
                "id": row["关系ID"],
                "type": row["关系类型"],
                "weight": float(row["计算权重"]),
                "source": row["起始节点名称"],
                "source_id": row["起始节点ID"],
                "source_type": row["起始节点类型"],
                "target": row["终止节点名称"],
                "target_id": row["终止节点ID"],
                "target_type": row["终止节点类型"],
            }
        )

    # 按权重排序取top关系
    relations.sort(key=lambda x: x["weight"], reverse=True)

    return {
        "community_id": community_id,
        "summary_stats": summary_info,
        "nodes_by_type": nodes_by_type,
        "total_nodes": len(community_nodes),
        "total_internal_rels": len(internal_rels),
        "total_cross_rels": len(cross_rels),
        "top_relations": relations[:80],  # 限制发给LLM的关系数量
        "all_node_ids": list(community_node_ids),
    }


def find_relevant_policy_text(
    community_data: Dict, nodes_dict: Dict, policy_text: str
) -> str:
    """
    根据社区中包含的Section/Chapter/Law节点，从政策原文中提取相关段落
    """
    if not policy_text:
        return ""

    # 收集社区相关的法条编号和法律名称
    relevant_terms = set()
    for ntype, node_list in community_data["nodes_by_type"].items():
        if ntype in ("Section", "Chapter", "Title", "Law"):
            for node in node_list:
                name = node["name"]
                if name:
                    relevant_terms.add(name)

    if not relevant_terms:
        return ""

    # 从原文中匹配相关段落
    matched_paragraphs = []
    policy_lines = policy_text.split("\n")

    for line in policy_lines:
        line_stripped = line.strip()
        if not line_stripped or line_stripped.startswith("==="):
            continue
        for term in relevant_terms:
            # 匹配条款编号，如"第四十一条"
            if term in line_stripped:
                matched_paragraphs.append(line_stripped)
                break

    # 限制长度
    result = "\n".join(matched_paragraphs[:30])
    if len(result) > 3000:
        result = result[:3000] + "\n...(已截断)"

    return result


# ==================== LLM API 调用模块 ====================


def call_deepseek_api(prompt: str, system_prompt: str) -> Optional[str]:
    """
    调用DeepSeek API

    Returns:
        模型返回的文本内容，失败返回None
    """
    if not API_KEY:
        print("  ⚠ 未设置GRAPHRAG_API_KEY环境变量，使用模拟模式")
        return None

    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {API_KEY}",
    }

    payload = {
        "model": MODEL_NAME,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": prompt},
        ],
        "max_tokens": MAX_TOKENS,
        "temperature": TEMPERATURE,
    }

    for attempt in range(MAX_RETRIES):
        try:
            response = requests.post(
                f"{API_BASE}/chat/completions",
                headers=headers,
                json=payload,
                timeout=120,
            )

            if response.status_code == 200:
                result = response.json()
                content = result["choices"][0]["message"]["content"]
                return content
            elif response.status_code == 429:
                wait_time = RETRY_DELAY * (attempt + 1)
                print(f"    ⚠ API频率限制，等待{wait_time}秒后重试...")
                time.sleep(wait_time)
            else:
                print(
                    f"    ⚠ API返回错误 [{response.status_code}]: {response.text[:200]}"
                )
                if attempt < MAX_RETRIES - 1:
                    time.sleep(RETRY_DELAY)

        except requests.exceptions.Timeout:
            print(f"    ⚠ API请求超时 (尝试 {attempt + 1}/{MAX_RETRIES})")
            if attempt < MAX_RETRIES - 1:
                time.sleep(RETRY_DELAY)
        except requests.exceptions.ConnectionError as e:
            print(f"    ⚠ API连接错误: {e}")
            if attempt < MAX_RETRIES - 1:
                time.sleep(RETRY_DELAY)
        except Exception as e:
            print(f"    ⚠ API调用异常: {e}")
            traceback.print_exc()
            if attempt < MAX_RETRIES - 1:
                time.sleep(RETRY_DELAY)

    return None


# ==================== Prompt 构建模块 ====================


def build_community_report_prompt(
    community_data: Dict,
    perspective_name: str,
    relevant_policy: str,
) -> Tuple[str, str]:
    """
    构建社区报告生成的system prompt和user prompt

    Returns:
        (system_prompt, user_prompt)
    """
    system_prompt = f"""你是一位资本市场法规监管领域的专业分析师。你的任务是基于知识图谱的社区聚类结果和相关政策法规原文，为指定社区生成结构化的分析报告。

当前分析视角：{perspective_name}社区

你需要严格按照JSON格式输出报告，包含以下字段：
- title: 社区报告标题（简洁概括社区核心主题，15-30字）
- summary: 执行摘要（一段精炼文字概述社区的整体结构、主要实体、实体间关系以及最显著的特征或洞察，150-300字）
- rating: 重要性/影响力评分（0.0-10.0的浮点数）
- rating_explanation: 评分解释（一句话说明评分依据）
- key_words: 关键词列表，从以下几个角度考虑：针对"信披违规、高价超募、抱团压价、重大舆情、操纵市场"中的哪类/哪几类场景，针对哪些责任承担者进行约束，主要约束的是哪些行为
- findings: 发现/洞察列表（3-6条），每条包含：
  - summary: 一句话概括（10-20字）
  - explanation: 详细解释 + 数据引用（80-200字，引用格式为 [Data: Entities (ID1, ID2), Relationships (ID1, ID2, +more)]）

请仅输出合法的JSON对象，不要包含任何其他文字、Markdown标记或代码块标记。"""

    # 构建社区数据描述
    nodes_desc_parts = []
    for ntype, node_list in community_data["nodes_by_type"].items():
        # 按加权度数排序
        sorted_nodes = sorted(node_list, key=lambda x: x["weighted_degree"], reverse=True)
        node_strs = []
        for n in sorted_nodes[:15]:  # 每类最多展示15个
            core_mark = " [核心]" if n["is_core"] else ""
            node_strs.append(
                f"  - {n['name']} (ID: {n['id']}, "
                f"度数={n['degree']}, 加权度数={n['weighted_degree']:.1f}, "
                f"入度={n['in_degree']}, 出度={n['out_degree']}){core_mark}"
            )
        remaining = len(node_list) - len(node_strs)
        if remaining > 0:
            node_strs.append(f"  ... 还有 {remaining} 个{ntype}类型节点")
        nodes_desc_parts.append(f"\n【{ntype}类型节点】({len(node_list)}个):\n" + "\n".join(node_strs))

    nodes_desc = "\n".join(nodes_desc_parts)

    # 关系描述
    rels_desc_parts = []
    for rel in community_data["top_relations"][:40]:  # prompt中最多40条关系
        rels_desc_parts.append(
            f"  - [{rel['type']}] {rel['source']}({rel['source_type']}) → "
            f"{rel['target']}({rel['target_type']}) "
            f"(ID: {rel['id']}, 权重={rel['weight']:.2f})"
        )
    if len(community_data["top_relations"]) > 40:
        rels_desc_parts.append(
            f"  ... 还有 {len(community_data['top_relations']) - 40} 条关系"
        )
    rels_desc = "\n".join(rels_desc_parts)

    # 统计信息
    stats = community_data["summary_stats"]
    stats_desc = (
        f"社区编号: {community_data['community_id']}\n"
        f"节点总数: {community_data['total_nodes']}\n"
        f"内部关系数: {community_data['total_internal_rels']}\n"
        f"跨社区关系数: {community_data['total_cross_rels']}\n"
        f"核心节点数: {stats.get('核心节点数', 'N/A')}\n"
        f"子图密度: {stats.get('子图密度', 'N/A')}\n"
        f"平均度数: {stats.get('平均度数', 'N/A')}\n"
        f"平均加权度数: {stats.get('平均加权度数', 'N/A')}\n"
        f"总权重: {stats.get('总权重', 'N/A')}\n"
        f"节点类型分布: {stats.get('节点类型分布', 'N/A')}"
    )

    # 组装user prompt
    user_prompt = f"""请为以下{perspective_name}社区生成结构化分析报告。

## 一、社区统计信息
{stats_desc}

## 二、社区节点详情
{nodes_desc}

## 三、社区内部关系（按权重降序）
{rels_desc}

"""

    if relevant_policy:
        user_prompt += f"""## 四、相关政策法规原文
{relevant_policy}

"""

    user_prompt += """请基于以上信息生成JSON格式的社区报告。注意：
1. findings中的explanation必须引用具体的实体ID和关系ID
2. key_words要从"信披违规、高价超募、抱团压价、重大舆情、操纵市场"等场景角度考虑
3. 评分(rating)需要综合考虑社区的规模、核心节点数、密度和法律影响力
4. 仅输出JSON对象，不要包含```json```标记或其他文字"""

    return system_prompt, user_prompt


# ==================== 报告解析模块 ====================


def parse_llm_response(response_text: str, community_id: int) -> Optional[Dict]:
    """
    解析LLM返回的JSON报告

    Returns:
        解析后的报告字典，失败返回None
    """
    if not response_text:
        return None

    # 尝试清理常见的格式问题
    text = response_text.strip()

    # 去除可能的markdown代码块标记
    if text.startswith("```json"):
        text = text[7:]
    elif text.startswith("```"):
        text = text[3:]
    if text.endswith("```"):
        text = text[:-3]
    text = text.strip()

    try:
        report = json.loads(text)
        return report
    except json.JSONDecodeError as e:
        print(f"    ⚠ JSON解析失败 (社区 {community_id}): {e}")
        # 尝试找到JSON对象的边界
        start_idx = text.find("{")
        end_idx = text.rfind("}")
        if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
            try:
                report = json.loads(text[start_idx : end_idx + 1])
                return report
            except json.JSONDecodeError:
                pass
        print(f"    ⚠ 无法解析LLM响应，将使用回退方案")
        return None


def generate_fallback_report(community_data: Dict, perspective_name: str) -> Dict:
    """
    当LLM调用失败时，基于图谱数据生成基础报告（回退方案）
    """
    stats = community_data["summary_stats"]

    # 收集核心节点名称
    core_names = []
    all_names = []
    for ntype, node_list in community_data["nodes_by_type"].items():
        for n in node_list:
            all_names.append(n["name"])
            if n["is_core"]:
                core_names.append(n["name"])

    # 收集关系类型
    rel_types = set()
    for rel in community_data["top_relations"]:
        rel_types.add(rel["type"])

    # 生成基础标题
    top_entities = core_names[:3] if core_names else all_names[:3]
    title = f"{perspective_name}社区{community_data['community_id']}: " + "、".join(
        top_entities[:3]
    )
    if len(title) > 50:
        title = title[:47] + "..."

    # 生成基础摘要
    summary = (
        f"本社区（{perspective_name}视角-社区{community_data['community_id']}）"
        f"包含{community_data['total_nodes']}个节点和"
        f"{community_data['total_internal_rels']}条内部关系。"
        f"核心节点包括：{'、'.join(core_names[:5]) if core_names else '无明确核心节点'}。"
        f"主要关系类型包括：{'、'.join(list(rel_types)[:5])}。"
        f"社区子图密度为{stats.get('子图密度', 'N/A')}，"
        f"平均加权度数为{stats.get('平均加权度数', 'N/A')}。"
    )

    # 基础发现
    findings = []
    if core_names:
        # 找到最高加权度数节点
        all_nodes_flat = []
        for ntype, node_list in community_data["nodes_by_type"].items():
            all_nodes_flat.extend(node_list)
        all_nodes_flat.sort(key=lambda x: x["weighted_degree"], reverse=True)
        top_node = all_nodes_flat[0] if all_nodes_flat else None

        if top_node:
            findings.append(
                {
                    "summary": f"{top_node['name']}为社区核心枢纽",
                    "explanation": (
                        f"{top_node['name']}在社区内加权度数最高"
                        f"（weightedDegree={top_node['weighted_degree']:.1f}），"
                        f"度数为{top_node['degree']}，"
                        f"表明其在该社区中处于核心连接位置。"
                        f"[Data: Entities ({top_node['id']})]"
                    ),
                }
            )

    if community_data["top_relations"]:
        top_rel = community_data["top_relations"][0]
        findings.append(
            {
                "summary": f"最强关系为「{top_rel['type']}」",
                "explanation": (
                    f"社区内权重最高的关系为 {top_rel['source']} → {top_rel['target']}，"
                    f"关系类型为「{top_rel['type']}」，权重={top_rel['weight']:.2f}。"
                    f"[Data: Relationships ({top_rel['id']})]"
                ),
            }
        )

    findings.append(
        {
            "summary": "社区结构特征",
            "explanation": (
                f"该社区包含{community_data['total_nodes']}个节点和"
                f"{community_data['total_internal_rels']}条内部关系，"
                f"跨社区关系{community_data['total_cross_rels']}条，"
                f"子图密度{stats.get('子图密度', 'N/A')}。"
                f"节点类型分布：{stats.get('节点类型分布', 'N/A')}。"
            ),
        }
    )

    # 基础关键词
    key_words = list(rel_types)[:5]
    for n in core_names[:3]:
        if n not in key_words:
            key_words.append(n)

    return {
        "title": title,
        "summary": summary,
        "rating": round(
            min(
                community_data["total_nodes"] * 0.3
                + community_data["total_internal_rels"] * 0.1
                + int(stats.get("核心节点数", 0)) * 0.5,
                10.0,
            ),
            1,
        ),
        "rating_explanation": f"基于社区规模({community_data['total_nodes']}节点)和核心节点数({stats.get('核心节点数', 0)})的自动评分",
        "key_words": key_words,
        "findings": findings,
    }


# ==================== 报告生成主流程 ====================


def generate_report_id(perspective: str, community_id: int) -> str:
    """生成全局唯一报告ID"""
    raw = f"{perspective}_{community_id}_{uuid.uuid4().hex[:8]}"
    return hashlib.md5(raw.encode()).hexdigest()[:16]


def generate_community_report(
    community_id: int,
    df_nodes: pd.DataFrame,
    df_rels: pd.DataFrame,
    df_summary: pd.DataFrame,
    nodes_dict: Dict,
    policy_text: str,
    perspective_key: str,
    perspective_name: str,
) -> Dict:
    """
    为单个社区生成完整报告

    Returns:
        包含所有必要字段的报告字典
    """
    # 提取社区数据
    community_data = extract_community_data(
        community_id, df_nodes, df_rels, df_summary, nodes_dict
    )

    # 获取相关政策原文
    relevant_policy = find_relevant_policy_text(community_data, nodes_dict, policy_text)

    # 构建prompt
    system_prompt, user_prompt = build_community_report_prompt(
        community_data, perspective_name, relevant_policy
    )

    # 调用LLM
    print(f"    → 调用API生成社区 {community_id} 报告...")
    llm_response = call_deepseek_api(user_prompt, system_prompt)

    # 解析响应
    report_content = parse_llm_response(llm_response, community_id)

    # 如果LLM失败，使用回退方案
    if report_content is None:
        print(f"    → 使用回退方案生成社区 {community_id} 报告")
        report_content = generate_fallback_report(community_data, perspective_name)

    # 组装完整报告记录
    report_id = generate_report_id(perspective_key, community_id)

    # 确保findings是正确格式
    findings = report_content.get("findings", [])
    if isinstance(findings, str):
        try:
            findings = json.loads(findings)
        except json.JSONDecodeError:
            findings = [{"summary": findings, "explanation": findings}]

    # 构建full_content_json
    full_content = {
        "title": report_content.get("title", f"{perspective_name}社区{community_id}"),
        "summary": report_content.get("summary", ""),
        "rating": float(report_content.get("rating", 5.0)),
        "rating_explanation": report_content.get("rating_explanation", ""),
        "key_words": report_content.get("key_words", []),
        "findings": findings,
    }

    report_record = {
        "id": report_id,
        "community": community_id,
        "level": 0,  # 单层级Leiden，所有社区为叶子层级
        "title": full_content["title"],
        "parent": -1,  # 无父社区（单层级）
        "children": json.dumps([], ensure_ascii=False),  # 无子社区
        "summary": full_content["summary"],
        "key_words": json.dumps(
            full_content["key_words"], ensure_ascii=False
        ),
        "findings": json.dumps(findings, ensure_ascii=False),
        "rank": full_content["rating"],
        "rating_explanation": full_content["rating_explanation"],
        "full_content_json": json.dumps(full_content, ensure_ascii=False),
    }

    return report_record


def generate_reports_for_perspective(
    perspective_key: str,
    perspective_name: str,
    result_dir: str,
    nodes_dict: Dict,
    policy_text: str,
) -> List[Dict]:
    """
    为某个视角的所有社区生成报告

    Returns:
        报告记录列表
    """
    print(f"\n{'='*60}")
    print(f"  生成{perspective_name}社区报告")
    print(f"{'='*60}")

    # 加载聚类结果
    df_nodes, df_rels, df_summary, vis_data = load_clustering_results(
        result_dir, perspective_key
    )

    # 获取所有社区ID
    community_ids = sorted(df_summary["社区ID"].unique().tolist())
    print(f"  共 {len(community_ids)} 个社区需要生成报告")

    reports = []
    for idx, cid in enumerate(community_ids):
        print(f"\n  [{idx + 1}/{len(community_ids)}] 社区 {cid}:")

        report = generate_community_report(
            community_id=cid,
            df_nodes=df_nodes,
            df_rels=df_rels,
            df_summary=df_summary,
            nodes_dict=nodes_dict,
            policy_text=policy_text,
            perspective_key=perspective_key,
            perspective_name=perspective_name,
        )
        reports.append(report)
        print(f"    ✓ 完成 - {report['title'][:40]}...")

        # 请求间隔
        if idx < len(community_ids) - 1:
            time.sleep(REQUEST_INTERVAL)

    return reports


# ==================== Excel 导出模块 ====================


def save_reports_to_excel(reports: List[Dict], output_path: str, perspective_name: str):
    """
    将社区报告保存为格式化的Excel文件
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{perspective_name}社区报告"

    # 列定义
    columns = [
        ("id", 18),
        ("community", 12),
        ("level", 8),
        ("title", 40),
        ("parent", 10),
        ("children", 12),
        ("summary", 60),
        ("key_words", 35),
        ("findings", 70),
        ("rank", 10),
        ("rating_explanation", 45),
        ("full_content_json", 80),
    ]

    # 表头样式
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 写表头
    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # 数据样式
    data_font = Font(name="Arial", size=10)
    data_alignment = Alignment(vertical="top", wrap_text=True)
    alt_fill = PatternFill("solid", fgColor="F2F7FB")

    # 写数据
    for row_idx, report in enumerate(reports, 2):
        is_alt = (row_idx % 2) == 0
        for col_idx, (col_name, _) in enumerate(columns, 1):
            value = report.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            if is_alt:
                cell.fill = alt_fill

    # 冻结首行
    ws.freeze_panes = "A2"
    # 自动筛选
    ws.auto_filter.ref = ws.dimensions

    # 设置行高
    ws.row_dimensions[1].height = 30
    for row_idx in range(2, len(reports) + 2):
        ws.row_dimensions[row_idx].height = 80

    wb.save(output_path)
    print(f"  ✓ 报告已保存: {output_path} ({len(reports)} 条记录)")


# ==================== 主函数 ====================


def main():
    """主函数"""
    print("=" * 60)
    print("   社区报告生成器")
    print("   基于Leiden聚类结果 + DeepSeek API")
    print("=" * 60)

    # ---------- 配置检查 ----------
    if not API_KEY:
        print("\n⚠ 警告: 未检测到GRAPHRAG_API_KEY环境变量")
        print("  将使用回退方案（基于图谱数据自动生成基础报告）")
        print("  若要使用LLM生成高质量报告，请设置环境变量:")
        print("  export GRAPHRAG_API_KEY='sk-0a57f72b50854ace9d134a5eb697c4dc'\n")

    # ---------- 检查聚类结果是否存在 ----------
    if not os.path.exists(CLUSTERING_RESULT_DIR):
        print(f"\n未找到聚类结果目录: {CLUSTERING_RESULT_DIR}")
        print("正在先运行Leiden聚类分析...")

        # 确保数据文件存在
        if not os.path.exists(GRAPH_DATA_FILE):
            print(f"错误: 图谱数据文件不存在: {GRAPH_DATA_FILE}")
            sys.exit(1)

        # 尝试运行聚类脚本
        import subprocess

        clustering_script = "weighted_leiden_clustering.py"
        if not os.path.exists(clustering_script):
            print(f"错误: 聚类脚本不存在: {clustering_script}")
            sys.exit(1)

        result = subprocess.run(
            [sys.executable, clustering_script], capture_output=True, text=True
        )
        if result.returncode != 0:
            print(f"聚类脚本执行失败:\n{result.stderr}")
            sys.exit(1)
        print("聚类分析完成\n")

    # ---------- 加载图谱数据 ----------
    print("\n[1/4] 加载知识图谱数据...")
    nodes_dict, relationships_list = load_graph_data(GRAPH_DATA_FILE)

    # ---------- 加载政策原文 ----------
    print("\n[2/4] 加载政策法规原文...")
    policy_text = load_policy_documents(POLICY_DOC_DIR)

    # ---------- 为三种视角生成报告 ----------
    print("\n[3/4] 为各视角社区生成报告...")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    all_perspective_reports = {}

    for perspective_key, perspective_name in PERSPECTIVES.items():
        # 检查该视角的聚类结果是否存在
        nodes_file = os.path.join(
            CLUSTERING_RESULT_DIR, f"{perspective_key}_1_节点聚类结果.csv"
        )
        if not os.path.exists(nodes_file):
            print(f"\n  ⚠ 跳过{perspective_name}视角: 聚类结果文件不存在")
            continue

        reports = generate_reports_for_perspective(
            perspective_key=perspective_key,
            perspective_name=perspective_name,
            result_dir=CLUSTERING_RESULT_DIR,
            nodes_dict=nodes_dict,
            policy_text=policy_text,
        )
        all_perspective_reports[perspective_key] = reports

    # ---------- 导出Excel ----------
    print(f"\n[4/4] 导出Excel报告文件...")

    for perspective_key, reports in all_perspective_reports.items():
        perspective_name = PERSPECTIVES[perspective_key]
        output_path = os.path.join(
            OUTPUT_DIR, f"{perspective_name}社区报告.xlsx"
        )
        save_reports_to_excel(reports, output_path, perspective_name)

    # ---------- 完成 ----------
    print(f"\n{'='*60}")
    print("              社区报告生成完成!")
    print(f"{'='*60}")
    print(f"输出目录: {OUTPUT_DIR}")
    for perspective_key, reports in all_perspective_reports.items():
        perspective_name = PERSPECTIVES[perspective_key]
        print(f"  - {perspective_name}社区报告.xlsx ({len(reports)} 个社区)")
    print(f"{'='*60}")

    return all_perspective_reports


if __name__ == "__main__":
    main()
