#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
监管违规穿透式查询系统
基于社区检测和知识图谱的事件分析系统
"""

import json
import pandas as pd
import networkx as nx
import matplotlib.pyplot as plt
from typing import Dict, List, Tuple, Set
import requests
from collections import defaultdict
import warnings
warnings.filterwarnings('ignore')

# 配置matplotlib支持中文
plt.rcParams['font.sans-serif'] = ['SimHei', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False

# DeepSeek API配置
API_BASE = "https://api.deepseek.com/v1"
API_KEY = "sk-0a57f72b50854ace9d134a5eb697c4dc"  # 用户需要填入API密钥

class RegulatoryQuerySystem:
    """监管违规穿透式查询系统"""
    
    def __init__(self, data_dir="data"):
        """初始化系统"""
        self.data_dir = data_dir
        self.load_all_data()
        
    def load_all_data(self):
        """加载所有数据文件"""
        print("正在加载数据...")
        
        # 加载三种视角的社区可视化数据
        with open(f"{self.data_dir}/regulatory_visualization_data.json", 'r', encoding='utf-8') as f:
            self.regulatory_data = json.load(f)
        with open(f"{self.data_dir}/responsibility_visualization_data.json", 'r', encoding='utf-8') as f:
            self.responsibility_data = json.load(f)
        with open(f"{self.data_dir}/violation_visualization_data.json", 'r', encoding='utf-8') as f:
            self.violation_data = json.load(f)
            
        # 加载社区报告
        self.regulatory_reports = pd.read_excel(f"{self.data_dir}/监管机构社区报告.xlsx")
        self.responsibility_reports = pd.read_excel(f"{self.data_dir}/责任方社区报告.xlsx")
        self.violation_reports = pd.read_excel(f"{self.data_dir}/违规行为社区报告.xlsx")
        
        # 加载社区层级关系
        self.community_hierarchy = pd.read_excel(f"{self.data_dir}/community_hierarchy_v2.xlsx")
        
        # 加载知识图谱数据
        self.kg_nodes = []
        self.kg_edges = []
        with open(f"{self.data_dir}/merged_regulatory_unified.txt", 'r', encoding='utf-8') as f:
            for line in f:
                data = json.loads(line.strip())
                if data['type'] == 'node':
                    self.kg_nodes.append(data)
                elif data['type'] == 'relationship':
                    self.kg_edges.append(data)
        
        # 构建索引
        self._build_indices()
        print(f"数据加载完成！")
        print(f"- 监管机构社区: {len(set([n['community'] for n in self.regulatory_data['nodes']]))} 个")
        print(f"- 责任方社区: {len(set([n['community'] for n in self.responsibility_data['nodes']]))} 个")
        print(f"- 违规行为社区: {len(set([n['community'] for n in self.violation_data['nodes']]))} 个")
        print(f"- 知识图谱节点: {len(self.kg_nodes)} 个")
        print(f"- 知识图谱关系: {len(self.kg_edges)} 个")
        
    def _build_indices(self):
        """构建快速检索索引"""
        # 构建节点名称到社区的映射
        self.node_to_community = {
            'regulatory': {},
            'responsibility': {},
            'violation': {}
        }
        
        for node in self.regulatory_data['nodes']:
            name = node['name'].lower()
            community = node['community']
            if name not in self.node_to_community['regulatory']:
                self.node_to_community['regulatory'][name] = set()
            self.node_to_community['regulatory'][name].add(community)
            
        for node in self.responsibility_data['nodes']:
            name = node['name'].lower()
            community = node['community']
            if name not in self.node_to_community['responsibility']:
                self.node_to_community['responsibility'][name] = set()
            self.node_to_community['responsibility'][name].add(community)
            
        for node in self.violation_data['nodes']:
            name = node['name'].lower()
            community = node['community']
            if name not in self.node_to_community['violation']:
                self.node_to_community['violation'][name] = set()
            self.node_to_community['violation'][name].add(community)
    
    def call_deepseek_api(self, prompt: str, max_tokens: int = 2000) -> str:
        """调用DeepSeek API"""
        if not API_KEY:
            raise ValueError("请先设置API_KEY！")
            
        headers = {
            "Authorization": f"Bearer {API_KEY}",
            "Content-Type": "application/json"
        }
        
        data = {
            "model": "deepseek-chat",
            "messages": [
                {"role": "user", "content": prompt}
            ],
            "max_tokens": max_tokens,
            "temperature": 0.3
        }
        
        response = requests.post(f"{API_BASE}/chat/completions", headers=headers, json=data)
        
        if response.status_code == 200:
            return response.json()['choices'][0]['message']['content']
        else:
            raise Exception(f"API调用失败: {response.status_code}, {response.text}")
    
    def extract_entities(self, event_description: str) -> Dict[str, List[str]]:
        """从事件描述中提取关键实体"""
        print("\n步骤1: 提取事件中的关键实体...")
        
        prompt = f"""请从以下事件描述中提取关键实体，并分类为：责任方、违规行为、监管机构。

事件描述：
{event_description}

请以JSON格式返回，格式如下：
{{
    "责任方": ["实体1", "实体2", ...],
    "违规行为": ["行为1", "行为2", ...],
    "监管机构": ["机构1", "机构2", ...]
}}

只返回JSON，不要其他解释。"""

        try:
            response = self.call_deepseek_api(prompt)
            # 提取JSON部分
            start = response.find('{')
            end = response.rfind('}') + 1
            json_str = response[start:end]
            entities = json.loads(json_str)
            
            print(f"提取到的实体：")
            for category, items in entities.items():
                print(f"  {category}: {items}")
            
            return entities
        except Exception as e:
            print(f"API调用失败，使用简单分词: {e}")
            # 备用方案：简单分词
            return {
                "责任方": [],
                "违规行为": [],
                "监管机构": []
            }
    
    def match_communities(self, entities: Dict[str, List[str]]) -> Dict[str, Set[int]]:
        """将提取的实体与社区节点进行匹配"""
        print("\n步骤2: 匹配相关社区...")
        
        matched_communities = {
            'regulatory': set(),
            'responsibility': set(),
            'violation': set()
        }
        
        # 匹配责任方
        for entity in entities.get('责任方', []):
            entity_lower = entity.lower()
            for node_name, communities in self.node_to_community['responsibility'].items():
                if entity_lower in node_name or node_name in entity_lower:
                    matched_communities['responsibility'].update(communities)
        
        # 匹配违规行为
        for entity in entities.get('违规行为', []):
            entity_lower = entity.lower()
            for node_name, communities in self.node_to_community['violation'].items():
                if entity_lower in node_name or node_name in entity_lower:
                    matched_communities['violation'].update(communities)
        
        # 匹配监管机构
        for entity in entities.get('监管机构', []):
            entity_lower = entity.lower()
            for node_name, communities in self.node_to_community['regulatory'].items():
                if entity_lower in node_name or node_name in entity_lower:
                    matched_communities['regulatory'].update(communities)
        
        print(f"匹配到的社区：")
        print(f"  责任方社区: {sorted(matched_communities['responsibility'])}")
        print(f"  违规行为社区: {sorted(matched_communities['violation'])}")
        print(f"  监管机构社区: {sorted(matched_communities['regulatory'])}")
        
        return matched_communities
    
    def build_paths(self, matched_communities: Dict[str, Set[int]]) -> List[Tuple]:
        """基于社区层级关系构建链路路径"""
        print("\n步骤3: 构建链路路径...")
        
        paths = []
        
        # 从社区层级关系中查找连接
        for _, row in self.community_hierarchy.iterrows():
            source_perspective = row['source_perspective']
            source_id = row['source_community_id']
            target_perspective = row['target_perspective']
            target_id = row['target_community_id']
            relation_type = row['relation_type']
            is_strong = row['is_strong_link']
            
            # 映射视角名称
            perspective_map = {
                'responsibility': 'responsibility',
                'violation': 'violation',
                'regulatory': 'regulatory'
            }
            
            source_key = perspective_map.get(source_perspective)
            target_key = perspective_map.get(target_perspective)
            
            if source_key and target_key:
                if (source_id in matched_communities[source_key] and 
                    target_id in matched_communities[target_key]):
                    paths.append((
                        source_perspective, source_id,
                        target_perspective, target_id,
                        relation_type, is_strong
                    ))
        
        print(f"找到 {len(paths)} 条路径连接")
        return paths
    
    def visualize_network(self, matched_communities: Dict[str, Set[int]], 
                         paths: List[Tuple], output_file: str):
        """可视化社区网络"""
        print("\n步骤4: 可视化网络...")
        
        G = nx.DiGraph()
        
        # 添加节点
        node_labels = {}
        node_colors = []
        color_map = {
            'responsibility': '#ff7f0e',  # 橙色
            'violation': '#d62728',       # 红色
            'regulatory': '#1f77b4'       # 蓝色
        }
        
        for perspective, communities in matched_communities.items():
            for comm_id in communities:
                node_id = f"{perspective}_{comm_id}"
                G.add_node(node_id)
                
                # 获取社区标题
                if perspective == 'responsibility':
                    report = self.responsibility_reports[
                        self.responsibility_reports['community'] == comm_id
                    ]
                elif perspective == 'violation':
                    report = self.violation_reports[
                        self.violation_reports['community'] == comm_id
                    ]
                else:
                    report = self.regulatory_reports[
                        self.regulatory_reports['community'] == comm_id
                    ]
                
                if not report.empty:
                    title = report.iloc[0]['title']
                    node_labels[node_id] = f"{perspective[:4]}\n#{comm_id}\n{title[:15]}..."
                else:
                    node_labels[node_id] = f"{perspective[:4]}\n#{comm_id}"
                
                node_colors.append(color_map[perspective])
        
        # 添加边
        for path in paths:
            source_p, source_id, target_p, target_id, rel_type, is_strong = path
            source_node = f"{source_p}_{source_id}"
            target_node = f"{target_p}_{target_id}"
            
            if source_node in G.nodes() and target_node in G.nodes():
                G.add_edge(source_node, target_node, 
                          relation=rel_type, 
                          weight=2 if is_strong else 1)
        
        # 绘图
        plt.figure(figsize=(16, 12))
        pos = nx.spring_layout(G, k=2, iterations=50)
        
        nx.draw_networkx_nodes(G, pos, node_color=node_colors, 
                              node_size=3000, alpha=0.8)
        nx.draw_networkx_labels(G, pos, node_labels, font_size=8)
        
        # 绘制边
        edges = G.edges()
        weights = [G[u][v]['weight'] for u, v in edges]
        nx.draw_networkx_edges(G, pos, width=weights, alpha=0.5,
                              edge_color='gray', arrows=True,
                              arrowsize=20, arrowstyle='->')
        
        plt.title("监管违规事件穿透式分析网络", fontsize=16, pad=20)
        plt.axis('off')
        plt.tight_layout()
        plt.savefig(output_file, dpi=300, bbox_inches='tight')
        print(f"网络图已保存至: {output_file}")
        plt.close()
        
        return G
    
    def get_community_reports(self, matched_communities: Dict[str, Set[int]]) -> str:
        """获取相关社区的报告内容"""
        reports = []
        
        # 责任方社区报告
        if matched_communities['responsibility']:
            reports.append("=" * 80)
            reports.append("责任方社区报告")
            reports.append("=" * 80)
            for comm_id in sorted(matched_communities['responsibility']):
                report = self.responsibility_reports[
                    self.responsibility_reports['community'] == comm_id
                ]
                if not report.empty:
                    row = report.iloc[0]
                    reports.append(f"\n【社区 #{comm_id}】{row['title']}")
                    reports.append(f"摘要: {row['summary']}")
                    reports.append(f"关键词: {row['key_words']}")
                    if pd.notna(row['findings']):
                        reports.append(f"发现: {row['findings']}")
        
        # 违规行为社区报告
        if matched_communities['violation']:
            reports.append("\n" + "=" * 80)
            reports.append("违规行为社区报告")
            reports.append("=" * 80)
            for comm_id in sorted(matched_communities['violation']):
                report = self.violation_reports[
                    self.violation_reports['community'] == comm_id
                ]
                if not report.empty:
                    row = report.iloc[0]
                    reports.append(f"\n【社区 #{comm_id}】{row['title']}")
                    reports.append(f"摘要: {row['summary']}")
                    reports.append(f"关键词: {row['key_words']}")
                    if pd.notna(row['findings']):
                        reports.append(f"发现: {row['findings']}")
        
        # 监管机构社区报告
        if matched_communities['regulatory']:
            reports.append("\n" + "=" * 80)
            reports.append("监管机构社区报告")
            reports.append("=" * 80)
            for comm_id in sorted(matched_communities['regulatory']):
                report = self.regulatory_reports[
                    self.regulatory_reports['community'] == comm_id
                ]
                if not report.empty:
                    row = report.iloc[0]
                    reports.append(f"\n【社区 #{comm_id}】{row['title']}")
                    reports.append(f"摘要: {row['summary']}")
                    reports.append(f"关键词: {row['key_words']}")
                    if pd.notna(row['findings']):
                        reports.append(f"发现: {row['findings']}")
        
        return "\n".join(reports)
    
    def generate_final_answer(self, event_description: str, 
                            community_reports: str,
                            paths: List[Tuple]) -> str:
        """结合社区报告和知识图谱生成最终答案"""
        print("\n步骤5: 生成综合分析报告...")
        
        # 构建路径描述
        path_descriptions = []
        for i, path in enumerate(paths, 1):
            source_p, source_id, target_p, target_id, rel_type, is_strong = path
            strength = "强关联" if is_strong else "弱关联"
            path_descriptions.append(
                f"{i}. {source_p}社区#{source_id} → ({rel_type}, {strength}) → "
                f"{target_p}社区#{target_id}"
            )
        
        prompt = f"""作为资本市场监管专家，请基于以下信息对给定事件进行深度分析：

【事件描述】
{event_description}

【识别的链路路径】
{chr(10).join(path_descriptions)}

【相关社区报告】
{community_reports}

请从以下几个方面进行分析：
1. 事件涉及的主要责任方及其法律地位
2. 涉及的违规行为类型及法律依据
3. 相关监管机构的职责和监管依据
4. 可能触发的法律责任和处罚措施
5. 合规建议和风险提示

请提供专业、全面、结构化的分析报告。"""

        try:
            response = self.call_deepseek_api(prompt, max_tokens=4000)
            return response
        except Exception as e:
            print(f"生成最终报告失败: {e}")
            return "无法生成最终报告，请检查API配置。"
    
    def query(self, event_description: str, output_dir: str = "output"):
        """执行完整的穿透式查询流程"""
        print("=" * 80)
        print("监管违规穿透式查询系统")
        print("=" * 80)
        print(f"\n事件描述:\n{event_description}\n")
        
        # 步骤1: 提取实体
        entities = self.extract_entities(event_description)
        
        # 步骤2: 匹配社区
        matched_communities = self.match_communities(entities)
        
        # 如果没有匹配到社区，尝试使用关键词匹配
        if not any(matched_communities.values()):
            print("\n未匹配到社区，尝试关键词匹配...")
            # 使用事件描述中的所有词进行模糊匹配
            keywords = event_description.replace('，', ' ').replace('。', ' ').split()
            for keyword in keywords:
                if len(keyword) >= 2:
                    kw_lower = keyword.lower()
                    for node_name, communities in self.node_to_community['responsibility'].items():
                        if kw_lower in node_name:
                            matched_communities['responsibility'].update(communities)
                    for node_name, communities in self.node_to_community['violation'].items():
                        if kw_lower in node_name:
                            matched_communities['violation'].update(communities)
                    for node_name, communities in self.node_to_community['regulatory'].items():
                        if kw_lower in node_name:
                            matched_communities['regulatory'].update(communities)
        
        # 步骤3: 构建路径
        paths = self.build_paths(matched_communities)
        
        # 步骤4: 可视化
        network_file = f"{output_dir}/event_network.png"
        self.visualize_network(matched_communities, paths, network_file)
        
        # 步骤5: 获取社区报告
        community_reports = self.get_community_reports(matched_communities)
        
        # 保存社区报告
        report_file = f"{output_dir}/community_reports.txt"
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write(community_reports)
        print(f"社区报告已保存至: {report_file}")
        
        # 步骤6: 生成最终答案
        final_answer = self.generate_final_answer(event_description, 
                                                  community_reports, paths)
        
        # 保存最终报告
        final_report_file = f"{output_dir}/final_analysis.txt"
        with open(final_report_file, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("监管违规事件综合分析报告\n")
            f.write("=" * 80 + "\n\n")
            f.write(f"【事件描述】\n{event_description}\n\n")
            f.write(f"【分析结果】\n{final_answer}\n")
        print(f"最终分析报告已保存至: {final_report_file}")
        
        # 生成汇总Excel
        self.generate_summary_excel(event_description, matched_communities, 
                                   paths, f"{output_dir}/analysis_summary.xlsx")
        
        print("\n" + "=" * 80)
        print("查询完成！")
        print("=" * 80)
        
        return {
            'entities': entities,
            'matched_communities': matched_communities,
            'paths': paths,
            'network_file': network_file,
            'report_file': report_file,
            'final_report_file': final_report_file,
            'final_answer': final_answer
        }
    
    def generate_summary_excel(self, event_description: str,
                              matched_communities: Dict[str, Set[int]],
                              paths: List[Tuple], output_file: str):
        """生成汇总Excel报告"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        
        # 工作表1: 事件概述
        ws1 = wb.active
        ws1.title = "事件概述"
        ws1['A1'] = "监管违规事件分析报告"
        ws1['A1'].font = Font(size=16, bold=True)
        ws1['A2'] = "事件描述"
        ws1['A2'].font = Font(bold=True)
        ws1['A3'] = event_description
        ws1.merge_cells('A3:D3')
        
        # 工作表2: 匹配的社区
        ws2 = wb.create_sheet("匹配社区")
        ws2.append(['视角', '社区ID', '社区标题', '摘要'])
        ws2['A1'].font = Font(bold=True)
        ws2['B1'].font = Font(bold=True)
        ws2['C1'].font = Font(bold=True)
        ws2['D1'].font = Font(bold=True)
        
        for perspective, communities in matched_communities.items():
            for comm_id in sorted(communities):
                if perspective == 'responsibility':
                    report = self.responsibility_reports[
                        self.responsibility_reports['community'] == comm_id
                    ]
                    p_name = "责任方"
                elif perspective == 'violation':
                    report = self.violation_reports[
                        self.violation_reports['community'] == comm_id
                    ]
                    p_name = "违规行为"
                else:
                    report = self.regulatory_reports[
                        self.regulatory_reports['community'] == comm_id
                    ]
                    p_name = "监管机构"
                
                if not report.empty:
                    row = report.iloc[0]
                    ws2.append([p_name, comm_id, row['title'], row['summary']])
        
        # 调整列宽
        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 10
        ws2.column_dimensions['C'].width = 40
        ws2.column_dimensions['D'].width = 60
        
        # 工作表3: 路径分析
        ws3 = wb.create_sheet("路径分析")
        ws3.append(['序号', '源视角', '源社区', '目标视角', '目标社区', '关系类型', '关联强度'])
        for i in range(1, 8):
            ws3.cell(1, i).font = Font(bold=True)
        
        for i, path in enumerate(paths, 1):
            source_p, source_id, target_p, target_id, rel_type, is_strong = path
            ws3.append([i, source_p, source_id, target_p, target_id, 
                       rel_type, "强" if is_strong else "弱"])
        
        ws3.column_dimensions['A'].width = 8
        ws3.column_dimensions['B'].width = 15
        ws3.column_dimensions['C'].width = 10
        ws3.column_dimensions['D'].width = 15
        ws3.column_dimensions['E'].width = 10
        ws3.column_dimensions['F'].width = 20
        ws3.column_dimensions['G'].width = 10
        
        wb.save(output_file)
        print(f"汇总Excel已保存至: {output_file}")


def main():
    """主函数"""
    print("初始化监管违规穿透式查询系统...")
    system = RegulatoryQuerySystem()
    
    print("\n" + "=" * 80)
    print("请输入事件描述（输入完成后按Enter，再输入'END'并按Enter结束）：")
    print("=" * 80)
    
    lines = []
    while True:
        line = input()
        if line.strip() == 'END':
            break
        lines.append(line)
    
    event_description = '\n'.join(lines)
    
    if not event_description.strip():
        print("错误: 事件描述不能为空！")
        return
    
    # 执行查询
    results = system.query(event_description)
    
    # 打印最终答案
    print("\n" + "=" * 80)
    print("最终分析结果:")
    print("=" * 80)
    print(results['final_answer'])


if __name__ == "__main__":
    main()
