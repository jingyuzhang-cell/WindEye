#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
监管违规穿透式查询系统 v2
================================================================
核心升级：
1. 节点级粒度穿透路径（原系统只到社区级别）
   - 合规路径：责任方节点 → 责任义务节点 → 监管方节点
   - 违规路径：责任方节点 → 违规行为节点 → 监管方节点
2. 双层可视化
   - 社区级别可视化（保留原 event_network.png）
   - 节点级别可视化（新增 node_level_path.png）
3. 优化 Prompt 构建
   - 同时注入：社区报告 + 路径节点的原始图谱三元组
================================================================
"""

import json
import os
import re
import warnings
from collections import defaultdict
from typing import Dict, List, Set, Tuple, Optional

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import FancyBboxPatch, FancyArrowPatch
import networkx as nx
import pandas as pd
import requests

warnings.filterwarnings('ignore')

plt.rcParams['font.sans-serif'] = ['SimHei', 'DejaVu Sans', 'Arial Unicode MS', 'sans-serif']
plt.rcParams['axes.unicode_minus'] = False

# ==================== 配置 ====================
API_BASE = "https://api.deepseek.com/v1"
API_KEY = "sk-0a57f72b50854ace9d134a5eb697c4dc"


# ==================== 工具函数 ====================

def remove_markdown_formatting(text: str) -> str:
    """去除文本中的 Markdown 格式"""
    if not text:
        return text
    text = re.sub(r'```[\w]*\n', '', text)
    text = re.sub(r'```', '', text)
    text = re.sub(r'\*\*(.+?)\*\*', r'\1', text)
    text = re.sub(r'__(.+?)__', r'\1', text)
    text = re.sub(r'\*(.+?)\*', r'\1', text)
    text = re.sub(r'_(.+?)_', r'\1', text)
    text = re.sub(r'^#{1,6}\s+', '', text, flags=re.MULTILINE)
    text = re.sub(r'\[([^\]]+)\]\([^\)]+\)', r'\1', text)
    text = re.sub(r'`([^`]+)`', r'\1', text)
    text = re.sub(r'^[\*\-_]{3,}\s*$', '', text, flags=re.MULTILINE)
    text = re.sub(r'^[\*\-\+]\s+', '• ', text, flags=re.MULTILINE)
    text = re.sub(r'^\d+\.\s+', '', text, flags=re.MULTILINE)
    text = re.sub(r'^>\s+', '', text, flags=re.MULTILINE)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


# ==================== 数据加载与索引 ====================

class DataLoader:
    """统一数据加载器"""

    def __init__(self, data_dir: str = "data"):
        self.data_dir = data_dir
        self.regulatory_data: dict = {}
        self.responsibility_data: dict = {}
        self.violation_data: dict = {}
        self.regulatory_reports: pd.DataFrame = pd.DataFrame()
        self.responsibility_reports: pd.DataFrame = pd.DataFrame()
        self.violation_reports: pd.DataFrame = pd.DataFrame()
        self.community_hierarchy: pd.DataFrame = pd.DataFrame()
        self.kg_nodes: List[dict] = []
        self.kg_edges: List[dict] = []
        # 构建后的索引
        self.kg_node_map: Dict[str, dict] = {}       # id -> node
        self.kg_edge_list: List[dict] = []
        self.node_name_map: Dict[str, list] = {}     # name_lower -> [node_id, ...]
        self.G_full: nx.DiGraph = nx.DiGraph()       # 完整知识图谱有向图

    def load(self):
        print("正在加载数据...")
        d = self.data_dir

        with open(f"{d}/regulatory_visualization_data.json", 'r', encoding='utf-8') as f:
            self.regulatory_data = json.load(f)
        with open(f"{d}/responsibility_visualization_data.json", 'r', encoding='utf-8') as f:
            self.responsibility_data = json.load(f)
        with open(f"{d}/violation_visualization_data.json", 'r', encoding='utf-8') as f:
            self.violation_data = json.load(f)

        self.regulatory_reports = pd.read_excel(f"{d}/监管机构社区报告.xlsx")
        self.responsibility_reports = pd.read_excel(f"{d}/责任方社区报告.xlsx")
        self.violation_reports = pd.read_excel(f"{d}/违规行为社区报告.xlsx")
        self.community_hierarchy = pd.read_excel(f"{d}/community_hierarchy_v3_fixed.xlsx")

        with open(f"{d}/merged_regulatory_unified.txt", 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                data = json.loads(line)
                if data['type'] == 'node':
                    self.kg_nodes.append(data)
                    self.kg_node_map[data['id']] = data
                elif data['type'] == 'relationship':
                    self.kg_edges.append(data)

        self._build_kg_graph()
        self._build_name_index()
        print(f"数据加载完成！节点 {len(self.kg_nodes)} 个，关系 {len(self.kg_edges)} 条")

    def _build_kg_graph(self):
        """构建完整知识图谱有向图"""
        for node in self.kg_nodes:
            ntype = node['labels'][0] if node['labels'] else 'Unknown'
            self.G_full.add_node(
                node['id'],
                name=node['properties'].get('name', ''),
                node_type=ntype,
                labels=node['labels'],
                properties=node['properties']
            )
        for rel in self.kg_edges:
            self.G_full.add_edge(
                rel['start']['id'],
                rel['end']['id'],
                rel_id=rel['id'],
                rel_type=rel['label'],
                properties=rel.get('properties', {})
            )

    def _build_name_index(self):
        """构建节点名称到节点ID的索引（支持模糊匹配）"""
        for node in self.kg_nodes:
            name = node['properties'].get('name', '')
            if name:
                key = name.lower()
                if key not in self.node_name_map:
                    self.node_name_map[key] = []
                self.node_name_map[key].append(node['id'])


# ==================== 实体提取与社区匹配 ====================

class EntityMatcher:
    """实体提取 + 社区匹配器"""

    # 三层社区核心节点类型
    CORE_TYPES = {
        'responsibility': {'PartyWithResponsibility', 'AdvantageHolder', 'Actor'},
        'violation': {'Action', 'Means'},
        'regulatory': {'RegulatoryAuthority'}
    }

    def __init__(self, loader: DataLoader):
        self.loader = loader
        # 构建视角 -> {节点名_lower: {community_id, ...}}
        self.node_to_community: Dict[str, Dict[str, Set[int]]] = {
            'regulatory': {},
            'responsibility': {},
            'violation': {}
        }
        self._build_community_index()

    def _build_community_index(self):
        for perspective, vis_data in [
            ('regulatory', self.loader.regulatory_data),
            ('responsibility', self.loader.responsibility_data),
            ('violation', self.loader.violation_data),
        ]:
            for node in vis_data['nodes']:
                name_lower = node['name'].lower()
                comm_id = node['community']
                if name_lower not in self.node_to_community[perspective]:
                    self.node_to_community[perspective][name_lower] = set()
                self.node_to_community[perspective][name_lower].add(comm_id)

    def call_deepseek(self, prompt: str, max_tokens: int = 1000) -> str:
        headers = {
            "Authorization": f"Bearer {API_KEY}",
            "Content-Type": "application/json"
        }
        payload = {
            "model": "deepseek-chat",
            "messages": [{"role": "user", "content": prompt}],
            "max_tokens": max_tokens,
            "temperature": 0.2
        }
        resp = requests.post(f"{API_BASE}/chat/completions", headers=headers, json=payload, timeout=60)
        if resp.status_code == 200:
            return remove_markdown_formatting(resp.json()['choices'][0]['message']['content'])
        raise Exception(f"API错误 {resp.status_code}: {resp.text[:200]}")

    def extract_entities(self, event_desc: str) -> Dict[str, List[str]]:
        """调用 DeepSeek 从事件描述中提取三类实体"""
        print("\n[步骤1] 提取事件关键实体...")
        prompt = f"""请从以下资本市场事件描述中提取关键实体，分为三类：
责任方（上市公司、自然人、中介机构等）、违规行为（具体行为或手段）、监管机构（证监会、交易所等）。

事件描述：
{event_desc}

严格以 JSON 格式返回，不要包含其他内容：
{{
    "责任方": ["实体1", "实体2"],
    "违规行为": ["行为1", "行为2"],
    "监管机构": ["机构1"]
}}"""
        try:
            raw = self.call_deepseek(prompt)
            s = raw.find('{')
            e = raw.rfind('}') + 1
            entities = json.loads(raw[s:e])
        except Exception as ex:
            print(f"  实体提取失败，使用空结果: {ex}")
            entities = {"责任方": [], "违规行为": [], "监管机构": []}

        print(f"  责任方: {entities.get('责任方', [])}")
        print(f"  违规行为: {entities.get('违规行为', [])}")
        print(f"  监管机构: {entities.get('监管机构', [])}")
        return entities

    def match_communities(self, entities: Dict[str, List[str]],
                          event_desc: str) -> Dict[str, Set[int]]:
        """将实体映射到对应视角的社区"""
        print("\n[步骤2] 匹配相关社区...")
        matched: Dict[str, Set[int]] = {
            'responsibility': set(),
            'violation': set(),
            'regulatory': set()
        }

        def fuzzy_match(keyword: str, perspective: str):
            kl = keyword.lower()
            for name_lower, comms in self.node_to_community[perspective].items():
                if kl in name_lower or name_lower in kl:
                    matched[perspective].update(comms)

        for e in entities.get('责任方', []):
            fuzzy_match(e, 'responsibility')
        for e in entities.get('违规行为', []):
            fuzzy_match(e, 'violation')
        for e in entities.get('监管机构', []):
            fuzzy_match(e, 'regulatory')

        # 若匹配为空，退化到关键词扫描
        if not any(matched.values()):
            print("  初次匹配为空，使用关键词扫描...")
            words = re.split(r'[，。、；\s]+', event_desc)
            for w in words:
                if len(w) >= 2:
                    for p in ['responsibility', 'violation', 'regulatory']:
                        fuzzy_match(w, p)

        print(f"  责任方社区: {sorted(matched['responsibility'])}")
        print(f"  违规行为社区: {sorted(matched['violation'])}")
        print(f"  监管机构社区: {sorted(matched['regulatory'])}")
        return matched


# ==================== 节点级路径提取 ====================

class NodePathExtractor:
    """
    节点级穿透路径提取器

    逻辑：
    1. 从匹配到的社区中取出核心节点（三类视角各取其核心类型）
    2. 在完整知识图谱有向图上，用多跳 BFS/最短路径连接三层核心节点
    3. 返回两类路径：
       - 合规路径：责任方核心节点 → (责任/义务关系) → 监管核心节点
       - 违规路径：责任方核心节点 → (违规行为节点) → 监管核心节点
    """

    # 合规路径优先关系类型（从 KG 关系类型中选）
    COMPLIANCE_REL_TYPES = {'监管', '依照', '履行', '规定', '包含责任方', '依据', '处理', '处以'}
    # 违规路径优先关系类型
    VIOLATION_REL_TYPES = {'执行', '做出', '实施', '针对', '产生', '控制', '侵害', '受到处罚', '处理'}

    MAX_PATH_DEPTH = 5        # 最大跳数
    MAX_PATHS_PER_PAIR = 2    # 每对核心节点最多保留路径数
    MAX_CORE_NODES = 8        # 每层最多取核心节点数

    def __init__(self, loader: DataLoader):
        self.loader = loader
        # 构建 社区节点集合索引 {(perspective, community_id): [node_id, ...]}
        self.community_nodes: Dict[Tuple[str, int], List[str]] = {}
        self._build_community_node_index()

    def _build_community_node_index(self):
        for perspective, vis_data in [
            ('responsibility', self.loader.responsibility_data),
            ('violation', self.loader.violation_data),
            ('regulatory', self.loader.regulatory_data),
        ]:
            by_comm: Dict[int, List[str]] = defaultdict(list)
            for node in vis_data['nodes']:
                by_comm[node['community']].append(node['id'])
            for comm_id, node_ids in by_comm.items():
                self.community_nodes[(perspective, comm_id)] = node_ids

    def _get_core_nodes(self, perspective: str, community_ids: Set[int]) -> List[str]:
        """从匹配社区中提取核心类型节点，按加权度数排序取 top-N"""
        core_types = EntityMatcher.CORE_TYPES.get(perspective, set())
        candidates = []

        vis_data_map = {
            'responsibility': self.loader.responsibility_data,
            'violation': self.loader.violation_data,
            'regulatory': self.loader.regulatory_data,
        }
        vis_nodes = vis_data_map[perspective]['nodes']
        node_degree: Dict[str, float] = {n['id']: n.get('weightedDegree', 0) for n in vis_nodes}

        for comm_id in community_ids:
            for node_id in self.community_nodes.get((perspective, comm_id), []):
                node = self.loader.kg_node_map.get(node_id)
                if node:
                    ntype = node['labels'][0] if node['labels'] else ''
                    if ntype in core_types:
                        candidates.append((node_id, node_degree.get(node_id, 0)))

        # 按加权度排序，取前 N 个
        candidates.sort(key=lambda x: x[1], reverse=True)
        return [nid for nid, _ in candidates[:self.MAX_CORE_NODES]]

    def _find_paths_bfs(self, source_ids: List[str], target_ids: List[str],
                        preferred_rels: Set[str]) -> List[List[str]]:
        """
        在 KG 有向图中从 source 节点集合 BFS 到 target 节点集合
        返回节点ID路径列表
        """
        G = self.loader.G_full
        target_set = set(target_ids)
        found_paths: List[List[str]] = []

        for src in source_ids:
            if src not in G:
                continue
            # BFS: queue = [(当前节点, 当前路径)]
            queue: List[Tuple[str, List[str]]] = [(src, [src])]
            visited_in_path: Set[str] = set()

            while queue:
                cur, path = queue.pop(0)
                if len(path) > self.MAX_PATH_DEPTH:
                    continue
                if cur in target_set and len(path) > 1:
                    found_paths.append(path)
                    if len(found_paths) >= self.MAX_PATHS_PER_PAIR * len(source_ids):
                        return found_paths
                    continue

                for nxt in G.successors(cur):
                    if nxt in path:  # 避免环
                        continue
                    edge_data = G[cur][nxt]
                    rel_type = edge_data.get('rel_type', '')
                    # 优先走 preferred 关系类型，但也允许走其他类型
                    new_path = path + [nxt]
                    queue.append((nxt, new_path))

        # 去重并按路径长度排序
        seen = set()
        unique = []
        for p in sorted(found_paths, key=len):
            key = tuple(p)
            if key not in seen:
                seen.add(key)
                unique.append(p)
        return unique[:self.MAX_PATHS_PER_PAIR * max(len(source_ids), 1)]

    def extract_node_paths(self, matched_communities: Dict[str, Set[int]]) -> Dict[str, List]:
        """
        提取节点级穿透路径

        Returns:
            {
                'compliance': [   # 合规路径：责任方 → 监管
                    {
                        'nodes': [node_id, ...],
                        'edges': [(src, tgt, rel_type), ...],
                        'type': 'compliance'
                    }, ...
                ],
                'violation': [   # 违规路径：责任方 → 违规行为 → 监管
                    {...}, ...
                ]
            }
        """
        print("\n[步骤3] 提取节点级穿透路径...")

        resp_nodes = self._get_core_nodes('responsibility', matched_communities['responsibility'])
        viol_nodes = self._get_core_nodes('violation', matched_communities['violation'])
        reg_nodes = self._get_core_nodes('regulatory', matched_communities['regulatory'])

        print(f"  核心节点数 — 责任方:{len(resp_nodes)}, 违规行为:{len(viol_nodes)}, 监管:{len(reg_nodes)}")

        result = {'compliance': [], 'violation': []}

        G = self.loader.G_full

        def build_path_record(path_nodes: List[str], ptype: str) -> dict:
            edges = []
            for i in range(len(path_nodes) - 1):
                src, tgt = path_nodes[i], path_nodes[i + 1]
                edge_data = G.get_edge_data(src, tgt, default={})
                rel_type = edge_data.get('rel_type', '?')
                edges.append((src, tgt, rel_type))
            return {'nodes': path_nodes, 'edges': edges, 'type': ptype}

        # ---- 合规路径：责任方 → 监管 ----
        if resp_nodes and reg_nodes:
            paths = self._find_paths_bfs(resp_nodes, reg_nodes, self.COMPLIANCE_REL_TYPES)
            for p in paths:
                result['compliance'].append(build_path_record(p, 'compliance'))

        # ---- 违规路径：责任方 → 违规行为 → 监管 ----
        # 分两段：责任方→违规行为 + 违规行为→监管
        if resp_nodes and viol_nodes:
            seg1_paths = self._find_paths_bfs(resp_nodes, viol_nodes, self.VIOLATION_REL_TYPES)
        else:
            seg1_paths = []

        if viol_nodes and reg_nodes:
            seg2_paths = self._find_paths_bfs(viol_nodes, reg_nodes, self.COMPLIANCE_REL_TYPES)
        else:
            seg2_paths = []

        # 拼接两段路径（取 seg1 和 seg2 的笛卡尔积，限制数量）
        MAX_VIOL_PATHS = 4
        for p1 in seg1_paths:
            for p2 in seg2_paths:
                if p1[-1] == p2[0]:  # 衔接点是违规行为节点
                    combined = p1 + p2[1:]
                    result['violation'].append(build_path_record(combined, 'violation'))
                    if len(result['violation']) >= MAX_VIOL_PATHS:
                        break
            if len(result['violation']) >= MAX_VIOL_PATHS:
                break

        # 若两段拼接为空，退化：直接责任方→监管，中间包含违规节点
        if not result['violation'] and resp_nodes and reg_nodes:
            # 尝试经过 viol_nodes 的完整路径
            all_targets = set(reg_nodes)
            for src in resp_nodes:
                queue = [(src, [src])]
                while queue:
                    cur, path = queue.pop(0)
                    if len(path) > self.MAX_PATH_DEPTH:
                        continue
                    if cur in all_targets and len(path) > 1:
                        # 检查路径是否经过违规节点
                        if any(n in set(viol_nodes) for n in path[1:-1]):
                            result['violation'].append(build_path_record(path, 'violation'))
                        break
                    for nxt in G.successors(cur):
                        if nxt not in path:
                            queue.append((nxt, path + [nxt]))
                if len(result['violation']) >= MAX_VIOL_PATHS:
                    break

        print(f"  合规路径: {len(result['compliance'])} 条，违规路径: {len(result['violation'])} 条")
        return result


# ==================== 社区级路径构建（保留原逻辑）====================

def build_community_paths(matched_communities: Dict[str, Set[int]],
                          hierarchy_df: pd.DataFrame) -> List[Tuple]:
    """基于社区层级关系构建社区级路径（保留原逻辑）"""
    paths = []
    for _, row in hierarchy_df.iterrows():
        sp = row['source_perspective']
        si = row['source_community_id']
        tp = row['target_perspective']
        ti = row['target_community_id']
        rel_type = row['relation_type']
        is_strong = row['is_strong_link']

        if (si in matched_communities.get(sp, set()) and
                ti in matched_communities.get(tp, set())):
            paths.append((sp, si, tp, ti, rel_type, is_strong))
    return paths


# ==================== 可视化模块 ====================

class Visualizer:
    """双层可视化：社区级 + 节点级"""

    PERSPECTIVE_COLORS = {
        'responsibility': '#FF6B6B',
        'violation': '#FFA94D',
        'regulatory': '#4ECDC4',
    }
    PERSPECTIVE_NAMES_ZH = {
        'responsibility': '责任方',
        'violation': '违规行为',
        'regulatory': '监管机构',
    }
    PATH_COLORS = {
        'compliance': '#2196F3',   # 蓝色：合规路径
        'violation': '#E53935',    # 红色：违规路径
    }

    def __init__(self, loader: DataLoader):
        self.loader = loader

    # ---------- 社区级可视化（原逻辑保留）----------

    def visualize_community_network(self, matched_communities: Dict[str, Set[int]],
                                    community_paths: List[Tuple],
                                    output_file: str):
        """社区级网络可视化（保留原 event_network.png）"""
        G = nx.DiGraph()
        node_colors, node_sizes, node_labels = [], [], {}

        for perspective, comms in matched_communities.items():
            for cid in comms:
                nid = f"{perspective}_{cid}"
                G.add_node(nid)
                node_colors.append(self.PERSPECTIVE_COLORS[perspective])
                node_sizes.append(3000)
                node_labels[nid] = f"{self.PERSPECTIVE_NAMES_ZH[perspective]}\n社区#{cid}"

        for sp, si, tp, ti, rel_type, is_strong in community_paths:
            sn, tn = f"{sp}_{si}", f"{tp}_{ti}"
            if sn in G.nodes() and tn in G.nodes():
                G.add_edge(sn, tn, relation=rel_type, width=3 if is_strong else 1)

        fig, ax = plt.subplots(figsize=(16, 10))
        if len(G.nodes()) == 0:
            ax.text(0.5, 0.5, '未找到匹配社区', ha='center', va='center', fontsize=16)
        else:
            pos = nx.spring_layout(G, k=2.5, iterations=60, seed=42)
            nx.draw_networkx_nodes(G, pos, node_color=node_colors,
                                   node_size=node_sizes, alpha=0.9, ax=ax)
            edges = list(G.edges())
            widths = [G[u][v].get('width', 1) for u, v in edges]
            nx.draw_networkx_edges(G, pos, width=widths, alpha=0.6,
                                   arrows=True, arrowsize=20,
                                   arrowstyle='->', ax=ax)
            nx.draw_networkx_labels(G, pos, node_labels, font_size=9, ax=ax)

            legend_elements = [
                mpatches.Patch(color=self.PERSPECTIVE_COLORS['responsibility'],
                               label='责任方社区'),
                mpatches.Patch(color=self.PERSPECTIVE_COLORS['violation'],
                               label='违规行为社区'),
                mpatches.Patch(color=self.PERSPECTIVE_COLORS['regulatory'],
                               label='监管机构社区'),
            ]
            ax.legend(handles=legend_elements, loc='upper left', fontsize=10)

        ax.set_title("社区级责任递进路径", fontsize=15, pad=15)
        ax.axis('off')
        plt.tight_layout()
        plt.savefig(output_file, dpi=200, bbox_inches='tight', facecolor='white')
        plt.close()
        print(f"  ✓ 社区级可视化: {output_file}")

    # ---------- 节点级可视化 ----------

    def _node_display_name(self, node_id: str) -> str:
        node = self.loader.kg_node_map.get(node_id, {})
        name = node.get('properties', {}).get('name', '') if node else ''
        ntype = (node.get('labels', ['?'])[0] if node else '?')
        display = name[:18] if name else node_id[:12]
        return f"{display}\n[{ntype}]"

    def _node_color(self, node_id: str,
                    resp_ids: Set[str], viol_ids: Set[str], reg_ids: Set[str]) -> str:
        if node_id in resp_ids:
            return self.PERSPECTIVE_COLORS['responsibility']
        elif node_id in viol_ids:
            return self.PERSPECTIVE_COLORS['violation']
        elif node_id in reg_ids:
            return self.PERSPECTIVE_COLORS['regulatory']
        return '#BDBDBD'

    def visualize_node_paths(self, node_paths: Dict[str, List],
                             matched_communities: Dict[str, Set[int]],
                             output_file: str):
        """
        节点级穿透路径可视化
        左侧：合规路径（蓝色箭头）
        右侧：违规路径（红色箭头）
        """
        # 收集各层核心节点集合用于着色
        def collect_ids(perspective):
            ids = set()
            vis_map = {
                'responsibility': self.loader.responsibility_data,
                'violation': self.loader.violation_data,
                'regulatory': self.loader.regulatory_data,
            }
            core_types = EntityMatcher.CORE_TYPES.get(perspective, set())
            for n in vis_map[perspective]['nodes']:
                if n['community'] in matched_communities.get(perspective, set()):
                    node = self.loader.kg_node_map.get(n['id'], {})
                    ntype = (node.get('labels', [''])[0] if node else '')
                    if ntype in core_types:
                        ids.add(n['id'])
            return ids

        resp_ids = collect_ids('responsibility')
        viol_ids = collect_ids('violation')
        reg_ids = collect_ids('regulatory')

        compliance_paths = node_paths.get('compliance', [])
        violation_paths = node_paths.get('violation', [])

        n_comp = len(compliance_paths)
        n_viol = len(violation_paths)
        total = n_comp + n_viol

        if total == 0:
            fig, ax = plt.subplots(figsize=(10, 4))
            ax.text(0.5, 0.5, '未找到节点级穿透路径\n（可能是图谱连通性不足）',
                    ha='center', va='center', fontsize=14)
            ax.axis('off')
            plt.tight_layout()
            plt.savefig(output_file, dpi=150, bbox_inches='tight', facecolor='white')
            plt.close()
            print(f"  ✓ 节点级可视化（空）: {output_file}")
            return

        # 每条路径占一行，双列布局（合规/违规各一列）
        rows = max(n_comp, n_viol, 1)
        fig, axes = plt.subplots(rows, 2, figsize=(22, max(4 * rows, 6)))
        if rows == 1:
            axes = axes.reshape(1, 2)

        fig.suptitle("节点级穿透路径可视化", fontsize=16, fontweight='bold', y=1.01)

        def draw_path_on_ax(ax, path_record: Optional[dict], col_title: str, path_color: str):
            ax.axis('off')
            ax.set_title(col_title, fontsize=12, color=path_color, pad=8)
            if path_record is None:
                ax.text(0.5, 0.5, '（无路径）', ha='center', va='center',
                        fontsize=10, color='#999')
                return

            nodes_in_path = path_record['nodes']
            edges_in_path = path_record['edges']
            n = len(nodes_in_path)

            # 横向布局：等间距
            xs = [i / max(n - 1, 1) for i in range(n)]
            y = 0.5

            # 画边（箭头）
            for i, (src, tgt, rtype) in enumerate(edges_in_path):
                x1, x2 = xs[i], xs[i + 1]
                ax.annotate(
                    '', xy=(x2 - 0.03, y), xytext=(x1 + 0.03, y),
                    arrowprops=dict(arrowstyle='->', color=path_color,
                                   lw=2.0, connectionstyle='arc3,rad=0')
                )
                # 关系标签
                mid_x = (x1 + x2) / 2
                ax.text(mid_x, y + 0.12, rtype, ha='center', va='bottom',
                        fontsize=7.5, color=path_color,
                        bbox=dict(boxstyle='round,pad=0.2', facecolor='white',
                                  edgecolor=path_color, alpha=0.85))

            # 画节点
            for i, node_id in enumerate(nodes_in_path):
                x = xs[i]
                color = self._node_color(node_id, resp_ids, viol_ids, reg_ids)
                circle = plt.Circle((x, y), 0.06, color=color, zorder=5, alpha=0.9)
                ax.add_patch(circle)
                label = self._node_display_name(node_id)
                ax.text(x, y - 0.16, label, ha='center', va='top',
                        fontsize=7, wrap=True,
                        bbox=dict(boxstyle='round,pad=0.3', facecolor='white',
                                  edgecolor='#ccc', alpha=0.9))

            ax.set_xlim(-0.12, 1.12)
            ax.set_ylim(0.1, 0.85)

        for row_i in range(rows):
            # 左列：合规路径
            comp_rec = compliance_paths[row_i] if row_i < n_comp else None
            draw_path_on_ax(axes[row_i][0], comp_rec,
                            f"合规路径 #{row_i + 1}（责任方→监管）",
                            self.PATH_COLORS['compliance'])

            # 右列：违规路径
            viol_rec = violation_paths[row_i] if row_i < n_viol else None
            draw_path_on_ax(axes[row_i][1], viol_rec,
                            f"违规路径 #{row_i + 1}（责任方→违规行为→监管）",
                            self.PATH_COLORS['violation'])

        # 图例
        legend_handles = [
            mpatches.Patch(color=self.PERSPECTIVE_COLORS['responsibility'], label='责任方节点'),
            mpatches.Patch(color=self.PERSPECTIVE_COLORS['violation'], label='违规行为节点'),
            mpatches.Patch(color=self.PERSPECTIVE_COLORS['regulatory'], label='监管机构节点'),
            mpatches.Patch(color='#BDBDBD', label='中间节点'),
        ]
        fig.legend(handles=legend_handles, loc='lower center',
                   ncol=4, fontsize=10, bbox_to_anchor=(0.5, -0.02))

        plt.tight_layout()
        plt.savefig(output_file, dpi=200, bbox_inches='tight', facecolor='white')
        plt.close()
        print(f"  ✓ 节点级可视化: {output_file}")


# ==================== Prompt 构建（v2 升级）====================

class PromptBuilder:
    """
    构建 RAG Prompt

    v2 升级：同时注入
    1. 社区报告（原有）
    2. 路径节点的原始 KG 三元组（新增）→ 提供精细的法律依据和关系链
    """

    def __init__(self, loader: DataLoader):
        self.loader = loader

    def get_community_reports_text(self, matched_communities: Dict[str, Set[int]],
                                   report_dfs: Dict[str, pd.DataFrame]) -> str:
        """提取匹配社区的报告文本（原逻辑保留）"""
        parts = []
        for perspective, pname in [('responsibility', '责任方'),
                                    ('violation', '违规行为'),
                                    ('regulatory', '监管机构')]:
            df = report_dfs[perspective]
            comms = sorted(matched_communities.get(perspective, set()))
            if not comms:
                continue
            parts.append(f"\n{'='*50}")
            parts.append(f"【{pname}社区报告】")
            parts.append('='*50)
            for cid in comms:
                rows = df[df['community'] == cid]
                if rows.empty:
                    continue
                row = rows.iloc[0]
                parts.append(f"\n社区 #{cid}：{row.get('title', '')}")
                parts.append(f"摘要：{row.get('summary', '')}")
                parts.append(f"关键词：{row.get('key_words', '')}")
                # 取 findings 的 summary 字段
                try:
                    findings = json.loads(row.get('findings', '[]'))
                    for fi, f_item in enumerate(findings[:3], 1):
                        parts.append(f"  发现{fi}：{f_item.get('summary', '')}  —  {f_item.get('explanation', '')[:120]}")
                except Exception:
                    pass
        return '\n'.join(parts)

    def get_node_path_triples(self, node_paths: Dict[str, List]) -> str:
        """
        提取路径节点的 KG 原始三元组（新增）

        格式：
        [主体] -[关系]-> [客体]  |  主体类型: X, 客体类型: Y
        """
        G = self.loader.G_full
        kg_node_map = self.loader.kg_node_map
        seen_triples: Set[Tuple] = set()
        lines = []

        def node_info(nid: str) -> str:
            node = kg_node_map.get(nid, {})
            name = node.get('properties', {}).get('name', nid[:12]) if node else nid[:12]
            ntype = (node.get('labels', ['?'])[0] if node else '?')
            return f"{name}（{ntype}）"

        all_paths = node_paths.get('compliance', []) + node_paths.get('violation', [])
        for path_rec in all_paths:
            ptype = path_rec['type']
            ptype_zh = '合规' if ptype == 'compliance' else '违规'
            path_lines = [f"\n[{ptype_zh}路径]"]
            for src_id, tgt_id, rel_type in path_rec['edges']:
                triple_key = (src_id, tgt_id, rel_type)
                if triple_key in seen_triples:
                    continue
                seen_triples.add(triple_key)
                path_lines.append(
                    f"  {node_info(src_id)}  -[{rel_type}]->  {node_info(tgt_id)}"
                )
            lines.extend(path_lines)

        if not lines:
            return "（未提取到节点级三元组）"
        return '\n'.join(lines)

    def build_final_prompt(self, event_desc: str,
                           community_reports_text: str,
                           node_triples_text: str,
                           community_paths: List[Tuple],
                           node_paths: Dict[str, List]) -> str:
        """构建最终 Prompt（v2 升级版）"""

        # 社区级路径描述
        comm_path_lines = []
        for sp, si, tp, ti, rel_type, is_strong in community_paths:
            strength = '强关联' if is_strong else '弱关联'
            sp_zh = {'responsibility': '责任方', 'violation': '违规行为',
                     'regulatory': '监管机构'}.get(sp, sp)
            tp_zh = {'responsibility': '责任方', 'violation': '违规行为',
                     'regulatory': '监管机构'}.get(tp, tp)
            comm_path_lines.append(
                f"  {sp_zh}社区#{si} --[{rel_type}, {strength}]--> {tp_zh}社区#{ti}"
            )

        # 节点级路径描述（摘要）
        node_path_lines = []
        kg_node_map = self.loader.kg_node_map

        def short_name(nid):
            node = kg_node_map.get(nid, {})
            return node.get('properties', {}).get('name', nid[:10]) if node else nid[:10]

        for i, rec in enumerate(node_paths.get('compliance', [])[:2], 1):
            step = ' → '.join([short_name(n) for n in rec['nodes']])
            node_path_lines.append(f"  合规路径{i}: {step}")
        for i, rec in enumerate(node_paths.get('violation', [])[:2], 1):
            step = ' → '.join([short_name(n) for n in rec['nodes']])
            node_path_lines.append(f"  违规路径{i}: {step}")

        prompt = f"""你是资本市场法规合规领域的资深专家，请基于以下结构化知识对给定事件进行深度穿透式分析。

【事件描述】
{event_desc}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
一、社区级责任链路（宏观）
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{chr(10).join(comm_path_lines) if comm_path_lines else '（未找到社区级路径）'}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
二、节点级穿透路径（微观精细）
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{chr(10).join(node_path_lines) if node_path_lines else '（未找到节点级路径）'}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
三、知识图谱原始三元组（精细法律依据）
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{node_triples_text}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
四、相关社区报告（法规主题概述）
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{community_reports_text}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【分析要求】
请从以下五个维度提供专业分析，不使用 Markdown 格式：

1. 责任主体认定
   - 指明事件涉及的具体责任主体及其类型（如上市公司、实控人、中介机构）
   - 引用知识图谱三元组中的法律关系为依据

2. 违规行为定性
   - 结合违规路径和三元组，判断该事件属于哪类违规行为（信披违规/操纵市场/内幕交易等）
   - 援引具体法条/条款名称

3. 监管机构职责与处罚依据
   - 说明哪些监管机构有权介入，依据是哪条法规
   - 可能触发的处罚措施（警告/罚款/暂停/吊销等）

4. 合规义务梳理
   - 基于合规路径，列举责任主体在该场景下应当履行的合规义务

5. 风险提示与建议
   - 针对上述分析，给出具体可操作的合规建议
   - 提示监管趋势和潜在扩大风险"""

        return prompt


# ==================== 主查询系统 ====================

class RegulatoryQuerySystemV2:
    """监管违规穿透式查询系统 v2（节点级粒度）"""

    def __init__(self, data_dir: str = "data"):
        self.loader = DataLoader(data_dir)
        self.loader.load()

        self.matcher = EntityMatcher(self.loader)
        self.path_extractor = NodePathExtractor(self.loader)
        self.visualizer = Visualizer(self.loader)
        self.prompt_builder = PromptBuilder(self.loader)

        self.report_dfs = {
            'responsibility': self.loader.responsibility_reports,
            'violation': self.loader.violation_reports,
            'regulatory': self.loader.regulatory_reports,
        }

    def call_deepseek(self, prompt: str, max_tokens: int = 4000) -> str:
        headers = {"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"}
        payload = {
            "model": "deepseek-chat",
            "messages": [{"role": "user", "content": prompt}],
            "max_tokens": max_tokens,
            "temperature": 0.3
        }
        resp = requests.post(f"{API_BASE}/chat/completions", headers=headers,
                             json=payload, timeout=120)
        if resp.status_code == 200:
            return remove_markdown_formatting(resp.json()['choices'][0]['message']['content'])
        raise Exception(f"API 错误 {resp.status_code}: {resp.text[:200]}")

    def query(self, event_desc: str, output_dir: str = "v2_output") -> dict:
        """执行完整的 v2 穿透式查询"""
        os.makedirs(output_dir, exist_ok=True)
        print("=" * 70)
        print("  监管违规穿透式查询系统 v2（节点级粒度）")
        print("=" * 70)
        print(f"\n事件描述：\n{event_desc}\n")

        # Step 1 & 2：实体提取 + 社区匹配
        entities = self.matcher.extract_entities(event_desc)
        matched_communities = self.matcher.match_communities(entities, event_desc)

        # Step 3：节点级路径提取
        node_paths = self.path_extractor.extract_node_paths(matched_communities)

        # Step 4：社区级路径（原逻辑）
        community_paths = build_community_paths(matched_communities,
                                                self.loader.community_hierarchy)
        print(f"\n[步骤4] 社区级路径: {len(community_paths)} 条")

        # Step 5：双层可视化
        print("\n[步骤5] 生成可视化...")
        comm_vis_file = os.path.join(output_dir, "event_network.png")
        node_vis_file = os.path.join(output_dir, "node_level_path.png")
        self.visualizer.visualize_community_network(
            matched_communities, community_paths, comm_vis_file)
        self.visualizer.visualize_node_paths(
            node_paths, matched_communities, node_vis_file)

        # Step 6：构建 Prompt 并调用 LLM
        print("\n[步骤6] 构建 Prompt 并调用 DeepSeek...")
        community_reports_text = self.prompt_builder.get_community_reports_text(
            matched_communities, self.report_dfs)
        node_triples_text = self.prompt_builder.get_node_path_triples(node_paths)
        final_prompt = self.prompt_builder.build_final_prompt(
            event_desc, community_reports_text, node_triples_text,
            community_paths, node_paths)

        try:
            final_answer = self.call_deepseek(final_prompt, max_tokens=4000)
        except Exception as e:
            print(f"  LLM 调用失败: {e}")
            final_answer = f"（LLM 调用失败，请检查 API Key）\n\n原始 Prompt 已保存，请手动调用。\n\n错误信息：{e}"

        # Step 7：保存所有输出
        print("\n[步骤7] 保存输出文件...")
        self._save_outputs(output_dir, event_desc, entities, matched_communities,
                           community_paths, node_paths, node_triples_text,
                           community_reports_text, final_answer, final_prompt)

        print("\n" + "=" * 70)
        print("  查询完成！")
        print("=" * 70)

        return {
            'entities': entities,
            'matched_communities': matched_communities,
            'community_paths': community_paths,
            'node_paths': node_paths,
            'community_vis': comm_vis_file,
            'node_vis': node_vis_file,
            'final_answer': final_answer,
        }

    def _save_outputs(self, output_dir: str, event_desc: str,
                      entities: dict, matched_communities: dict,
                      community_paths: list, node_paths: dict,
                      node_triples_text: str, community_reports_text: str,
                      final_answer: str, final_prompt: str):
        kg_node_map = self.loader.kg_node_map

        def node_name(nid):
            n = kg_node_map.get(nid, {})
            return n.get('properties', {}).get('name', nid[:16]) if n else nid[:16]

        def node_type(nid):
            n = kg_node_map.get(nid, {})
            return (n.get('labels', ['?'])[0] if n else '?')

        # ---- 节点路径文本报告 ----
        path_report_lines = [
            "=" * 70,
            "节点级穿透路径报告",
            "=" * 70,
            f"\n事件描述：\n{event_desc}\n",
        ]

        path_report_lines.append("\n【合规路径（责任方 → 监管机构）】")
        for i, rec in enumerate(node_paths.get('compliance', []), 1):
            path_report_lines.append(f"\n  合规路径 #{i}：")
            for src, tgt, rt in rec['edges']:
                path_report_lines.append(
                    f"    {node_name(src)}（{node_type(src)}） -[{rt}]-> "
                    f"{node_name(tgt)}（{node_type(tgt)}）")

        path_report_lines.append("\n【违规路径（责任方 → 违规行为 → 监管机构）】")
        for i, rec in enumerate(node_paths.get('violation', []), 1):
            path_report_lines.append(f"\n  违规路径 #{i}：")
            for src, tgt, rt in rec['edges']:
                path_report_lines.append(
                    f"    {node_name(src)}（{node_type(src)}） -[{rt}]-> "
                    f"{node_name(tgt)}（{node_type(tgt)}）")

        with open(os.path.join(output_dir, "node_paths_report.txt"),
                  'w', encoding='utf-8') as f:
            f.write('\n'.join(path_report_lines))

        # ---- 社区报告 ----
        with open(os.path.join(output_dir, "community_reports.txt"),
                  'w', encoding='utf-8') as f:
            f.write(community_reports_text)

        # ---- KG 三元组 ----
        with open(os.path.join(output_dir, "node_triples.txt"),
                  'w', encoding='utf-8') as f:
            f.write(node_triples_text)

        # ---- 最终分析报告 ----
        with open(os.path.join(output_dir, "final_analysis.txt"),
                  'w', encoding='utf-8') as f:
            f.write("=" * 70 + "\n")
            f.write("监管违规事件综合分析报告（v2）\n")
            f.write("=" * 70 + "\n\n")
            f.write(f"【事件描述】\n{event_desc}\n\n")
            f.write(f"【分析结果】\n{final_answer}\n")

        # ---- Debug Prompt ----
        with open(os.path.join(output_dir, "debug_prompt.txt"),
                  'w', encoding='utf-8') as f:
            f.write(final_prompt)

        # ---- 汇总 Excel ----
        self._save_excel(output_dir, event_desc, entities, matched_communities,
                         community_paths, node_paths, final_answer)

        print(f"  ✓ 输出目录: {output_dir}")
        for fname in ['event_network.png', 'node_level_path.png',
                      'node_paths_report.txt', 'community_reports.txt',
                      'node_triples.txt', 'final_analysis.txt',
                      'analysis_summary_v2.xlsx']:
            fpath = os.path.join(output_dir, fname)
            if os.path.exists(fpath):
                print(f"    - {fname}")

    def _save_excel(self, output_dir: str, event_desc: str, entities: dict,
                    matched_communities: dict, community_paths: list,
                    node_paths: dict, final_answer: str):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        kg_node_map = self.loader.kg_node_map

        def node_name(nid):
            n = kg_node_map.get(nid, {})
            return n.get('properties', {}).get('name', nid[:16]) if n else nid[:16]

        def node_type(nid):
            n = kg_node_map.get(nid, {})
            return (n.get('labels', ['?'])[0] if n else '?')

        header_font = Font(bold=True, color="FFFFFF")
        header_fill_blue = PatternFill("solid", fgColor="1565C0")
        header_fill_red = PatternFill("solid", fgColor="B71C1C")
        header_fill_green = PatternFill("solid", fgColor="1B5E20")

        # Sheet 1: 事件概述
        ws1 = wb.active
        ws1.title = "事件概述"
        ws1['A1'] = "监管违规事件穿透分析报告 v2"
        ws1['A1'].font = Font(size=14, bold=True)
        ws1.append([])
        ws1.append(["事件描述", event_desc])
        ws1.append([])
        ws1.append(["提取实体", ""])
        ws1.append(["  责任方", str(entities.get('责任方', []))])
        ws1.append(["  违规行为", str(entities.get('违规行为', []))])
        ws1.append(["  监管机构", str(entities.get('监管机构', []))])
        ws1.append([])
        ws1.append(["最终分析结论", final_answer[:2000]])
        ws1.column_dimensions['A'].width = 20
        ws1.column_dimensions['B'].width = 100

        # Sheet 2: 社区级路径
        ws2 = wb.create_sheet("社区级路径")
        headers2 = ['序号', '源视角', '源社区', '目标视角', '目标社区', '关系类型', '关联强度']
        ws2.append(headers2)
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill_blue
        for i, (sp, si, tp, ti, rt, strong) in enumerate(community_paths, 1):
            pname = {'responsibility': '责任方', 'violation': '违规行为',
                     'regulatory': '监管机构'}
            ws2.append([i, pname.get(sp, sp), si, pname.get(tp, tp), ti, rt,
                        "强" if strong else "弱"])
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws2.column_dimensions[col].width = 16

        # Sheet 3: 节点级路径
        ws3 = wb.create_sheet("节点级路径")
        headers3 = ['路径类型', '路径序号', '步骤', '源节点ID', '源节点名称',
                    '源节点类型', '关系类型', '目标节点ID', '目标节点名称', '目标节点类型']
        ws3.append(headers3)
        for cell in ws3[1]:
            cell.font = header_font
            cell.fill = header_fill_red

        row_i = 1
        for ptype, ptype_zh in [('compliance', '合规路径'), ('violation', '违规路径')]:
            for path_no, rec in enumerate(node_paths.get(ptype, []), 1):
                for step_i, (src, tgt, rt) in enumerate(rec['edges'], 1):
                    ws3.append([
                        ptype_zh, path_no, step_i,
                        src, node_name(src), node_type(src),
                        rt,
                        tgt, node_name(tgt), node_type(tgt)
                    ])

        for col, width in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'],
                               [12, 10, 8, 30, 25, 20, 15, 30, 25, 20]):
            ws3.column_dimensions[col].width = width

        # Sheet 4: 匹配社区详情
        ws4 = wb.create_sheet("匹配社区详情")
        headers4 = ['视角', '社区ID', '社区标题', '摘要', '关键词']
        ws4.append(headers4)
        for cell in ws4[1]:
            cell.font = header_font
            cell.fill = header_fill_green

        report_map = {
            'responsibility': (self.loader.responsibility_reports, '责任方'),
            'violation': (self.loader.violation_reports, '违规行为'),
            'regulatory': (self.loader.regulatory_reports, '监管机构'),
        }
        for perspective, (df, pname) in report_map.items():
            for cid in sorted(matched_communities.get(perspective, set())):
                rows = df[df['community'] == cid]
                if not rows.empty:
                    row = rows.iloc[0]
                    ws4.append([pname, cid, row.get('title', ''),
                                row.get('summary', '')[:300],
                                row.get('key_words', '')])

        for col, width in zip(['A', 'B', 'C', 'D', 'E'],
                               [12, 10, 40, 80, 40]):
            ws4.column_dimensions[col].width = width

        excel_path = os.path.join(output_dir, "analysis_summary_v2.xlsx")
        wb.save(excel_path)


# ==================== 主入口 ====================

def main():
    print("初始化监管违规穿透式查询系统 v2...")

    # 数据目录（根据实际情况修改）
    system = RegulatoryQuerySystemV2(data_dir="data")

    print("\n" + "=" * 70)
    print("请输入事件描述（输入完成后按 Enter，再输入 END 结束）：")
    print("=" * 70)

    lines = []
    while True:
        line = input()
        if line.strip() == 'END':
            break
        lines.append(line)

    event_description = '\n'.join(lines).strip()
    if not event_description:
        print("错误：事件描述不能为空！")
        return

    results = system.query(event_description, output_dir="v2_output")

    print("\n" + "=" * 70)
    print("最终分析结果：")
    print("=" * 70)
    print(results['final_answer'])


if __name__ == "__main__":
    main()
