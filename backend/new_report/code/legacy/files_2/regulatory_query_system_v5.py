#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
监管违规穿透式查询系统 v5
================================================================
在 v4 基础上对【步骤5：节点级证据路径查找】进行全面升级。

核心升级内容（对应用户需求1/2/3）：

【升级1：搜索起点改为违规行为节点，强制三层约束】
  v4问题：从责任方出发，BFS 容易找到"责任方→监管机构"的通用捷径，
          违规行为节点反而绕道进入，导致路径缺乏事件针对性。
  v5修复：以违规行为节点为枢纽（Pivot）：
    - 反向束搜索：违规节点 ←← 责任方节点（沿反向边/无向边）
    - 正向束搜索：违规节点 →→ 监管机构节点（沿正向边）
    - 拼接：责任方→...→违规→...→监管机构
    - 强制约束：最终路径必须同时包含责任方、违规行为、监管机构三层节点

【升级2：用"违规中心双向语义束搜索"(VCBSBS)替代BFS】
  v4问题：BFS等权展开，语义无关节点与相关节点被平等扩展，
          搜索空间爆炸且结果质量低。
  v5算法：Violation-Centered Bidirectional Semantic Beam Search (VCBSBS)
    - 束(Beam)：每轮仅扩展语义评分最高的 Top-K 路径，大幅剪枝
    - 语义引导：扩展时用关键词匹配分实时评估每条候选路径的价值
    - 多样性去重：最终选路径时用节点集合重叠度保证结果多样性
    - 四级回退：有向束搜索 → 无向束搜索 → 直接回退搜索 → 原v4BFS兜底

【升级3：新版七维路径质量评分细则】
  v4问题：评分维度少（仅3项），且没有强制三层约束。
  v5评分（7个维度，总分约40+分）：
    维度1 三层覆盖完整性  (0/5/10分)：三层齐全10分，两层5分，一层0分
    维度2 关键词语义匹配密度 (上限20分)：精确匹配3分/词，子串匹配1分/词
    维度3 层次跳转质量  (上限8分)：责任方→违规→监管正序加分，逆序扣分
    维度4 节点特异性   (惩罚/奖励)：黑名单-3分/节点，名称>4字+0.5分
    维度5 路径紧凑性   (0-3分)：4-6跳最优3分，过短过长均递减
    维度6 边关系多样性  (0-2分)：不同关系类型数 × 0.5，上限2分
    维度7 法律条款覆盖  (0-3分)：包含Section/Chapter/Title节点加分
================================================================
"""

import json
import os
import re
import time
import warnings
from collections import defaultdict, deque
from typing import Dict, List, Set, Tuple, Optional, Any

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import networkx as nx
import numpy as np
import pandas as pd
import requests

warnings.filterwarnings('ignore')

import matplotlib.font_manager as fm
_FONT_PATH = "/root/miniconda3/lib/python3.8/site-packages/matplotlib/mpl-data/fonts/ttf/华文中宋.ttf"
try:
    fm.fontManager.addfont(_FONT_PATH)
    _font_prop = fm.FontProperties(fname=_FONT_PATH)
    plt.rcParams['font.sans-serif'] = [_font_prop.get_name(), 'sans-serif']
except Exception:
    plt.rcParams['font.sans-serif'] = ['SimHei', 'sans-serif']
plt.rcParams['axes.unicode_minus'] = False


# ==================== 配置 ====================
API_BASE   = "https://api.deepseek.com/v1"
API_KEY    = "sk-0a57f72b50854ace9d134a5eb697c4dc"
DATA_DIR   = "data"
OUTPUT_DIR = "v5_output"

# ---- 社区匹配参数 ----
COMMUNITY_CANDIDATE_TOPK  = 8
COMMUNITY_SCORE_THRESHOLD = 5
FAST_MODE = False

# ---- 节点锚点参数 ----
ANCHOR_CANDIDATE_SIZE = 20
ANCHOR_TOPK           = 8

# ---- v5 路径搜索参数（新增/修改）----
BEAM_WIDTH          = 15    # 束宽：每轮保留的最优路径数
MAX_SEGMENT_DEPTH   = 6     # 单段（违规→责任方 or 违规→监管）最大深度
MIN_TOTAL_HOPS      = 3     # 完整路径最少跳数
MAX_TOTAL_HOPS      = 10    # 完整路径最多跳数
COMPLIANCE_MIN_HOPS = 3     # 合规路径最少跳数
MAX_PATHS_PER_QUERY = 5     # 最终返回路径数
PATH_DIVERSITY_THRESHOLD = 0.7  # 路径相似度阈值（超过则认为重复）

# ---- 通用节点黑名单 ----
GENERIC_NODE_BLACKLIST = {
    '上市公司', '发行人', '本所', '本办法', '违反本办法',
    '公告', '报告', '审议通过', '相关规定', '有关规定',
}

# ---- 节点类型分层 ----
LAYER_TYPES = {
    'responsibility': {'PartyWithResponsibility', 'AdvantageHolder', 'Actor'},
    'violation':      {'Action', 'Means'},
    'regulatory':     {'RegulatoryAuthority'},
    'legal':          {'Section', 'Chapter', 'Title', 'Law'},
    'other':          {'Restriction', 'Event', 'Penalty', 'Obligation'},
}
TYPE_TO_LAYER: Dict[str, str] = {}
for _layer, _types in LAYER_TYPES.items():
    for _t in _types:
        TYPE_TO_LAYER[_t] = _layer

PERSPECTIVE_COLOR = {
    'responsibility': '#E53935',
    'violation':      '#FF8F00',
    'regulatory':     '#1E88E5',
    'legal':          '#43A047',
    'other':          '#8E24AA',
}
PERSPECTIVE_ZH = {
    'responsibility': '责任方',
    'violation':      '违规行为',
    'regulatory':     '监管机构',
    'legal':          '法律条款',
    'other':          '其他',
}
PATH_COLOR = {
    'compliance':         '#1565C0',
    'violation':          '#B71C1C',
    'violation_partial':  '#E65100',
}

# 层次顺序（用于跳转质量评分）
LAYER_ORDER = {
    'responsibility': 0,
    'violation':      1,
    'legal':          1.5,
    'other':          1.5,
    'regulatory':     2,
}


# ==================== 工具 ====================

def remove_markdown(text: str) -> str:
    if not text:
        return text
    text = re.sub(r'```[\w]*\n?', '', text)
    text = re.sub(r'```', '', text)
    text = re.sub(r'\*\*(.+?)\*\*', r'\1', text)
    text = re.sub(r'__(.+?)__', r'\1', text)
    text = re.sub(r'\*(.+?)\*', r'\1', text)
    text = re.sub(r'^#{1,6}\s+', '', text, flags=re.MULTILINE)
    text = re.sub(r'\[([^\]]+)\]\([^\)]+\)', r'\1', text)
    text = re.sub(r'`([^`]+)`', r'\1', text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def call_deepseek(prompt: str, max_tokens: int = 1500,
                  temperature: float = 0.1, retries: int = 3) -> str:
    headers = {"Authorization": f"Bearer {API_KEY}",
               "Content-Type": "application/json"}
    payload = {
        "model": "deepseek-chat",
        "messages": [{"role": "user", "content": prompt}],
        "max_tokens": max_tokens,
        "temperature": temperature,
    }
    for attempt in range(retries):
        try:
            resp = requests.post(f"{API_BASE}/chat/completions",
                                 headers=headers, json=payload, timeout=90)
            if resp.status_code == 200:
                return remove_markdown(
                    resp.json()['choices'][0]['message']['content'])
            print(f"    API 状态码: {resp.status_code}")
        except Exception as e:
            print(f"    API 调用异常 (尝试{attempt+1}): {e}")
        time.sleep(4)
    raise RuntimeError("DeepSeek API 调用失败")


def parse_json_safe(text: str) -> Any:
    for start_char, end_char in [('{', '}'), ('[', ']')]:
        s = text.find(start_char)
        e = text.rfind(end_char) + 1
        if s >= 0 and e > s:
            try:
                return json.loads(text[s:e])
            except Exception:
                pass
    return None


# ==================== 数据加载（与v4相同）====================

class DataLoader:
    def __init__(self, data_dir: str = DATA_DIR):
        self.data_dir      = data_dir
        self.vis:          Dict[str, dict]                  = {}
        self.reports:      Dict[str, pd.DataFrame]          = {}
        self.hierarchy:    pd.DataFrame                     = pd.DataFrame()
        self.kg_node_map:  Dict[str, dict]                  = {}
        self.kg_edges:     List[dict]                       = []
        self.G:            nx.DiGraph                       = nx.DiGraph()
        self.G_ud:         nx.Graph                         = nx.Graph()
        self.comm_nodes:   Dict[str, Dict[int, List[str]]]  = {}
        self.node_comm:    Dict[str, Dict[str, int]]        = {}
        self.comm_summary: Dict[str, Dict[int, str]]        = {}

    def load(self):
        d = self.data_dir
        print("正在加载数据...")
        for p in ['responsibility', 'violation', 'regulatory']:
            with open(f"{d}/{p}_visualization_data.json", 'r', encoding='utf-8') as f:
                self.vis[p] = json.load(f)

        for p, fname in [
            ('responsibility', '责任方社区报告.xlsx'),
            ('violation',      '违规行为社区报告.xlsx'),
            ('regulatory',     '监管机构社区报告.xlsx'),
        ]:
            self.reports[p] = pd.read_excel(f"{d}/{fname}")

        self.hierarchy = pd.read_excel(f"{d}/community_hierarchy_v3_fixed.xlsx")

        with open(f"{d}/merged_regulatory_unified.txt", 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                data = json.loads(line)
                if data['type'] == 'node':
                    self.kg_node_map[data['id']] = data
                elif data['type'] == 'relationship':
                    self.kg_edges.append(data)

        self._build_graph()
        self._build_community_index()
        self._build_summary_index()
        print(f"加载完成：{len(self.kg_node_map)} 节点  "
              f"{self.G.number_of_edges()} 有向边")

    def _build_graph(self):
        for node in self.kg_node_map.values():
            ntype = node['labels'][0] if node['labels'] else 'Unknown'
            name  = node['properties'].get('name', '')
            self.G.add_node(node['id'], name=name, node_type=ntype)
        for rel in self.kg_edges:
            sid = rel['start']['id']
            eid = rel['end']['id']
            if sid in self.G and eid in self.G:
                self.G.add_edge(sid, eid, rel_type=rel['label'])
        self.G_ud = self.G.to_undirected()

    def _build_community_index(self):
        for p in ['responsibility', 'violation', 'regulatory']:
            self.comm_nodes[p] = defaultdict(list)
            self.node_comm[p]  = {}
            for vnode in self.vis[p]['nodes']:
                nid = vnode['id']
                cid = int(vnode['community'])
                self.comm_nodes[p][cid].append(nid)
                self.node_comm[p][nid] = cid

    def _build_summary_index(self):
        for p in ['responsibility', 'violation', 'regulatory']:
            self.comm_summary[p] = {}
            df = self.reports[p]
            for _, row in df.iterrows():
                cid     = int(row['community'])
                title   = str(row.get('title', ''))
                summary = str(row.get('summary', ''))
                kw      = str(row.get('key_words', ''))
                self.comm_summary[p][cid] = f"{title}。{summary}。关键词：{kw}"

    def node_name(self, nid: str) -> str:
        n = self.kg_node_map.get(nid, {})
        return n.get('properties', {}).get('name', nid[:16]) if n else nid[:16]

    def node_type(self, nid: str) -> str:
        n = self.kg_node_map.get(nid, {})
        return (n.get('labels', ['?'])[0] if n else '?')

    def node_layer(self, nid: str) -> str:
        return TYPE_TO_LAYER.get(self.node_type(nid), 'other')

    def community_all_node_names(self, perspective: str, cid: int) -> List[str]:
        return [self.node_name(nid)
                for nid in self.comm_nodes[perspective].get(cid, [])]


# ==================== 精准实体提取（与v4相同）====================

class EntityExtractor:
    def __init__(self, loader: DataLoader):
        self.loader = loader

    def extract(self, event_desc: str) -> Dict[str, List[str]]:
        print("\n[步骤1] 提取结构化实体关键词...")
        prompt = f"""你是资本市场法规合规专家。请从下列事件中提取关键实体，只给出知识图谱中会出现的规范法律术语（不要给出人名/公司名等具体实体，要给出其法律分类名称）。

事件描述：
{event_desc}

返回严格 JSON（不含其他内容）：
{{
  "责任方":   ["法律意义上的责任主体类型术语1", "类型术语2"],
  "违规行为": ["违规行为类型或手段的规范术语1", "术语2"],
  "监管机构": ["有权监管此类事件的机构名称1", "名称2"],
  "核心法律概念": ["本事件涉及的核心法律概念1", "概念2", "概念3"]
}}

注意：
- 责任方举例："内幕信息知情人"、"持股5%%以上股东"、"董事、监事、高级管理人员"、"实际控制人"
- 违规行为举例："内幕交易"、"短线交易"、"操纵证券市场"、"虚假陈述"、"信息披露违规"
- 监管机构举例："中国证监会"、"证券交易所"、"派出机构"
- 核心法律概念举例："内幕信息"、"重大事项"、"公开前"、"利用"""
        try:
            raw    = call_deepseek(prompt, max_tokens=500)
            result = parse_json_safe(raw) or {}
        except Exception as ex:
            print(f"  提取失败: {ex}")
            result = {}

        all_kw: List[str] = []
        for key in ["责任方", "违规行为", "监管机构", "核心法律概念"]:
            all_kw.extend(result.get(key, []))
        result['_all'] = list(dict.fromkeys(all_kw))

        print(f"  责任方词:     {result.get('责任方', [])}")
        print(f"  违规行为词:   {result.get('违规行为', [])}")
        print(f"  监管机构词:   {result.get('监管机构', [])}")
        print(f"  核心法律概念: {result.get('核心法律概念', [])}")
        return result


# ==================== 精准社区匹配（与v4相同）====================

class PrecisionCommunityMatcher:
    HIGH_FREQ_BLACKLIST = {
        '证券', '公司', '规定', '法律', '法规', '本办法', '本规定',
        '依法', '依据', '应当', '不得', '行为', '处理', '相关',
        '违反', '违规', '合规', '管理', '监督', '机构',
    }

    def __init__(self, loader: DataLoader):
        self.loader = loader

    def _score_community(self, perspective: str, cid: int,
                         core_kws: List[str], general_kws: List[str]) -> float:
        summary    = self.loader.comm_summary[perspective].get(cid, '')
        node_names = self.loader.community_all_node_names(perspective, cid)
        full_text  = summary + ' ' + ' '.join(node_names)
        score = 0.0
        for kw in core_kws:
            if len(kw) < 2 or kw in self.HIGH_FREQ_BLACKLIST:
                continue
            if kw in full_text:
                score += 3.0
            else:
                for sub_len in range(len(kw), 2, -1):
                    for i in range(len(kw) - sub_len + 1):
                        sub = kw[i:i+sub_len]
                        if sub not in self.HIGH_FREQ_BLACKLIST and sub in full_text:
                            score += 0.8
                            break
                    else:
                        continue
                    break
        for kw in general_kws:
            if len(kw) < 2 or kw in self.HIGH_FREQ_BLACKLIST:
                continue
            if kw in full_text:
                score += 1.0
        return score

    def match_candidates(self, entities: Dict[str, List[str]],
                         topk: int = COMMUNITY_CANDIDATE_TOPK
                         ) -> Dict[str, List[Tuple[int, float]]]:
        kw_map = {
            'responsibility': {
                'core':    entities.get('责任方', []),
                'general': entities.get('核心法律概念', []),
            },
            'violation': {
                'core':    entities.get('违规行为', []),
                'general': entities.get('核心法律概念', []),
            },
            'regulatory': {
                'core':    entities.get('监管机构', []),
                'general': entities.get('违规行为', []),
            },
        }
        candidates: Dict[str, List[Tuple[int, float]]] = {}
        for p in ['responsibility', 'violation', 'regulatory']:
            scored = []
            for cid in self.loader.comm_nodes[p].keys():
                sc = self._score_community(p, cid, kw_map[p]['core'], kw_map[p]['general'])
                if sc > 0:
                    scored.append((cid, sc))
            scored.sort(key=lambda x: x[1], reverse=True)
            candidates[p] = scored[:topk]
        return candidates

    def llm_filter(self, candidates: Dict[str, List[Tuple[int, float]]],
                   event_desc: str,
                   threshold: float = COMMUNITY_SCORE_THRESHOLD
                   ) -> Dict[str, Set[int]]:
        matched: Dict[str, Set[int]] = {
            'responsibility': set(),
            'violation':      set(),
            'regulatory':     set(),
        }
        if FAST_MODE:
            for p, scored in candidates.items():
                matched[p] = {cid for cid, _ in scored}
            return matched

        pname = {'responsibility': '责任方',
                 'violation':      '违规行为',
                 'regulatory':     '监管机构'}
        for p, scored in candidates.items():
            if not scored:
                continue
            comm_info_lines = []
            for cid, pre_score in scored:
                summary    = self.loader.comm_summary[p].get(cid, '')[:150]
                node_names = self.loader.community_all_node_names(p, cid)[:8]
                comm_info_lines.append(
                    f"  社区#{cid}（预评分{pre_score:.1f}）：{summary}\n"
                    f"    代表节点：{', '.join(node_names)}")

            prompt = f"""你是资本市场法规专家。请评估以下每个{pname[p]}社区与事件的相关性（0-10分，10分最相关）。

【事件描述】
{event_desc}

【待评分的{pname[p]}社区】
{chr(10).join(comm_info_lines)}

评分标准：
- 9-10分：社区内容直接描述了事件涉及的{pname[p]}类型/行为/机构
- 7-8分：社区内容与事件存在明显关联
- 5-6分：社区内容与事件有一定关联
- 0-4分：社区内容与事件关联性弱

返回严格 JSON（社区编号->分数），不含其他内容：
{{"0": 7, "3": 9, "5": 2}}"""
            try:
                raw    = call_deepseek(prompt, max_tokens=300, temperature=0.1)
                scores = parse_json_safe(raw) or {}
                for cid_str, score in scores.items():
                    try:
                        cid   = int(cid_str)
                        score = float(score)
                        if score >= threshold:
                            matched[p].add(cid)
                    except (ValueError, TypeError):
                        pass
                print(f"  {pname[p]} LLM精筛分数: {scores}")
            except Exception as ex:
                print(f"  {pname[p]} LLM精筛失败: {ex}，使用词汇候选")
                matched[p] = {cid for cid, _ in scored}
        return matched

    def match(self, entities: Dict[str, List[str]],
              event_desc: str) -> Dict[str, Set[int]]:
        print("\n[步骤2] 精准社区匹配（词汇评分 + LLM精筛）...")
        candidates = self.match_candidates(entities)

        pname = {'responsibility': '责任方',
                 'violation':      '违规行为',
                 'regulatory':     '监管机构'}
        for p, scored in candidates.items():
            print(f"  {pname[p]} 词汇候选: "
                  f"{[(cid, round(sc,1)) for cid, sc in scored]}")

        matched = self.llm_filter(candidates, event_desc)

        print(f"  精筛后责任方社区:   {sorted(matched['responsibility'])}")
        print(f"  精筛后违规行为社区: {sorted(matched['violation'])}")
        print(f"  精筛后监管机构社区: {sorted(matched['regulatory'])}")

        matched = self._complement_missing(matched, candidates)
        return matched

    def _complement_missing(self,
                            matched:    Dict[str, Set[int]],
                            candidates: Dict[str, List[Tuple[int, float]]]
                            ) -> Dict[str, Set[int]]:
        h = self.loader.hierarchy

        def top_candidate(p: str) -> Optional[int]:
            lst = candidates.get(p, [])
            return lst[0][0] if lst else None

        for p in ['responsibility', 'violation', 'regulatory']:
            if matched[p]:
                continue
            other_ps = [op for op in ['responsibility', 'violation', 'regulatory']
                        if op != p and matched[op]]
            found: Set[int] = set()
            for op in other_ps:
                rows = h[(h['source_perspective'] == op) &
                         (h['source_community_id'].isin(matched[op])) &
                         (h['target_perspective'] == p)]
                found.update(rows['target_community_id'].tolist())
                rows = h[(h['target_perspective'] == op) &
                         (h['target_community_id'].isin(matched[op])) &
                         (h['source_perspective'] == p)]
                found.update(rows['source_community_id'].tolist())
            if found:
                matched[p] = found
                print(f"  → 层级推导补全 {p}: {sorted(found)}")
            else:
                top = top_candidate(p)
                if top is not None:
                    matched[p] = {top}
                    print(f"  → 词汇兜底补全 {p}: 社区#{top}")
        return matched


# ==================== LLM 精选锚点节点（与v4相同）====================

class LLMAnchorSelector:
    def __init__(self, loader: DataLoader):
        self.loader = loader

    def select(self, entities: Dict[str, List[str]],
               matched_communities: Dict[str, Set[int]],
               event_desc: str) -> Dict[str, List[str]]:
        print("\n[步骤3] LLM精选节点锚点...")
        anchors: Dict[str, List[str]] = {
            'responsibility': [],
            'violation':      [],
            'regulatory':     [],
        }
        for p in ['responsibility', 'violation', 'regulatory']:
            anchors[p] = self._select_for(p, matched_communities[p],
                                          entities, event_desc)
            names = [self.loader.node_name(n) for n in anchors[p][:6]]
            print(f"  {p} 锚点({len(anchors[p])}个): {names}")

        anchors = self._expand_empty(anchors)
        return anchors

    def _select_for(self, perspective: str, community_ids: Set[int],
                    entities: Dict[str, List[str]],
                    event_desc: str) -> List[str]:
        target_types = LAYER_TYPES[perspective]
        candidates: List[Tuple[str, str, float]] = []
        for cid in community_ids:
            for nid in self.loader.comm_nodes[perspective].get(cid, []):
                if self.loader.node_type(nid) not in target_types:
                    continue
                name = self.loader.node_name(nid)
                if name in GENERIC_NODE_BLACKLIST:
                    continue
                candidates.append((nid, name, self.loader.G.degree(nid)))

        if not candidates:
            return []

        candidates.sort(key=lambda x: x[2], reverse=True)
        top = candidates[:ANCHOR_CANDIDATE_SIZE]

        if FAST_MODE or len(top) <= 3:
            return [nid for nid, _, _ in top[:ANCHOR_TOPK]]

        pname = {'responsibility': '责任方',
                 'violation':      '违规行为',
                 'regulatory':     '监管机构'}
        candidate_lines = [f"  [{i}] {name}（频次:{int(deg)}）"
                           for i, (nid, name, deg) in enumerate(top)]
        prompt = f"""你是资本市场法规专家。下列是知识图谱中的{pname[perspective]}类型节点，请选出与事件最相关的节点编号（最多选{ANCHOR_TOPK}个）。

【事件描述】
{event_desc}

【候选节点列表】
{chr(10).join(candidate_lines)}

选择标准：节点名称与事件描述的责任主体/违规行为/监管场景直接相关。
请优先选择具体的、专指性强的节点，避免选"上市公司"、"发行人"等过于通用的节点（除非它确实是唯一选项）。

返回严格 JSON，只含选中的编号列表，不含其他内容：
[0, 3, 7]"""
        try:
            raw     = call_deepseek(prompt, max_tokens=200, temperature=0.1)
            indices = parse_json_safe(raw)
            if isinstance(indices, list):
                selected = [top[idx][0] for idx in indices
                            if isinstance(idx, int) and 0 <= idx < len(top)]
                if selected:
                    return selected[:ANCHOR_TOPK]
        except Exception as ex:
            print(f"    LLM锚点选取失败({perspective}): {ex}")

        return [nid for nid, _, _ in top[:ANCHOR_TOPK]]

    def _expand_empty(self, anchors: Dict[str, List[str]]) -> Dict[str, List[str]]:
        G = self.loader.G_ud
        for p in ['responsibility', 'violation', 'regulatory']:
            if anchors[p]:
                continue
            target_types = LAYER_TYPES[p]
            found: List[Tuple[str, float]] = []
            for op in ['responsibility', 'violation', 'regulatory']:
                if op == p:
                    continue
                for nid in anchors[op]:
                    for nbr in G.neighbors(nid):
                        if self.loader.node_type(nbr) in target_types:
                            name = self.loader.node_name(nbr)
                            if name not in GENERIC_NODE_BLACKLIST:
                                found.append((nbr, self.loader.G.degree(nbr)))
            found.sort(key=lambda x: x[1], reverse=True)
            anchors[p] = [nid for nid, _ in found[:ANCHOR_TOPK]]
            if anchors[p]:
                names = [self.loader.node_name(n) for n in anchors[p][:4]]
                print(f"  → {p} 邻居扩展锚点: {names}")
        return anchors


# ==================== v5 核心升级：违规中心双向语义束搜索 ====================

class PathQualityScorer:
    """
    七维路径质量评分器
    
    评分维度：
    1. 三层覆盖完整性  (0/5/10分)
    2. 关键词语义匹配密度  (每节点上限6分，总路径不设上限)
    3. 层次跳转质量  (正序加分，逆序扣分，标准序列额外奖励)
    4. 节点特异性  (黑名单扣分，名称长度奖励)
    5. 路径紧凑性  (最优4-6跳得3分，偏离递减)
    6. 边关系多样性  (不同关系类型数 × 0.5，上限2分)
    7. 法律条款覆盖  (每个法律节点+1.5分，上限3分)
    """

    def __init__(self, loader: DataLoader, event_keywords: List[str]):
        self.loader  = loader
        self.kw_set  = {kw.lower() for kw in event_keywords if len(kw) >= 2}
        self._cache: Dict[str, float] = {}

    def node_keyword_score(self, nid: str) -> float:
        """计算单节点与事件关键词的语义匹配分（带缓存）"""
        if nid in self._cache:
            return self._cache[nid]
        name = self.loader.node_name(nid).lower()
        score = 0.0
        for kw in self.kw_set:
            if len(kw) < 2:
                continue
            if kw in name:
                score += 3.0  # 精确包含：高分
            elif len(kw) >= 3:
                # 三字以上子串匹配：部分分
                matched = False
                for i in range(len(kw) - 2):
                    if kw[i:i+3] in name:
                        score += 1.0
                        matched = True
                        break
        # 节点名称特异性奖励：名称越长越具体
        raw_name = self.loader.node_name(nid)
        if len(raw_name) > 6:
            score += 0.8
        elif len(raw_name) > 4:
            score += 0.3
        # 黑名单惩罚
        if raw_name in GENERIC_NODE_BLACKLIST:
            score -= 4.0
        self._cache[nid] = score
        return score

    def compute(self, node_ids: List[str],
                edges: Optional[List[Tuple[str, str, str]]] = None) -> float:
        """
        计算完整路径质量分（七维综合）
        
        参数:
            node_ids: 路径节点ID列表
            edges: [(src_id, tgt_id, rel_type), ...] 可选，用于计算边多样性
        
        返回:
            综合质量分（浮点数，越高越好）
        """
        if len(node_ids) < 2:
            return 0.0

        layers = [self.loader.node_layer(nid) for nid in node_ids]
        unique_layers = set(layers)

        # ── 维度1：三层覆盖完整性 ──────────────────────────────
        has_resp = 'responsibility' in unique_layers
        has_viol = 'violation'      in unique_layers
        has_reg  = 'regulatory'     in unique_layers
        covered  = sum([has_resp, has_viol, has_reg])

        if covered == 3:
            dim1 = 10.0   # 三层完整：满分
        elif covered == 2:
            dim1 = 5.0    # 两层：半分
        else:
            dim1 = 0.0    # 单层：无分（路径无意义）

        # ── 维度2：关键词语义匹配密度 ──────────────────────────
        dim2 = 0.0
        for nid in node_ids:
            node_s = self.node_keyword_score(nid)
            # 每个节点贡献上限6分（避免单节点过度主导）
            dim2 += min(max(node_s, 0.0), 6.0)

        # ── 维度3：层次跳转质量 ────────────────────────────────
        dim3 = 0.0
        prev_order = LAYER_ORDER.get(layers[0], 1.5)
        for layer in layers[1:]:
            cur_order = LAYER_ORDER.get(layer, 1.5)
            if cur_order > prev_order:
                dim3 += 1.5   # 顺序递进：正向跳转
            elif abs(cur_order - prev_order) < 0.1:
                dim3 -= 0.5   # 同层横向：轻微扣分
            else:
                dim3 -= 1.0   # 逆序回退：扣分
            prev_order = cur_order

        # 标准三层顺序序列额外奖励
        key_layers = [l for l in layers
                      if l in {'responsibility', 'violation', 'regulatory'}]
        if key_layers == ['responsibility', 'violation', 'regulatory']:
            dim3 += 5.0   # 完美三层顺序
        elif (len(key_layers) >= 3
              and key_layers[0]  == 'responsibility'
              and 'violation'    in key_layers
              and key_layers[-1] == 'regulatory'):
            dim3 += 3.0   # 含三层但有中间穿插

        # ── 维度4：节点特异性 ──────────────────────────────────
        dim4 = 0.0
        for nid in node_ids:
            name = self.loader.node_name(nid)
            if name in GENERIC_NODE_BLACKLIST:
                dim4 -= 3.0  # 黑名单惩罚
            else:
                if len(name) > 6:
                    dim4 += 0.5  # 较长名称（具体性高）
                elif len(name) > 4:
                    dim4 += 0.2

        # ── 维度5：路径紧凑性 ──────────────────────────────────
        hops = len(node_ids) - 1
        if 4 <= hops <= 6:
            dim5 = 3.0    # 最优区间
        elif hops == 3:
            dim5 = 2.0    # 略短但可接受
        elif hops == 7:
            dim5 = 1.5
        elif hops == 8:
            dim5 = 0.5
        elif hops < 3:
            dim5 = -1.0   # 过短：缺乏语义
        else:
            dim5 = max(-2.0, 3.0 - (hops - 6) * 1.0)

        # ── 维度6：边关系多样性 ────────────────────────────────
        dim6 = 0.0
        if edges:
            rel_types = {e[2].replace('(↔)', '') for e in edges if e[2]}
            dim6 = min(len(rel_types) * 0.5, 2.0)

        # ── 维度7：法律条款覆盖 ────────────────────────────────
        legal_types = LAYER_TYPES['legal']
        legal_count = sum(1 for nid in node_ids
                          if self.loader.node_type(nid) in legal_types)
        dim6_legal = min(legal_count * 1.5, 3.0)

        total = dim1 + dim2 + dim3 + dim4 + dim5 + dim6 + dim6_legal

        return round(total, 2)

    def score_detail(self, node_ids: List[str],
                     edges: Optional[List[Tuple]] = None) -> Dict[str, float]:
        """返回各维度得分明细（调试用）"""
        layers = [self.loader.node_layer(nid) for nid in node_ids]
        unique_layers = set(layers)
        has_resp = 'responsibility' in unique_layers
        has_viol = 'violation'      in unique_layers
        has_reg  = 'regulatory'     in unique_layers
        covered  = sum([has_resp, has_viol, has_reg])

        dim1 = {3: 10.0, 2: 5.0}.get(covered, 0.0)
        dim2 = sum(min(max(self.node_keyword_score(n), 0.0), 6.0) for n in node_ids)

        dim3 = 0.0
        prev_order = LAYER_ORDER.get(layers[0], 1.5)
        for layer in layers[1:]:
            cur_order = LAYER_ORDER.get(layer, 1.5)
            dim3 += 1.5 if cur_order > prev_order else (-0.5 if abs(cur_order - prev_order) < 0.1 else -1.0)
            prev_order = cur_order
        key_layers = [l for l in layers if l in {'responsibility', 'violation', 'regulatory'}]
        if key_layers == ['responsibility', 'violation', 'regulatory']:
            dim3 += 5.0
        elif len(key_layers) >= 3 and key_layers[0] == 'responsibility' and key_layers[-1] == 'regulatory':
            dim3 += 3.0

        dim4 = sum(-3.0 if self.loader.node_name(n) in GENERIC_NODE_BLACKLIST
                   else (0.5 if len(self.loader.node_name(n)) > 6 else 0.2)
                   for n in node_ids)

        hops = len(node_ids) - 1
        dim5_map = {3: 2.0, 4: 3.0, 5: 3.0, 6: 3.0, 7: 1.5, 8: 0.5}
        dim5 = dim5_map.get(hops, -1.0 if hops < 3 else max(-2.0, 3.0 - (hops-6)))

        dim6 = min(len({e[2].replace('(↔)','') for e in edges if e[2]}) * 0.5, 2.0) if edges else 0.0
        legal_count = sum(1 for n in node_ids if self.loader.node_type(n) in LAYER_TYPES['legal'])
        dim7 = min(legal_count * 1.5, 3.0)

        return {
            '维度1_三层覆盖': round(dim1, 2),
            '维度2_关键词匹配': round(dim2, 2),
            '维度3_层次跳转': round(dim3, 2),
            '维度4_节点特异性': round(dim4, 2),
            '维度5_路径紧凑性': round(dim5, 2),
            '维度6_边关系多样性': round(dim6, 2),
            '维度7_法律条款覆盖': round(dim7, 2),
            '总分': round(dim1+dim2+dim3+dim4+dim5+dim6+dim7, 2),
        }


class VCBSBSPathFinder:
    """
    违规中心双向语义束搜索路径查找器
    Violation-Centered Bidirectional Semantic Beam Search (VCBSBS)

    算法流程：
    ┌─────────────────────────────────────────────────────────────┐
    │  违规行为锚点（Pivot）                                        │
    │       ↑ 反向束搜索（沿反向边）                               │
    │  责任方节点（segment_1_reversed）                            │
    │       ↓ 正向束搜索（沿正向边）                               │
    │  监管机构节点（segment_2）                                   │
    │                                                             │
    │  拼接：[责任方 → ... → 违规] + [违规 → ... → 监管机构]       │
    │  约束：最终路径必须包含三层节点                               │
    └─────────────────────────────────────────────────────────────┘

    束搜索核心思想：
    - 每一步仅扩展语义评分 Top-BEAM_WIDTH 的候选路径
    - 语义分 = 路径上所有节点的关键词匹配分之和
    - 大幅减少搜索空间，优先探索语义相关的路径方向

    四级回退策略：
    Level 1: 有向束搜索（反向/正向）
    Level 2: 无向图束搜索（图结构回退）
    Level 3: 直接端到端束搜索（含违规节点约束）
    Level 4: 原v4 BFS（最终兜底）
    """

    def __init__(self, loader: DataLoader, event_keywords: List[str]):
        self.loader  = loader
        self.G       = loader.G
        self.G_ud    = loader.G_ud
        self.scorer  = PathQualityScorer(loader, event_keywords)
        self.kw_set  = {kw.lower() for kw in event_keywords if len(kw) >= 2}

    # ── 内部：核心束搜索 ────────────────────────────────────────

    def _beam_search_directed(
            self,
            start_nodes:   List[str],
            target_types:  Set[str],
            max_depth:     int,
            beam_width:    int,
            reverse:       bool,
    ) -> List[Tuple[List[str], float]]:
        """
        有向束搜索。
        reverse=True  → 沿反向边（predecessor）搜索
        reverse=False → 沿正向边（successor）搜索

        返回: [(path, cumulative_score), ...] 其中 path[0] = start_node
        """
        # 初始束：每个起始节点一条单节点路径
        beams: List[Tuple[List[str], float]] = []
        for nid in start_nodes:
            if nid not in self.G:
                continue
            s = self.scorer.node_keyword_score(nid)
            beams.append(([nid], s))

        if not beams:
            return []

        beams.sort(key=lambda x: x[1], reverse=True)
        beams = beams[:beam_width]

        found: List[Tuple[List[str], float]] = []

        for _depth in range(max_depth):
            candidates: List[Tuple[List[str], float]] = []

            for path, path_score in beams:
                cur     = path[-1]
                visited = set(path)

                # 获取邻居（方向敏感）
                if reverse:
                    nbrs = list(self.G.predecessors(cur)) if cur in self.G else []
                else:
                    nbrs = list(self.G.successors(cur)) if cur in self.G else []

                for nbr in nbrs:
                    if nbr in visited:
                        continue
                    nbr_s    = self.scorer.node_keyword_score(nbr)
                    new_score = path_score + nbr_s
                    new_path  = path + [nbr]

                    nbr_type = self.loader.node_type(nbr)
                    if nbr_type in target_types:
                        if self.loader.node_name(nbr) not in GENERIC_NODE_BLACKLIST:
                            found.append((new_path, new_score))

                    candidates.append((new_path, new_score))

            # 剪枝：只保留 Top beam_width
            candidates.sort(key=lambda x: x[1], reverse=True)
            beams = candidates[:beam_width]
            if not beams:
                break

        return found

    def _beam_search_undirected(
            self,
            start_nodes:  List[str],
            target_types: Set[str],
            max_depth:    int,
            beam_width:   int,
    ) -> List[Tuple[List[str], float]]:
        """无向图束搜索（有向搜索失败时的回退）"""
        beams: List[Tuple[List[str], float]] = []
        for nid in start_nodes:
            if nid not in self.G_ud:
                continue
            s = self.scorer.node_keyword_score(nid)
            beams.append(([nid], s))

        if not beams:
            return []

        beams.sort(key=lambda x: x[1], reverse=True)
        beams = beams[:beam_width]
        found: List[Tuple[List[str], float]] = []

        for _depth in range(max_depth):
            candidates: List[Tuple[List[str], float]] = []
            for path, path_score in beams:
                cur     = path[-1]
                visited = set(path)
                nbrs    = list(self.G_ud.neighbors(cur)) if cur in self.G_ud else []
                for nbr in nbrs:
                    if nbr in visited:
                        continue
                    nbr_s    = self.scorer.node_keyword_score(nbr)
                    new_score = path_score + nbr_s
                    new_path  = path + [nbr]
                    if self.loader.node_type(nbr) in target_types:
                        if self.loader.node_name(nbr) not in GENERIC_NODE_BLACKLIST:
                            found.append((new_path, new_score))
                    candidates.append((new_path, new_score))
            candidates.sort(key=lambda x: x[1], reverse=True)
            beams = candidates[:beam_width]
            if not beams:
                break

        return found

    def _assemble_tri_layer_paths(
            self,
            resp_segs: List[Tuple[List[str], float]],
            reg_segs:  List[Tuple[List[str], float]],
    ) -> List[List[str]]:
        """
        路径拼接：将"违规→责任方"段（反转后变为"责任方→违规"）
        与"违规→监管机构"段拼接成完整三层路径。

        约束：
        - 两段共享同一个违规节点（path[0]）
        - 合并后无重复节点
        - 总跳数在 [MIN_TOTAL_HOPS, MAX_TOTAL_HOPS] 内
        """
        # 按违规节点分组
        seg1_by_viol: Dict[str, List[Tuple[List[str], float]]] = defaultdict(list)
        for path, sc in resp_segs:
            seg1_by_viol[path[0]].append((path, sc))

        seg2_by_viol: Dict[str, List[Tuple[List[str], float]]] = defaultdict(list)
        for path, sc in reg_segs:
            seg2_by_viol[path[0]].append((path, sc))

        common_pivots = set(seg1_by_viol.keys()) & set(seg2_by_viol.keys())
        if not common_pivots:
            return []

        assembled: List[List[str]] = []
        seen:       Set[Tuple]     = set()

        for pivot in common_pivots:
            for (seg1, _s1) in seg1_by_viol[pivot]:
                # 反转：[pivot, ..., resp] → [resp, ..., pivot]
                resp_to_viol = list(reversed(seg1))

                for (seg2, _s2) in seg2_by_viol[pivot]:
                    viol_to_reg = seg2  # [pivot, ..., reg]

                    # 拼接（去掉 seg2 首节点避免重复）
                    combined = resp_to_viol + viol_to_reg[1:]

                    # 基本检查
                    hops = len(combined) - 1
                    if hops < MIN_TOTAL_HOPS or hops > MAX_TOTAL_HOPS:
                        continue

                    # 无重复节点检查
                    if len(set(combined)) < len(combined):
                        continue

                    key = tuple(combined)
                    if key in seen:
                        continue
                    seen.add(key)

                    assembled.append(combined)

        return assembled

    def _direct_beam_fallback(
            self,
            resp_anchors: List[str],
            viol_anchors: List[str],
            reg_anchors:  List[str],
    ) -> List[List[str]]:
        """
        Level-3 回退：从责任方出发端到端束搜索监管机构，
        要求路径中间必须经过违规行为节点。
        """
        viol_set  = set(viol_anchors)
        reg_types = LAYER_TYPES['regulatory']
        found:    List[List[str]] = []
        seen:     Set[Tuple]      = set()

        beams: List[Tuple[List[str], float]] = [
            ([nid], self.scorer.node_keyword_score(nid))
            for nid in resp_anchors if nid in self.G_ud
        ]
        beams.sort(key=lambda x: x[1], reverse=True)
        beams = beams[:BEAM_WIDTH]

        max_depth = MAX_SEGMENT_DEPTH * 2

        for _depth in range(max_depth):
            candidates: List[Tuple[List[str], float]] = []
            for path, score in beams:
                cur     = path[-1]
                visited = set(path)
                nbrs    = (list(self.G.successors(cur)) if cur in self.G else []) \
                          or (list(self.G_ud.neighbors(cur)) if cur in self.G_ud else [])
                for nbr in nbrs:
                    if nbr in visited:
                        continue
                    nbr_s    = self.scorer.node_keyword_score(nbr)
                    new_path  = path + [nbr]
                    new_score = score + nbr_s
                    hops      = len(new_path) - 1
                    if (self.loader.node_type(nbr) in reg_types
                            and hops >= MIN_TOTAL_HOPS
                            and self.loader.node_name(nbr) not in GENERIC_NODE_BLACKLIST
                            and any(n in viol_set for n in new_path[1:-1])):
                        key = tuple(new_path)
                        if key not in seen and len(set(new_path)) == len(new_path):
                            seen.add(key)
                            found.append(new_path)
                    if hops < max_depth:
                        candidates.append((new_path, new_score))
            candidates.sort(key=lambda x: x[1], reverse=True)
            beams = candidates[:BEAM_WIDTH]
            if not beams:
                break

        return found

    def _v4_bfs_fallback(
            self,
            resp_anchors: List[str],
            reg_anchors:  List[str],
            min_hops:     int = MIN_TOTAL_HOPS,
    ) -> List[List[str]]:
        """
        Level-4 终极回退：原v4 BFS（仅在所有束搜索均失败时使用）
        """
        tgt_set = set(reg_anchors)
        found:   List[List[str]] = []
        seen:    Set[Tuple]      = set()

        for src in resp_anchors[:ANCHOR_TOPK]:
            if src not in self.G_ud:
                continue
            queue = deque([(src, [src], {src})])
            while queue:
                cur, path, visited = queue.popleft()
                hops = len(path) - 1
                if cur in tgt_set and hops >= min_hops:
                    key = tuple(path)
                    if key not in seen:
                        seen.add(key)
                        found.append(path)
                    continue
                if hops >= BFS_MAX_DEPTH:
                    continue
                for nxt in list(self.G_ud.neighbors(cur)):
                    if nxt not in visited:
                        queue.append((nxt, path + [nxt], visited | {nxt}))
        return found

    # ── 内部：多样性最优路径选择 ─────────────────────────────────

    def _select_diverse_best(
            self,
            paths: List[List[str]],
            n:     int,
    ) -> List[List[str]]:
        """
        选择 n 条质量最高且互相多样化的路径。
        多样性度量：两条路径的节点集合 Jaccard 相似度 < PATH_DIVERSITY_THRESHOLD
        才视为足够不同。
        """
        if not paths:
            return []

        # 按质量分降序排列
        scored = [(p, self.scorer.compute(p)) for p in paths]
        scored.sort(key=lambda x: x[1], reverse=True)

        selected:        List[List[str]]   = []
        selected_sets:   List[Set[str]]    = []
        seen_keys:       Set[Tuple]        = set()

        for path, score in scored:
            key = tuple(path)
            if key in seen_keys:
                continue
            seen_keys.add(key)

            path_set = set(path)

            # 多样性检查
            too_similar = False
            for existing_set in selected_sets:
                union_size = len(path_set | existing_set)
                if union_size == 0:
                    continue
                jaccard = len(path_set & existing_set) / union_size
                if jaccard >= PATH_DIVERSITY_THRESHOLD:
                    too_similar = True
                    break

            if not too_similar or len(selected) == 0:
                selected.append(path)
                selected_sets.append(path_set)

            if len(selected) >= n:
                break

        # 若多样性约束太严导致数量不足，降级补充
        if len(selected) < n:
            for path, _ in scored:
                if path not in selected:
                    selected.append(path)
                if len(selected) >= n:
                    break

        return selected[:n]

    # ── 内部：路径记录构建 ───────────────────────────────────────

    def _build_record(self, node_ids: List[str], ptype: str) -> dict:
        """构建路径记录，包含节点、边、质量分及各维度明细"""
        edges: List[Tuple[str, str, str]] = []
        for i in range(len(node_ids) - 1):
            src, tgt = node_ids[i], node_ids[i+1]
            data = self.G.get_edge_data(src, tgt)
            if data is None:
                data = self.G.get_edge_data(tgt, src)
                rt   = (data or {}).get('rel_type', '关联')
                edges.append((src, tgt, f"{rt}(↔)"))
            else:
                edges.append((src, tgt, data.get('rel_type', '关联')))

        score   = self.scorer.compute(node_ids, edges)
        detail  = self.scorer.score_detail(node_ids, edges)
        layers  = list({self.loader.node_layer(n) for n in node_ids})

        return {
            'nodes':         node_ids,
            'edges':         edges,
            'type':          ptype,
            'score':         score,
            'score_detail':  detail,
            'layer_coverage': layers,
        }

    # ── 公开接口 ─────────────────────────────────────────────────

    def find_violation_paths(
            self,
            resp_anchors: List[str],
            viol_anchors: List[str],
            reg_anchors:  List[str],
    ) -> List[dict]:
        """
        违规路径搜索（核心方法）

        以违规节点为枢纽，双向束搜索：
          反向 → 责任方
          正向 → 监管机构
        强制三层约束，四级回退保证鲁棒性。
        """
        print("    [VCBSBS] 启动违规路径搜索...")
        resp_types = LAYER_TYPES['responsibility']
        reg_types  = LAYER_TYPES['regulatory']

        # Level 1a: 有向反向束搜索（违规 ←← 责任方）
        print("    [Level 1] 有向束搜索：违规←责任方 / 违规→监管...")
        resp_segs = self._beam_search_directed(
            viol_anchors, resp_types,
            MAX_SEGMENT_DEPTH, BEAM_WIDTH, reverse=True)

        # Level 1b: 有向正向束搜索（违规 →→ 监管机构）
        reg_segs = self._beam_search_directed(
            viol_anchors, reg_types,
            MAX_SEGMENT_DEPTH, BEAM_WIDTH, reverse=False)

        # Level 2: 有向失败 → 无向束搜索回退
        if not resp_segs:
            print("    [Level 2] 责任方段回退：无向束搜索...")
            resp_segs = self._beam_search_undirected(
                viol_anchors, resp_types, MAX_SEGMENT_DEPTH + 2, BEAM_WIDTH)
        if not reg_segs:
            print("    [Level 2] 监管机构段回退：无向束搜索...")
            reg_segs = self._beam_search_undirected(
                viol_anchors, reg_types, MAX_SEGMENT_DEPTH + 2, BEAM_WIDTH)

        print(f"    责任方段: {len(resp_segs)} 条，监管机构段: {len(reg_segs)} 条")

        # 路径拼接（三层约束）
        assembled = self._assemble_tri_layer_paths(resp_segs, reg_segs)
        print(f"    三层拼接路径: {len(assembled)} 条")

        # Level 3: 拼接失败 → 端到端束搜索（含违规节点约束）
        if not assembled:
            print("    [Level 3] 回退：端到端束搜索（要求经过违规节点）...")
            assembled = self._direct_beam_fallback(
                resp_anchors, viol_anchors, reg_anchors)
            print(f"    端到端路径: {len(assembled)} 条")

        # Level 4: 终极 BFS 兜底
        if not assembled:
            print("    [Level 4] 终极回退：v4 BFS...")
            assembled = self._v4_bfs_fallback(resp_anchors, reg_anchors)

        best = self._select_diverse_best(assembled, MAX_PATHS_PER_QUERY)
        result = [self._build_record(p, 'violation') for p in best]

        # 如果完全没有三层路径，降级返回半段
        if not result:
            print("    → 未找到三层路径，返回责任方→违规半段...")
            partial_paths = [list(reversed(path))
                             for path, _ in resp_segs[:MAX_PATHS_PER_QUERY]]
            result = [self._build_record(p, 'violation_partial')
                      for p in partial_paths[:MAX_PATHS_PER_QUERY]]

        return result

    def find_compliance_paths(
            self,
            resp_anchors: List[str],
            reg_anchors:  List[str],
    ) -> List[dict]:
        """
        合规路径搜索：责任方 → 监管机构（经过法律/义务节点）
        使用单向语义束搜索，仍允许途经违规或法律节点。
        """
        print("    [VCBSBS] 启动合规路径搜索...")
        reg_types = LAYER_TYPES['regulatory']

        # Level 1: 有向正向束搜索
        print("    [Level 1] 有向束搜索：责任方→监管机构...")
        segs = self._beam_search_directed(
            resp_anchors, reg_types,
            MAX_SEGMENT_DEPTH + 2, BEAM_WIDTH, reverse=False)
        paths = [path for path, _ in segs
                 if len(path) - 1 >= COMPLIANCE_MIN_HOPS]

        # Level 2: 无向回退
        if not paths:
            print("    [Level 2] 合规路径回退：无向束搜索...")
            segs = self._beam_search_undirected(
                resp_anchors, reg_types,
                MAX_SEGMENT_DEPTH + 3, BEAM_WIDTH)
            paths = [path for path, _ in segs
                     if len(path) - 1 >= COMPLIANCE_MIN_HOPS]

        # Level 3: BFS 兜底
        if not paths:
            print("    [Level 3] 合规路径终极回退：BFS...")
            paths = self._v4_bfs_fallback(resp_anchors, reg_anchors,
                                          min_hops=COMPLIANCE_MIN_HOPS)

        best = self._select_diverse_best(paths, MAX_PATHS_PER_QUERY)
        return [self._build_record(p, 'compliance') for p in best]


# ==================== 可视化（与v4相同，新增分数明细展示）====================

BFS_MAX_DEPTH = 4   # BFS兜底参数（与v4一致）


def visualize_community_network(loader: DataLoader,
                                matched: Dict[str, Set[int]],
                                comm_paths: List[Tuple],
                                out_file: str):
    G = nx.DiGraph()
    colors, sizes, labels = [], [], {}
    for p, comms in matched.items():
        for cid in comms:
            nid = f"{p}_{cid}"
            G.add_node(nid)
            colors.append(PERSPECTIVE_COLOR.get(p, '#999'))
            sizes.append(3200)
            labels[nid] = f"{PERSPECTIVE_ZH.get(p,p)}\n社区#{cid}"
    for sp, si, tp, ti, rt, strong in comm_paths:
        sn, tn = f"{sp}_{si}", f"{tp}_{ti}"
        if sn in G and tn in G:
            G.add_edge(sn, tn, width=3 if strong else 1)

    fig, ax = plt.subplots(figsize=(16, 10))
    if len(G) == 0:
        ax.text(0.5, 0.5, '未找到匹配社区', ha='center', va='center', fontsize=16)
    else:
        pos    = nx.spring_layout(G, k=2.5, iterations=80, seed=42)
        nx.draw_networkx_nodes(G, pos, node_color=colors,
                               node_size=sizes, alpha=0.88, ax=ax)
        es     = list(G.edges())
        widths = [G[u][v].get('width', 1) for u, v in es]
        nx.draw_networkx_edges(G, pos, width=widths, alpha=0.6,
                               arrows=True, arrowsize=22, ax=ax)
        nx.draw_networkx_labels(G, pos, labels, font_size=9, ax=ax)
        legend_handles = [
            mpatches.Patch(color=PERSPECTIVE_COLOR[p], label=PERSPECTIVE_ZH[p])
            for p in ['responsibility', 'violation', 'regulatory']
        ]
        ax.legend(handles=legend_handles, loc='upper left', fontsize=10)
    ax.set_title("社区级责任递进路径（v5）", fontsize=14, pad=12)
    ax.axis('off')
    plt.tight_layout()
    plt.savefig(out_file, dpi=180, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f"  ✓ 社区级: {out_file}")


def visualize_node_paths(loader: DataLoader,
                         node_paths: Dict[str, List],
                         out_file: str):
    comp_paths = node_paths.get('compliance', [])
    viol_paths = (node_paths.get('violation', []) +
                  node_paths.get('violation_partial', []))
    n_rows = max(len(comp_paths), len(viol_paths), 1)

    fig, axes = plt.subplots(n_rows, 2,
                              figsize=(28, max(5.5 * n_rows, 5.5)))
    if n_rows == 1:
        axes = axes.reshape(1, 2)
    fig.suptitle(
        "节点级穿透路径可视化（v5）\n"
        "蓝=合规路径（责任方→监管）  红=违规路径（责任方→违规→监管）",
        fontsize=14, fontweight='bold', y=1.01)

    def get_color(nid: str) -> str:
        return PERSPECTIVE_COLOR.get(loader.node_layer(nid), '#9E9E9E')

    def draw(ax, rec: Optional[dict], title: str, arrow_color: str):
        ax.axis('off')
        ax.set_title(title, fontsize=9.5, color=arrow_color,
                     pad=6, fontweight='bold')
        if rec is None:
            ax.text(0.5, 0.5, '（无路径）', ha='center', va='center',
                    fontsize=10, color='#BDBDBD')
            return
        nodes = rec['nodes']
        edges = rec['edges']
        n     = len(nodes)
        if n == 0:
            return
        xs     = np.linspace(0.06, 0.94, n)
        y_node = 0.62
        y_lbl  = 0.24

        for i, (s, t, rt) in enumerate(edges):
            ax.annotate('', xy=(xs[i+1]-0.025, y_node),
                        xytext=(xs[i]+0.025, y_node),
                        arrowprops=dict(arrowstyle='->', color=arrow_color,
                                        lw=2.2, mutation_scale=18))
            mid_x = (xs[i] + xs[i+1]) / 2
            ax.text(mid_x, y_node+0.14, rt, ha='center', va='center',
                    fontsize=7.0, color=arrow_color,
                    bbox=dict(boxstyle='round,pad=0.25', facecolor='white',
                              edgecolor=arrow_color, alpha=0.9, linewidth=1.2))

        for i, nid in enumerate(nodes):
            x     = xs[i]
            color = get_color(nid)
            layer = loader.node_layer(nid)
            circle = plt.Circle((x, y_node), 0.045,
                                  color=color, zorder=5, alpha=0.92)
            ax.add_patch(circle)
            ax.text(x, y_node+0.005, PERSPECTIVE_ZH.get(layer, '?')[:3],
                    ha='center', va='center', fontsize=6.5,
                    color='white', fontweight='bold', zorder=6)
            name   = loader.node_name(nid)
            name_d = name[:14] + '..' if len(name) > 14 else name
            ax.text(x, y_lbl,
                    f"{name_d}\n({loader.node_type(nid)})",
                    ha='center', va='top', fontsize=7.0,
                    bbox=dict(boxstyle='round,pad=0.3', facecolor='#FAFAFA',
                              edgecolor='#CCCCCC', alpha=0.95))

        # 显示七维分数明细
        detail = rec.get('score_detail', {})
        if detail:
            dim_text = (
                f"三层:{detail.get('维度1_三层覆盖',0):.1f} "
                f"关键词:{detail.get('维度2_关键词匹配',0):.1f} "
                f"跳转:{detail.get('维度3_层次跳转',0):.1f} "
                f"特异:{detail.get('维度4_节点特异性',0):.1f} "
                f"紧凑:{detail.get('维度5_路径紧凑性',0):.1f} "
                f"法律:{detail.get('维度7_法律条款覆盖',0):.1f}"
            )
            ax.text(0.5, 0.08, dim_text, ha='center', va='center',
                    fontsize=6.5, color='#555555',
                    bbox=dict(boxstyle='round,pad=0.3', facecolor='#F5F5F5',
                              edgecolor='#DDDDDD', alpha=0.9))

        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)

    ptype_zh = {'violation': '违规', 'violation_partial': '违规(半段)'}
    for row_i in range(n_rows):
        cr = comp_paths[row_i] if row_i < len(comp_paths) else None
        vr = viol_paths[row_i] if row_i < len(viol_paths) else None

        layer_tag_c = (f"[{'/'.join(cr.get('layer_coverage',[])[:3])}]"
                       if cr else "")
        layer_tag_v = (f"[{'/'.join(vr.get('layer_coverage',[])[:3])}]"
                       if vr else "")

        ct = (f"合规路径#{row_i+1}（{len(cr['nodes'])}节点  "
              f"质量分:{cr['score']}）{layer_tag_c}"
              if cr else f"合规路径#{row_i+1}")
        vt_pfx = ptype_zh.get((vr or {}).get('type', ''), '违规')
        vt = (f"{vt_pfx}路径#{row_i+1}（{len(vr['nodes'])}节点  "
              f"质量分:{vr['score']}）{layer_tag_v}"
              if vr else f"违规路径#{row_i+1}")

        draw(axes[row_i][0], cr, ct, PATH_COLOR['compliance'])
        draw(axes[row_i][1], vr, vt,
             PATH_COLOR.get((vr or {}).get('type', 'violation'),
                            PATH_COLOR['violation']))

    legend_handles = [
        mpatches.Patch(color=PERSPECTIVE_COLOR[p],
                       label=f"{PERSPECTIVE_ZH[p]}节点")
        for p in ['responsibility', 'violation', 'regulatory', 'legal', 'other']
    ]
    fig.legend(handles=legend_handles, loc='lower center', ncol=5,
               fontsize=9, bbox_to_anchor=(0.5, -0.03))
    plt.tight_layout()
    plt.savefig(out_file, dpi=180, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f"  ✓ 节点级: {out_file}")


# ==================== 社区级路径（与v4相同）====================

def build_community_paths(matched: Dict[str, Set[int]],
                          hierarchy: pd.DataFrame) -> List[Tuple]:
    paths = []
    for _, row in hierarchy.iterrows():
        sp     = row['source_perspective']
        si     = int(row['source_community_id'])
        tp     = row['target_perspective']
        ti     = int(row['target_community_id'])
        rt     = row['relation_type']
        strong = bool(row['is_strong_link'])
        if (si in matched.get(sp, set()) and
                ti in matched.get(tp, set())):
            paths.append((sp, si, tp, ti, rt, strong))
    return paths


# ==================== Prompt + 文本构建（与v4相同）====================

def build_prompt(event_desc: str, community_text: str,
                 triples_text: str, comm_paths: List[Tuple],
                 node_paths: Dict[str, List],
                 loader: DataLoader) -> str:
    pzh = {'responsibility': '责任方',
           'violation':      '违规行为',
           'regulatory':     '监管机构'}
    comm_lines = [
        f"  {pzh.get(sp,sp)}社区#{si} --[{rt},{'强' if s else '弱'}]--> "
        f"{pzh.get(tp,tp)}社区#{ti}"
        for sp, si, tp, ti, rt, s in comm_paths
    ]
    node_lines = []
    for ptype, plabel in [('compliance',        '合规'),
                          ('violation',         '违规'),
                          ('violation_partial', '违规(部分)')]:
        for i, rec in enumerate(node_paths.get(ptype, [])[:2], 1):
            step = ' → '.join(loader.node_name(n) for n in rec['nodes'])
            node_lines.append(
                f"  {plabel}路径{i}（质量分{rec['score']}）: {step}")
    return f"""你是资本市场法规合规领域资深专家，请对以下事件进行五维穿透式分析。

【事件描述】
{event_desc}

━━━━ 一、社区级责任链路（宏观）━━━━
{chr(10).join(comm_lines[:20]) if comm_lines else '（未找到社区级路径）'}

━━━━ 二、节点级穿透路径（微观）━━━━
{chr(10).join(node_lines) if node_lines else '（未找到节点级路径）'}

━━━━ 三、知识图谱原始三元组（精细法律依据）━━━━
{triples_text}

━━━━ 四、相关社区法规报告 ━━━━
{community_text}

━━━━ 分析要求（五个维度，纯文本不使用 Markdown）━━━━

1. 责任主体认定：具体指出责任主体类型，引用三元组中的关系为依据
2. 违规行为定性：判断违规类型，援引相关法条
3. 监管机构职责与处罚依据：说明哪些机构可介入，可触发哪些处罚
4. 合规义务梳理：基于合规路径列出应履行的合规义务
5. 风险提示与建议：具体可操作的合规建议及监管趋势提示"""


def get_community_text(loader: DataLoader,
                       matched: Dict[str, Set[int]]) -> str:
    parts = []
    for p, pname in [('responsibility', '责任方'),
                     ('violation',      '违规行为'),
                     ('regulatory',     '监管机构')]:
        comms = sorted(matched.get(p, set()))
        if not comms:
            continue
        parts.append(f"\n{'='*50}\n【{pname}社区报告】\n{'='*50}")
        for cid in comms[:4]:
            df   = loader.reports[p]
            rows = df[df['community'] == cid]
            if rows.empty:
                continue
            row = rows.iloc[0]
            parts.append(f"\n社区#{cid}：{row.get('title','')}")
            parts.append(f"摘要：{str(row.get('summary',''))[:250]}")
            try:
                findings = json.loads(row.get('findings', '[]'))
                for fi, f_item in enumerate(findings[:2], 1):
                    parts.append(
                        f"  发现{fi}：{f_item.get('summary','')} — "
                        f"{f_item.get('explanation','')[:100]}")
            except Exception:
                pass
    return '\n'.join(parts)


def get_triples_text(loader: DataLoader,
                     node_paths: Dict[str, List]) -> str:
    seen:  Set[Tuple] = set()
    lines: List[str]  = []
    ptype_zh = {'compliance':        '合规',
                'violation':         '违规',
                'violation_partial': '违规(部分)'}
    for ptype in ['compliance', 'violation', 'violation_partial']:
        recs = node_paths.get(ptype, [])
        if not recs:
            continue
        for rec in recs:
            lines.append(f"\n[{ptype_zh.get(ptype,ptype)}路径  "
                         f"质量分:{rec['score']}  层次覆盖:{rec.get('layer_coverage',[])}]")
            for src, tgt, rt in rec['edges']:
                key = (src, tgt, rt)
                if key in seen:
                    continue
                seen.add(key)
                lines.append(
                    f"  {loader.node_name(src)}（{loader.node_type(src)}）"
                    f" -[{rt}]-> "
                    f"{loader.node_name(tgt)}（{loader.node_type(tgt)}）")
    return '\n'.join(lines) if lines else '（未找到三元组）'


# ==================== 主系统 ====================

class RegulatoryQuerySystemV5:
    """
    监管违规穿透式查询系统 v5

    相比 v4 的核心升级（步骤4节点路径搜索）：
    - 搜索起点：违规行为节点（而非责任方节点）
    - 搜索算法：VCBSBS（违规中心双向语义束搜索）
    - 路径约束：强制三层（责任方+违规+监管）
    - 质量评分：七维综合评分
    - 结果选择：Jaccard 相似度多样性保证
    """

    def __init__(self, data_dir: str = DATA_DIR):
        self.loader = DataLoader(data_dir)
        self.loader.load()

    def query(self, event_desc: str, output_dir: str = OUTPUT_DIR) -> dict:
        os.makedirs(output_dir, exist_ok=True)
        print("\n" + "=" * 70)
        print("  监管违规穿透式查询系统 v5")
        print("  VCBSBS 违规中心双向语义束搜索 + 七维质量评分")
        print("=" * 70)

        # Step 1: 实体提取
        entities = EntityExtractor(self.loader).extract(event_desc)

        # Step 2: 精准社区匹配
        matched = PrecisionCommunityMatcher(self.loader).match(
            entities, event_desc)

        # Step 3: LLM锚点精选
        anchors = LLMAnchorSelector(self.loader).select(
            entities, matched, event_desc)

        # Step 4: v5 核心 - VCBSBS 路径搜索
        print("\n[步骤4] VCBSBS 节点路径搜索（违规中心双向语义束搜索）...")
        finder = VCBSBSPathFinder(self.loader, entities.get('_all', []))

        print("  → 搜索违规路径（三层约束）...")
        viol_paths = finder.find_violation_paths(
            anchors['responsibility'],
            anchors['violation'],
            anchors['regulatory'],
        )

        print("  → 搜索合规路径...")
        comp_paths = finder.find_compliance_paths(
            anchors['responsibility'],
            anchors['regulatory'],
        )

        node_paths: Dict[str, List] = {
            'compliance': comp_paths,
            'violation':  viol_paths,
        }
        # 分离 violation_partial
        viol_partial = [r for r in viol_paths if r['type'] == 'violation_partial']
        viol_full    = [r for r in viol_paths if r['type'] == 'violation']
        if viol_full or viol_partial:
            node_paths['violation']         = viol_full
            node_paths['violation_partial'] = viol_partial

        print(f"\n  合规路径: {len(comp_paths)} 条")
        print(f"  违规路径（完整三层）: {len(viol_full)} 条")
        print(f"  违规路径（半段）: {len(viol_partial)} 条")

        all_paths = comp_paths + viol_paths
        for i, rec in enumerate(all_paths, 1):
            steps = ' → '.join(self.loader.node_name(n) for n in rec['nodes'])
            detail = rec.get('score_detail', {})
            print(f"\n  路径#{i} [{rec['type']}] 总分:{rec['score']}")
            print(f"    层次:{rec.get('layer_coverage',[])}  跳数:{len(rec['nodes'])-1}")
            print(f"    节点: {steps[:100]}")
            if detail:
                print(f"    得分明细: 三层={detail.get('维度1_三层覆盖',0):.1f} "
                      f"关键词={detail.get('维度2_关键词匹配',0):.1f} "
                      f"跳转={detail.get('维度3_层次跳转',0):.1f} "
                      f"特异={detail.get('维度4_节点特异性',0):.1f} "
                      f"紧凑={detail.get('维度5_路径紧凑性',0):.1f} "
                      f"法律={detail.get('维度7_法律条款覆盖',0):.1f}")

        # Step 5: 社区级路径
        comm_paths = build_community_paths(matched, self.loader.hierarchy)
        print(f"\n[步骤5] 社区级路径: {len(comm_paths)} 条")

        # Step 6: 可视化
        print("\n[步骤6] 生成可视化...")
        comm_vis = os.path.join(output_dir, "event_network.png")
        node_vis = os.path.join(output_dir, "node_level_path.png")
        visualize_community_network(self.loader, matched, comm_paths, comm_vis)
        visualize_node_paths(self.loader, node_paths, node_vis)

        # Step 7: LLM综合分析
        print("\n[步骤7] 调用 DeepSeek 生成综合分析...")
        community_text = get_community_text(self.loader, matched)
        triples_text   = get_triples_text(self.loader, node_paths)
        prompt         = build_prompt(
            event_desc, community_text, triples_text,
            comm_paths, node_paths, self.loader)
        try:
            final_answer = call_deepseek(
                prompt, max_tokens=4000, temperature=0.3)
        except Exception as e:
            final_answer = f"LLM 调用失败：{e}"
            print(f"  错误: {e}")

        # Step 8: 保存输出
        print("\n[步骤8] 保存输出...")
        self._save(output_dir, event_desc, entities, matched, anchors,
                   comm_paths, node_paths, triples_text, community_text,
                   final_answer, prompt)

        print("\n" + "=" * 70)
        print("  查询完成！输出目录:", output_dir)
        print("=" * 70)
        return dict(entities=entities, matched=matched, anchors=anchors,
                    comm_paths=comm_paths, node_paths=node_paths,
                    final_answer=final_answer)

    def _save(self, output_dir, event_desc, entities, matched, anchors,
              comm_paths, node_paths, triples_text, community_text,
              final_answer, prompt):
        ld = self.loader

        def wp(fname, content):
            with open(os.path.join(output_dir, fname), 'w', encoding='utf-8') as f:
                f.write(content)

        # 节点路径文本报告
        lines = ["=" * 65, "节点级穿透路径报告（v5 - VCBSBS）", "=" * 65,
                 f"\n事件：{event_desc}\n"]
        for ptype, pname in [('compliance',        '合规路径'),
                              ('violation',         '违规路径（完整三层）'),
                              ('violation_partial', '违规路径（半段）')]:
            recs = node_paths.get(ptype, [])
            if not recs:
                continue
            lines.append(f"\n【{pname}】")
            for i, rec in enumerate(recs, 1):
                detail = rec.get('score_detail', {})
                lines.append(f"  路径#{i}（{len(rec['nodes'])}跳  "
                              f"总分:{rec['score']}  层次:{rec.get('layer_coverage',[])}）：")
                if detail:
                    lines.append(
                        f"    得分明细：三层覆盖={detail.get('维度1_三层覆盖',0):.1f}  "
                        f"关键词匹配={detail.get('维度2_关键词匹配',0):.1f}  "
                        f"层次跳转={detail.get('维度3_层次跳转',0):.1f}  "
                        f"节点特异={detail.get('维度4_节点特异性',0):.1f}  "
                        f"路径紧凑={detail.get('维度5_路径紧凑性',0):.1f}  "
                        f"边多样={detail.get('维度6_边关系多样性',0):.1f}  "
                        f"法律节点={detail.get('维度7_法律条款覆盖',0):.1f}")
                for src, tgt, rt in rec['edges']:
                    lines.append(
                        f"    {ld.node_name(src)}（{ld.node_type(src)}）"
                        f" -[{rt}]-> "
                        f"{ld.node_name(tgt)}（{ld.node_type(tgt)}）")

        wp("node_paths_report.txt", '\n'.join(lines))
        wp("community_reports.txt", community_text)
        wp("node_triples.txt", triples_text)
        wp("debug_prompt.txt", prompt)
        wp("final_analysis.txt",
           "=" * 65 + "\n监管违规综合分析报告（v5）\n" + "=" * 65 +
           f"\n\n【事件】\n{event_desc}\n\n【分析结果】\n{final_answer}\n")

        self._save_excel(output_dir, event_desc, entities, matched, anchors,
                         comm_paths, node_paths, final_answer)

        for fname in ['event_network.png', 'node_level_path.png',
                      'node_paths_report.txt', 'node_triples.txt',
                      'final_analysis.txt', 'analysis_summary_v5.xlsx']:
            if os.path.exists(os.path.join(output_dir, fname)):
                print(f"    ✓ {fname}")

    def _save_excel(self, output_dir, event_desc, entities, matched, anchors,
                    comm_paths, node_paths, final_answer):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb  = Workbook()
        ld  = self.loader
        pzh = {'responsibility': '责任方',
               'violation':      '违规行为',
               'regulatory':     '监管机构'}

        def hdr(ws, row_num, colors):
            for cell, color in zip(ws[row_num], colors):
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=color)
                cell.alignment = Alignment(wrap_text=True, vertical='center')

        # Sheet1: 事件概述
        ws1 = wb.active
        ws1.title = "事件概述"
        ws1.append(["监管违规穿透分析报告 v5 (VCBSBS)"])
        ws1['A1'].font = Font(size=13, bold=True)
        ws1.append([])
        ws1.append(["事件描述", event_desc])
        ws1.append([])
        for k in ["责任方", "违规行为", "监管机构", "核心法律概念"]:
            ws1.append([k, str(entities.get(k, []))])
        ws1.append([])
        ws1.append(["LLM分析摘要", final_answer[:2000]])
        ws1.column_dimensions['A'].width = 18
        ws1.column_dimensions['B'].width = 100

        # Sheet2: 精筛社区
        ws2 = wb.create_sheet("精筛社区")
        ws2.append(['视角', '社区ID', '社区标题', '摘要', '关键词'])
        hdr(ws2, 1, ['1565C0'] * 5)
        for p in ['responsibility', 'violation', 'regulatory']:
            df = ld.reports[p]
            for cid in sorted(matched.get(p, set())):
                rows = df[df['community'] == cid]
                if rows.empty:
                    continue
                row = rows.iloc[0]
                ws2.append([pzh.get(p, p), cid,
                             row.get('title', ''),
                             str(row.get('summary', ''))[:300],
                             row.get('key_words', '')])
        for col, w in zip('ABCDE', [12, 8, 40, 80, 40]):
            ws2.column_dimensions[col].width = w

        # Sheet3: 节点锚点
        ws3 = wb.create_sheet("节点锚点")
        ws3.append(['视角', '节点ID', '名称', '类型', '层', '度数'])
        hdr(ws3, 1, ['1B5E20'] * 6)
        for p in ['responsibility', 'violation', 'regulatory']:
            for nid in anchors.get(p, []):
                ws3.append([pzh.get(p, p), nid,
                             ld.node_name(nid),
                             ld.node_type(nid),
                             PERSPECTIVE_ZH.get(ld.node_layer(nid), '?'),
                             ld.G.degree(nid)])
        for col, w in zip('ABCDEF', [12, 28, 22, 18, 12, 8]):
            ws3.column_dimensions[col].width = w

        # Sheet4: 节点路径（新增七维得分列）
        ws4 = wb.create_sheet("节点级路径")
        ws4.append(['路径类型', '路径#', '总质量分',
                    '维度1_三层覆盖', '维度2_关键词', '维度3_层次跳转',
                    '维度4_节点特异', '维度5_路径紧凑', '维度6_边多样', '维度7_法律节点',
                    '步骤#', '源节点', '源类型', '关系', '目标节点', '目标类型'])
        hdr(ws4, 1, ['B71C1C'] * 16)

        ptype_zh = {'compliance':        '合规',
                    'violation':         '违规(三层)',
                    'violation_partial': '违规(半段)'}
        for ptype in ['compliance', 'violation', 'violation_partial']:
            for pno, rec in enumerate(node_paths.get(ptype, []), 1):
                detail = rec.get('score_detail', {})
                for step_i, (src, tgt, rt) in enumerate(rec['edges'], 1):
                    ws4.append([
                        ptype_zh.get(ptype, ptype),
                        pno,
                        rec['score'] if step_i == 1 else '',
                        detail.get('维度1_三层覆盖', '') if step_i == 1 else '',
                        detail.get('维度2_关键词匹配', '') if step_i == 1 else '',
                        detail.get('维度3_层次跳转', '') if step_i == 1 else '',
                        detail.get('维度4_节点特异性', '') if step_i == 1 else '',
                        detail.get('维度5_路径紧凑性', '') if step_i == 1 else '',
                        detail.get('维度6_边关系多样性', '') if step_i == 1 else '',
                        detail.get('维度7_法律条款覆盖', '') if step_i == 1 else '',
                        step_i,
                        ld.node_name(src), ld.node_type(src),
                        rt,
                        ld.node_name(tgt), ld.node_type(tgt),
                    ])
        for col, w in zip('ABCDEFGHIJKLMNOP',
                          [14, 6, 8, 10, 10, 10, 10, 10, 8, 10, 6, 24, 20, 14, 24, 20]):
            ws4.column_dimensions[col].width = w

        # Sheet5: 社区级路径
        ws5 = wb.create_sheet("社区级路径")
        ws5.append(['序号', '源视角', '源社区', '目标视角',
                    '目标社区', '关系类型', '强度'])
        hdr(ws5, 1, ['4A148C'] * 7)
        for i, (sp, si, tp, ti, rt, strong) in enumerate(comm_paths, 1):
            ws5.append([i, pzh.get(sp, sp), si,
                        pzh.get(tp, tp), ti, rt,
                        "强" if strong else "弱"])
        for col, w in zip('ABCDEFG', [6, 14, 10, 14, 10, 20, 6]):
            ws5.column_dimensions[col].width = w

        excel_path = os.path.join(output_dir, "analysis_summary_v5.xlsx")
        wb.save(excel_path)


# ==================== 入口 ====================

def main():
    print("=" * 70)
    print("  监管违规穿透式查询系统 v5")
    print("  VCBSBS 算法 | 七维质量评分 | 三层强制约束")
    print("=" * 70)
    system = RegulatoryQuerySystemV5(data_dir=DATA_DIR)
    print("\n请输入事件描述（输入完成后按 Enter，再输入 END 结束）：")
    lines = []
    while True:
        line = input()
        if line.strip().upper() == 'END':
            break
        lines.append(line)
    event_desc = '\n'.join(lines).strip()
    if not event_desc:
        print("错误：事件描述不能为空！")
        return
    results = system.query(event_desc, output_dir=OUTPUT_DIR)
    print("\n" + "=" * 70)
    print("最终分析结果：")
    print("=" * 70)
    print(results['final_answer'])


if __name__ == "__main__":
    main()
