#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
增强版监管违规穿透式查询系统 v2.0
================================================================
在原有社区级责任递进路径基础上，新增节点级证据路径查找：

阶段1. 输入事件描述 → 粗粒度社区激活（LLM实体提取 + 社区匹配）
阶段2. 在激活社区内做细粒度节点召回（BM25 + 模糊匹配）
阶段3. 构建跨视角"小规模证据子图"（1跳扩展 + 强连接边）
阶段4. 约束路径搜索（A* / Dijkstra + 社区流向偏置）
阶段5. 双层输出 + 双层可视化（宏观社区图 + 微观节点图）
================================================================
"""

import json
import os
import re
import time
import warnings
from collections import defaultdict
from typing import Dict, List, Optional, Set, Tuple

import matplotlib
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import FancyArrowPatch, FancyBboxPatch
import networkx as nx
import numpy as np
import pandas as pd
import requests

matplotlib.use("Agg")
plt.rcParams["font.sans-serif"] = ["SimHei", "DejaVu Sans", "Arial Unicode MS"]
plt.rcParams["axes.unicode_minus"] = False
warnings.filterwarnings("ignore")

# ==================== 配置 ====================

API_BASE = "https://api.deepseek.com/v1"
API_KEY = "sk-0a57f72b50854ace9d134a5eb697c4dc"

# 社区流向偏置权重（用于路径搜索打分）
COMMUNITY_FLOW_BONUS = {
    ("responsibility", "violation"): 0.4,
    ("violation", "regulatory"): 0.5,
    ("responsibility", "regulatory"): 0.2,
}

# 关键关系类型加分（越高越优先走这条边）
KEY_RELATION_BONUS = {
    "监管": 0.5,
    "处以": 0.5,
    "执行": 0.4,
    "实施": 0.4,
    "违反": 0.45,
    "做出": 0.35,
    "处理": 0.4,
    "规定": 0.3,
    "包含责任方": 0.3,
    "包含违规主体": 0.3,
    "包含监管机构": 0.3,
    "受到": 0.35,
    "针对": 0.3,
    "依据": 0.25,
    "产生": 0.3,
    "旨在导致": 0.3,
}

# 每个被激活社区最多召回的节点数
TOP_K_NODES_PER_COMMUNITY = 12

# 最终输出的最大路径数
MAX_PATHS_OUTPUT = 5

# 子图最大节点数（防止过大）
MAX_SUBGRAPH_NODES = 200

# 路径搜索最大跳数
MAX_PATH_LENGTH = 8


# ==================== 工具函数 ====================


def remove_markdown_formatting(text: str) -> str:
    """去除 Markdown 格式"""
    if not text:
        return text
    text = re.sub(r"```[\w]*\n?", "", text)
    text = re.sub(r"\*\*(.+?)\*\*", r"\1", text)
    text = re.sub(r"__(.+?)__", r"\1", text)
    text = re.sub(r"\*(.+?)\*", r"\1", text)
    text = re.sub(r"_(.+?)_", r"\1", text)
    text = re.sub(r"^#{1,6}\s+", "", text, flags=re.MULTILINE)
    text = re.sub(r"\[([^\]]+)\]\([^\)]+\)", r"\1", text)
    text = re.sub(r"`([^`]+)`", r"\1", text)
    text = re.sub(r"^[\*\-\+]\s+", "• ", text, flags=re.MULTILINE)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def call_deepseek_api(prompt: str, system_prompt: str = "", max_tokens: int = 2000) -> str:
    """调用 DeepSeek API"""
    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json",
    }
    messages = []
    if system_prompt:
        messages.append({"role": "system", "content": system_prompt})
    messages.append({"role": "user", "content": prompt})

    data = {
        "model": "deepseek-chat",
        "messages": messages,
        "max_tokens": max_tokens,
        "temperature": 0.3,
    }

    for attempt in range(3):
        try:
            resp = requests.post(f"{API_BASE}/chat/completions", headers=headers, json=data, timeout=60)
            if resp.status_code == 200:
                content = resp.json()["choices"][0]["message"]["content"]
                return remove_markdown_formatting(content)
            elif resp.status_code == 429:
                time.sleep(5 * (attempt + 1))
            else:
                print(f"  ⚠ API错误 [{resp.status_code}]: {resp.text[:200]}")
                time.sleep(3)
        except Exception as e:
            print(f"  ⚠ API调用异常: {e}")
            time.sleep(3)
    return ""


# ==================== 阶段1：社区粗匹配 ====================


class CommunityMatcher:
    """阶段1 + 阶段2：社区粗匹配 & 节点细召回"""

    def __init__(
        self,
        regulatory_vis: dict,
        responsibility_vis: dict,
        violation_vis: dict,
        community_hierarchy: pd.DataFrame,
    ):
        self.vis_data = {
            "regulatory": regulatory_vis,
            "responsibility": responsibility_vis,
            "violation": violation_vis,
        }
        self.hierarchy = community_hierarchy

        # 构建节点索引：{perspective: {node_name_lower: [node_dict, ...]}}
        self.node_index: Dict[str, Dict[str, List[dict]]] = {}
        # 构建社区节点列表：{perspective: {community_id: [node_dict, ...]}}
        self.community_nodes: Dict[str, Dict[int, List[dict]]] = {}

        self._build_indices()

    def _build_indices(self):
        for perspective, vis in self.vis_data.items():
            self.node_index[perspective] = defaultdict(list)
            self.community_nodes[perspective] = defaultdict(list)
            for node in vis.get("nodes", []):
                name_lower = node["name"].lower()
                self.node_index[perspective][name_lower].append(node)
                comm_id = int(node["community"])
                self.community_nodes[perspective][comm_id].append(node)

    def extract_entities_llm(self, event_description: str) -> Dict[str, List[str]]:
        """LLM 实体提取"""
        print("  [阶段1] LLM实体提取...")
        prompt = f"""请从以下资本市场事件描述中提取关键实体，分类为：责任方、违规行为、监管机构。

事件描述：
{event_description}

请严格以JSON格式返回，不要包含任何其他文字：
{{
    "责任方": ["实体1", "实体2"],
    "违规行为": ["行为1", "行为2"],
    "监管机构": ["机构1", "机构2"]
}}"""
        resp = call_deepseek_api(prompt)
        try:
            start = resp.find("{")
            end = resp.rfind("}") + 1
            entities = json.loads(resp[start:end])
            print(f"    提取实体: {entities}")
            return entities
        except Exception:
            print("    LLM实体提取失败，使用空实体")
            return {"责任方": [], "违规行为": [], "监管机构": []}

    def fuzzy_match_score(self, keyword: str, node_name: str) -> float:
        """简单模糊匹配分数（包含关系）"""
        kw = keyword.lower().strip()
        nm = node_name.lower().strip()
        if kw == nm:
            return 1.0
        if kw in nm or nm in kw:
            return 0.8
        # 逐字符匹配
        common = sum(1 for c in kw if c in nm)
        if len(kw) > 0:
            return common / max(len(kw), len(nm)) * 0.6
        return 0.0

    def bm25_score(self, query_terms: List[str], node_name: str) -> float:
        """轻量级 BM25-like 打分（无需外部库）"""
        k1, b = 1.5, 0.75
        avg_len = 4.0  # 平均节点名长度（字）
        name_lower = node_name.lower()
        score = 0.0
        for term in query_terms:
            tf = name_lower.count(term.lower())
            if tf == 0:
                continue
            doc_len = len(node_name)
            tf_norm = tf * (k1 + 1) / (tf + k1 * (1 - b + b * doc_len / avg_len))
            score += tf_norm
        return score

    def match_communities(
        self, entities: Dict[str, List[str]], event_description: str
    ) -> Dict[str, Set[int]]:
        """将实体映射到社区（阶段1）"""
        print("  [阶段1] 社区粗匹配...")
        entity_map = {
            "责任方": "responsibility",
            "违规行为": "violation",
            "监管机构": "regulatory",
        }
        matched: Dict[str, Set[int]] = {
            "responsibility": set(),
            "violation": set(),
            "regulatory": set(),
        }

        # 从事件描述中提取所有2字以上词
        event_keywords = list(set(re.findall(r"[\u4e00-\u9fff]{2,6}", event_description)))

        for entity_type, perspective in entity_map.items():
            keywords = entities.get(entity_type, []) + event_keywords
            for kw in keywords:
                if len(kw) < 2:
                    continue
                for node_name, node_list in self.node_index[perspective].items():
                    score = self.fuzzy_match_score(kw, node_name)
                    if score >= 0.7:
                        for node in node_list:
                            matched[perspective].add(int(node["community"]))

        # 沿层级关系做1步传播
        matched = self._propagate_along_hierarchy(matched)

        for p, comms in matched.items():
            print(f"    {p}: 激活社区 {sorted(comms)}")
        return matched

    def _propagate_along_hierarchy(
        self, matched: Dict[str, Set[int]]
    ) -> Dict[str, Set[int]]:
        """沿层级边做1步前向传播，激活下游社区"""
        new_matched = {k: set(v) for k, v in matched.items()}
        for _, row in self.hierarchy.iterrows():
            sp = row.get("source_perspective", "")
            sc = int(row.get("source_community_id", -1))
            tp = row.get("target_perspective", "")
            tc = int(row.get("target_community_id", -1))
            score = float(row.get("score", 0))
            if score < 0.25:
                continue
            if sc in matched.get(sp, set()):
                new_matched.setdefault(tp, set()).add(tc)
            if tc in matched.get(tp, set()):
                new_matched.setdefault(sp, set()).add(sc)
        return new_matched

    def recall_nodes_from_communities(
        self,
        matched_communities: Dict[str, Set[int]],
        entities: Dict[str, List[str]],
        event_description: str,
        top_k: int = TOP_K_NODES_PER_COMMUNITY,
    ) -> Dict[str, List[dict]]:
        """
        阶段2：在激活社区内做节点细召回
        使用 BM25 + 模糊匹配打分，每个社区取 Top-K 节点
        """
        print("  [阶段2] 节点细粒度召回...")
        entity_map = {
            "责任方": "responsibility",
            "违规行为": "violation",
            "监管机构": "regulatory",
        }
        # 查询词
        query_terms = list(set(re.findall(r"[\u4e00-\u9fff]{2,8}", event_description)))
        for etype, keywords in entities.items():
            query_terms.extend(keywords)
        query_terms = list(set(query_terms))

        recalled: Dict[str, List[dict]] = {
            "responsibility": [],
            "violation": [],
            "regulatory": [],
        }

        for entity_type, perspective in entity_map.items():
            for comm_id in matched_communities.get(perspective, set()):
                nodes_in_comm = self.community_nodes[perspective].get(comm_id, [])
                if not nodes_in_comm:
                    continue

                # 打分
                scored = []
                for node in nodes_in_comm:
                    bm25 = self.bm25_score(query_terms, node["name"])
                    fuzzy = max(
                        (self.fuzzy_match_score(kw, node["name"]) for kw in query_terms),
                        default=0.0,
                    )
                    # 核心节点加成
                    core_bonus = 0.3 if node.get("isCore", False) else 0.0
                    # 度数加成（归一化）
                    max_deg = max((n.get("degree", 1) for n in nodes_in_comm), default=1)
                    deg_bonus = node.get("degree", 0) / max(max_deg, 1) * 0.2
                    total = bm25 * 0.4 + fuzzy * 0.3 + core_bonus + deg_bonus
                    scored.append((total, node))

                scored.sort(key=lambda x: x[0], reverse=True)
                top_nodes = [n for _, n in scored[:top_k]]

                # 为每个召回节点标注perspective和community
                for node in top_nodes:
                    node_copy = dict(node)
                    node_copy["perspective"] = perspective
                    node_copy["community_id"] = comm_id
                    recalled[perspective].append(node_copy)

        for p, nodes in recalled.items():
            print(f"    {p}: 召回 {len(nodes)} 个节点")
        return recalled


# ==================== 阶段3：证据子图构建 ====================


class EvidenceSubgraphBuilder:
    """阶段3：构建跨视角证据子图"""

    def __init__(self, kg_nodes: List[dict], kg_edges: List[dict]):
        # 构建完整知识图谱
        self.full_graph = nx.DiGraph()
        self.node_info: Dict[str, dict] = {}

        for node in kg_nodes:
            nid = node["id"]
            self.node_info[nid] = {
                "id": nid,
                "name": node["properties"].get("name", ""),
                "type": node["labels"][0] if node["labels"] else "",
                "labels": node["labels"],
                "properties": node["properties"],
            }
            self.full_graph.add_node(nid, **self.node_info[nid])

        for edge in kg_edges:
            sid = edge["start"]["id"]
            eid = edge["end"]["id"]
            if sid in self.node_info and eid in self.node_info:
                self.full_graph.add_edge(
                    sid,
                    eid,
                    rel_type=edge["label"],
                    weight=edge.get("calculated_weight", 0.5),
                    edge_id=edge["id"],
                )

        print(f"    完整图谱: {self.full_graph.number_of_nodes()} 节点, {self.full_graph.number_of_edges()} 边")

    def _get_node_id_by_name(self, name: str) -> Optional[str]:
        """根据节点名称找KG节点ID"""
        for nid, info in self.node_info.items():
            if info["name"] == name or info["name"].lower() == name.lower():
                return nid
        return None

    def build_subgraph(
        self,
        recalled_nodes: Dict[str, List[dict]],
        community_hierarchy: pd.DataFrame,
        max_nodes: int = MAX_SUBGRAPH_NODES,
    ) -> Tuple[nx.DiGraph, Dict[str, str]]:
        """
        构建证据子图：
        1. 以召回节点为种子，做1跳扩展
        2. 允许跨社区强连接边
        返回：(subgraph, node_perspective_map {node_id -> perspective})
        """
        print("  [阶段3] 构建证据子图...")

        # 收集种子节点（通过名称匹配KG节点ID）
        seed_ids: Set[str] = set()
        node_perspective: Dict[str, str] = {}  # node_id -> perspective

        for perspective, nodes in recalled_nodes.items():
            for node in nodes:
                nid = self._get_node_id_by_name(node["name"])
                if nid:
                    seed_ids.add(nid)
                    node_perspective[nid] = perspective
                else:
                    # 模糊匹配
                    for kg_id, kg_info in self.node_info.items():
                        if (
                            node["name"][:4] in kg_info["name"]
                            or kg_info["name"][:4] in node["name"]
                        ):
                            seed_ids.add(kg_id)
                            node_perspective[kg_id] = perspective
                            break

        print(f"    种子节点: {len(seed_ids)} 个")

        # 1跳扩展
        expanded_ids: Set[str] = set(seed_ids)
        for nid in seed_ids:
            if nid not in self.full_graph:
                continue
            # 出边
            for successor in self.full_graph.successors(nid):
                expanded_ids.add(successor)
            # 入边
            for predecessor in self.full_graph.predecessors(nid):
                expanded_ids.add(predecessor)

        # 限制规模
        if len(expanded_ids) > max_nodes:
            # 优先保留种子节点和度数高的节点
            non_seed = expanded_ids - seed_ids
            deg_sorted = sorted(
                non_seed,
                key=lambda n: self.full_graph.degree(n) if n in self.full_graph else 0,
                reverse=True,
            )
            expanded_ids = seed_ids | set(deg_sorted[: max_nodes - len(seed_ids)])

        # 构建子图
        subgraph = self.full_graph.subgraph(expanded_ids).copy()

        # 为子图节点标注perspective
        perspective_order = ["responsibility", "violation", "regulatory"]
        for nid in subgraph.nodes():
            if nid not in node_perspective:
                # 根据节点类型推断视角
                ntype = self.node_info.get(nid, {}).get("type", "")
                if ntype in {"PartyWithResponsibility", "AdvantageHolder", "Actor"}:
                    node_perspective[nid] = "responsibility"
                elif ntype in {"Action", "Means"}:
                    node_perspective[nid] = "violation"
                elif ntype == "RegulatoryAuthority":
                    node_perspective[nid] = "regulatory"
                else:
                    node_perspective[nid] = "other"

        print(
            f"    子图构建完成: {subgraph.number_of_nodes()} 节点, {subgraph.number_of_edges()} 边"
        )
        return subgraph, node_perspective


# ==================== 阶段4：约束路径搜索 ====================


class ConstrainedPathFinder:
    """阶段4：在证据子图上做约束路径搜索"""

    def __init__(
        self,
        subgraph: nx.DiGraph,
        node_perspective: Dict[str, str],
        node_info: Dict[str, dict],
    ):
        self.G = subgraph
        self.node_perspective = node_perspective
        self.node_info = node_info

        # 构建带综合打分的无向图（用于路径搜索）
        self._build_scored_graph()

    def _edge_cost(self, u: str, v: str) -> float:
        """边代价：越小越优先走"""
        edge_data = self.G.edges.get((u, v), {})
        rel_type = edge_data.get("rel_type", "")
        base_weight = edge_data.get("weight", 0.1)

        # 关系类型加分
        rel_bonus = KEY_RELATION_BONUS.get(rel_type, 0.0)

        # 社区流向加分
        p_u = self.node_perspective.get(u, "other")
        p_v = self.node_perspective.get(v, "other")
        flow_bonus = COMMUNITY_FLOW_BONUS.get((p_u, p_v), 0.0)

        # cost = 1 / (weight + bonus)，分越高cost越低
        score = base_weight + rel_bonus + flow_bonus
        return 1.0 / max(score, 0.01)

    def _build_scored_graph(self):
        """构建带打分的无向图（Dijkstra需要无向或允许双向）"""
        self.scored_G = nx.DiGraph()
        for u, v, data in self.G.edges(data=True):
            cost = self._edge_cost(u, v)
            self.scored_G.add_edge(u, v, cost=cost, **data)

    def find_paths(
        self,
        start_nodes: List[str],
        end_nodes: List[str],
        max_paths: int = MAX_PATHS_OUTPUT,
        max_length: int = MAX_PATH_LENGTH,
    ) -> List[Dict]:
        """
        在子图上找从责任方节点到监管机构节点的最优路径
        使用 Dijkstra（加权最短路径）
        """
        print("  [阶段4] 约束路径搜索...")

        all_paths = []

        # 过滤存在于子图中的节点
        valid_starts = [n for n in start_nodes if n in self.scored_G]
        valid_ends = [n for n in end_nodes if n in self.scored_G]

        print(f"    起点候选: {len(valid_starts)}, 终点候选: {len(valid_ends)}")

        tried = set()
        for s in valid_starts:
            for e in valid_ends:
                if s == e or (s, e) in tried:
                    continue
                tried.add((s, e))
                try:
                    path = nx.dijkstra_path(
                        self.scored_G, s, e, weight="cost"
                    )
                    if len(path) > max_length:
                        continue
                    cost = nx.dijkstra_path_length(self.scored_G, s, e, weight="cost")
                    # 检查路径是否经过违规行为视角
                    perspectives_on_path = [
                        self.node_perspective.get(n, "other") for n in path
                    ]
                    has_violation = "violation" in perspectives_on_path
                    all_paths.append(
                        {
                            "path": path,
                            "cost": cost,
                            "has_violation": has_violation,
                            "length": len(path),
                            "perspectives": perspectives_on_path,
                        }
                    )
                except nx.NetworkXNoPath:
                    pass
                except nx.NodeNotFound:
                    pass

        # 如果有向路径太少，尝试无向路径
        if len(all_paths) < 2:
            undirected = self.scored_G.to_undirected()
            for s in valid_starts:
                for e in valid_ends:
                    if s == e:
                        continue
                    try:
                        path = nx.dijkstra_path(undirected, s, e, weight="cost")
                        if len(path) > max_length:
                            continue
                        cost = nx.dijkstra_path_length(undirected, s, e, weight="cost")
                        perspectives_on_path = [
                            self.node_perspective.get(n, "other") for n in path
                        ]
                        has_violation = "violation" in perspectives_on_path
                        all_paths.append(
                            {
                                "path": path,
                                "cost": cost,
                                "has_violation": has_violation,
                                "length": len(path),
                                "perspectives": perspectives_on_path,
                                "undirected": True,
                            }
                        )
                    except Exception:
                        pass

        # 去重 + 排序：优先经过违规行为视角、cost低、路径短
        seen_path_tuples = set()
        unique_paths = []
        for p in all_paths:
            key = tuple(p["path"])
            if key not in seen_path_tuples:
                seen_path_tuples.add(key)
                unique_paths.append(p)

        unique_paths.sort(
            key=lambda x: (
                0 if x["has_violation"] else 1,
                x["cost"],
                x["length"],
            )
        )

        top_paths = unique_paths[:max_paths]
        print(f"    找到 {len(top_paths)} 条有效路径")
        return top_paths

    def get_path_details(self, path_nodes: List[str]) -> List[Dict]:
        """获取路径上每步的详细信息（节点+边）"""
        details = []
        for i, nid in enumerate(path_nodes):
            node_detail = {
                "index": i,
                "node_id": nid,
                "node_name": self.node_info.get(nid, {}).get("name", nid),
                "node_type": self.node_info.get(nid, {}).get("type", ""),
                "perspective": self.node_perspective.get(nid, "other"),
                "relation_to_next": None,
                "next_node_id": None,
            }
            if i < len(path_nodes) - 1:
                next_id = path_nodes[i + 1]
                edge_data = self.G.edges.get((nid, next_id), self.G.edges.get((next_id, nid), {}))
                node_detail["relation_to_next"] = edge_data.get("rel_type", "→")
                node_detail["next_node_id"] = next_id
            details.append(node_detail)
        return details


# ==================== 阶段5：双层可视化 ====================


class DualLayerVisualizer:
    """阶段5：双层可视化"""

    PERSPECTIVE_COLORS = {
        "responsibility": "#FF6B6B",
        "violation": "#4ECDC4",
        "regulatory": "#95E1D3",
        "other": "#C8C8C8",
    }

    PERSPECTIVE_NAMES = {
        "responsibility": "责任方",
        "violation": "违规行为",
        "regulatory": "监管机构",
        "other": "其他",
    }

    def visualize_macro_community_paths(
        self,
        matched_communities: Dict[str, Set[int]],
        community_hierarchy: pd.DataFrame,
        output_file: str,
    ):
        """
        宏观层：社区级责任递进路径可视化
        高亮激活的社区和路径
        """
        print("  [阶段5] 生成宏观社区层可视化...")

        fig, ax = plt.subplots(figsize=(18, 10))
        ax.set_xlim(-1, 16)
        ax.set_ylim(-0.5, 3.8)
        ax.axis("off")

        # 计算节点位置
        perspective_levels = {
            "responsibility": (3.2, 0),
            "violation": (1.8, 1),
            "regulatory": (0.4, 2),
        }

        # 收集所有涉及的社区
        all_communities: Dict[str, Set[int]] = {
            "responsibility": set(),
            "violation": set(),
            "regulatory": set(),
        }
        for _, row in community_hierarchy.iterrows():
            all_communities[row["source_perspective"]].add(int(row["source_community_id"]))
            all_communities[row["target_perspective"]].add(int(row["target_community_id"]))

        # 计算节点位置
        node_positions: Dict[str, Tuple[float, float]] = {}
        for perspective, comm_set in all_communities.items():
            comms = sorted(comm_set)
            n = len(comms)
            if n == 0:
                continue
            y, _ = perspective_levels[perspective]
            xs = np.linspace(1, 14, n) if n > 1 else [7.5]
            for i, cid in enumerate(comms):
                key = f"{perspective}_{cid}"
                node_positions[key] = (xs[i], y)

        # 绘制连接边（先画，避免遮挡节点）
        for _, row in community_hierarchy.iterrows():
            sp = row["source_perspective"]
            sc = int(row["source_community_id"])
            tp = row["target_perspective"]
            tc = int(row["target_community_id"])
            score = float(row.get("score", 0))

            sk = f"{sp}_{sc}"
            tk = f"{tp}_{tc}"
            if sk not in node_positions or tk not in node_positions:
                continue

            x1, y1 = node_positions[sk]
            x2, y2 = node_positions[tk]

            # 判断是否为激活路径
            is_activated = (
                sc in matched_communities.get(sp, set())
                and tc in matched_communities.get(tp, set())
            )

            if is_activated:
                color, lw, alpha, ls = "#E74C3C", 3.0, 0.95, "-"
            elif score >= 0.5:
                color, lw, alpha, ls = "#555", 1.5, 0.5, "-"
            else:
                color, lw, alpha, ls = "#BBB", 0.8, 0.3, "--"

            arrow = FancyArrowPatch(
                (x1, y1 - 0.18),
                (x2, y2 + 0.18),
                arrowstyle="->,head_width=0.25,head_length=0.25",
                color=color,
                linewidth=lw,
                linestyle=ls,
                alpha=alpha,
                zorder=2,
            )
            ax.add_patch(arrow)

        # 绘制节点
        for key, (x, y) in node_positions.items():
            perspective, cid_str = key.split("_", 1)
            cid = int(cid_str)
            is_activated = cid in matched_communities.get(perspective, set())

            base_color = self.PERSPECTIVE_COLORS[perspective]
            edge_color = "#E74C3C" if is_activated else "#333"
            lw = 3.5 if is_activated else 1.5

            box = FancyBboxPatch(
                (x - 0.38, y - 0.20),
                0.76,
                0.40,
                boxstyle="round,pad=0.05",
                edgecolor=edge_color,
                facecolor=base_color,
                linewidth=lw,
                alpha=0.9,
                zorder=3,
            )
            ax.add_patch(box)

            pname = self.PERSPECTIVE_NAMES[perspective]
            ax.text(x, y + 0.08, pname, ha="center", va="center", fontsize=9, weight="bold", zorder=4)
            ax.text(x, y - 0.08, f"C{cid}", ha="center", va="center", fontsize=9, zorder=4)

        # 层级标签
        for perspective, (y, _) in perspective_levels.items():
            level_map = {"responsibility": "L0", "violation": "L1", "regulatory": "L2"}
            ax.text(-0.5, y, level_map[perspective], fontsize=16, weight="bold", color="#444")

        # 图例
        legend_elements = [
            mpatches.Patch(color=self.PERSPECTIVE_COLORS["responsibility"], label="L0: 责任方"),
            mpatches.Patch(color=self.PERSPECTIVE_COLORS["violation"], label="L1: 违规行为"),
            mpatches.Patch(color=self.PERSPECTIVE_COLORS["regulatory"], label="L2: 监管机构"),
            mpatches.Patch(facecolor="none", edgecolor="#E74C3C", label="激活路径", linewidth=2.5),
        ]
        ax.legend(handles=legend_elements, loc="upper right", fontsize=10, framealpha=0.9)

        plt.title("宏观社区层：责任递进路径（红框/红线=激活）", fontsize=16, weight="bold", pad=20)
        plt.tight_layout()
        plt.savefig(output_file, dpi=200, bbox_inches="tight", facecolor="white")
        plt.close()
        print(f"    宏观图已保存: {output_file}")

    def visualize_micro_evidence_paths(
        self,
        subgraph: nx.DiGraph,
        node_perspective: Dict[str, str],
        node_info: Dict[str, dict],
        top_paths: List[Dict],
        output_file: str,
    ):
        """
        微观层：节点级证据路径可视化
        使用 spring layout，高亮路径节点和边
        """
        print("  [阶段5] 生成微观节点层可视化...")

        if not top_paths:
            print("    无有效路径，跳过微观可视化")
            return

        # 收集路径上所有节点和边
        path_nodes: Set[str] = set()
        path_edges: Set[Tuple[str, str]] = set()
        for path_info in top_paths:
            pnodes = path_info["path"]
            for i, n in enumerate(pnodes):
                path_nodes.add(n)
                if i < len(pnodes) - 1:
                    path_edges.add((pnodes[i], pnodes[i + 1]))
                    path_edges.add((pnodes[i + 1], pnodes[i]))  # 双向标记

        # 只显示子图中与路径相关的节点（路径节点 + 1跳邻居）
        display_nodes: Set[str] = set(path_nodes)
        for n in list(path_nodes):
            if n in subgraph:
                for nb in list(subgraph.successors(n)) + list(subgraph.predecessors(n)):
                    display_nodes.add(nb)
        if len(display_nodes) > 80:
            display_nodes = path_nodes  # 只显示路径节点

        display_subgraph = subgraph.subgraph(display_nodes).copy()

        fig, ax = plt.subplots(figsize=(20, 14))
        ax.axis("off")

        # 布局（路径节点放中心）
        pos = {}
        if len(display_nodes) > 0:
            try:
                # 固定路径节点位置（横向排列）
                path_node_list = top_paths[0]["path"] if top_paths else []
                n_path = len(path_node_list)
                for i, nid in enumerate(path_node_list):
                    pos[nid] = np.array([i * 2.0, 0.0])

                # 其他节点用spring layout
                remaining = [n for n in display_nodes if n not in pos]
                if remaining:
                    sub_remaining = display_subgraph.subgraph(remaining)
                    remaining_pos = nx.spring_layout(
                        sub_remaining, k=1.5, seed=42
                    )
                    # 平移到下方
                    for nid, p in remaining_pos.items():
                        pos[nid] = np.array([p[0] * n_path, p[1] - 2.5])
            except Exception:
                pos = nx.spring_layout(display_subgraph, k=2.0, seed=42)

        # 节点颜色和大小
        node_colors = []
        node_sizes = []
        for nid in display_subgraph.nodes():
            persp = node_perspective.get(nid, "other")
            color = self.PERSPECTIVE_COLORS[persp]
            node_colors.append(color)
            size = 2500 if nid in path_nodes else 800
            node_sizes.append(size)

        # 边颜色和宽度
        edge_colors = []
        edge_widths = []
        for u, v in display_subgraph.edges():
            if (u, v) in path_edges or (v, u) in path_edges:
                edge_colors.append("#E74C3C")
                edge_widths.append(4.0)
            else:
                edge_colors.append("#CCCCCC")
                edge_widths.append(0.8)

        # 绘制
        valid_pos = {n: p for n, p in pos.items() if n in display_subgraph.nodes()}

        nx.draw_networkx_nodes(
            display_subgraph, valid_pos,
            node_color=node_colors,
            node_size=node_sizes,
            alpha=0.85,
            ax=ax,
        )
        nx.draw_networkx_edges(
            display_subgraph, valid_pos,
            edge_color=edge_colors,
            width=edge_widths,
            alpha=0.7,
            arrows=True,
            arrowsize=15,
            ax=ax,
        )

        # 节点标签（只标路径节点）
        path_labels = {}
        for nid in path_nodes:
            if nid in valid_pos:
                name = node_info.get(nid, {}).get("name", nid)
                path_labels[nid] = name[:12] + "..." if len(name) > 12 else name

        nx.draw_networkx_labels(
            display_subgraph, valid_pos,
            labels=path_labels,
            font_size=8,
            font_family="sans-serif",
            ax=ax,
        )

        # 边关系标签（路径边）
        path_edge_labels = {}
        for u, v in display_subgraph.edges():
            if (u, v) in path_edges:
                edge_data = display_subgraph.edges[u, v]
                path_edge_labels[(u, v)] = edge_data.get("rel_type", "")
        nx.draw_networkx_edge_labels(
            display_subgraph, valid_pos,
            edge_labels=path_edge_labels,
            font_size=7,
            ax=ax,
        )

        # 路径文本注解
        path_texts = []
        for i, path_info in enumerate(top_paths[:3]):
            details = [
                node_info.get(nid, {}).get("name", nid)[:8]
                for nid in path_info["path"]
            ]
            path_texts.append(f"路径{i+1}: {' → '.join(details)}")
        if path_texts:
            ax.text(
                0.01, 0.02,
                "\n".join(path_texts),
                transform=ax.transAxes,
                fontsize=8,
                verticalalignment="bottom",
                bbox=dict(boxstyle="round", facecolor="wheat", alpha=0.8),
            )

        # 图例
        legend_elements = [
            mpatches.Patch(color=self.PERSPECTIVE_COLORS["responsibility"], label="责任方节点"),
            mpatches.Patch(color=self.PERSPECTIVE_COLORS["violation"], label="违规行为节点"),
            mpatches.Patch(color=self.PERSPECTIVE_COLORS["regulatory"], label="监管机构节点"),
            mpatches.Patch(color=self.PERSPECTIVE_COLORS["other"], label="其他节点"),
            mpatches.Patch(facecolor="none", edgecolor="#E74C3C", label="证据路径（红线）", linewidth=3),
        ]
        ax.legend(handles=legend_elements, loc="upper right", fontsize=9, framealpha=0.9)

        plt.title("微观节点层：证据路径（红线=最优路径，大节点=路径节点）", fontsize=14, weight="bold")
        plt.tight_layout()
        plt.savefig(output_file, dpi=200, bbox_inches="tight", facecolor="white")
        plt.close()
        print(f"    微观图已保存: {output_file}")


# ==================== 最终报告生成 ====================


def format_path_text(
    path_details: List[Dict],
    node_info: Dict[str, dict],
) -> str:
    """将路径详情格式化为可读文本"""
    steps = []
    for detail in path_details:
        name = detail["node_name"]
        ntype = detail["node_type"]
        perspective = detail["perspective"]
        persp_names = {
            "responsibility": "责任方",
            "violation": "违规",
            "regulatory": "监管",
            "other": "其他",
        }
        step = f"{name}({ntype}, {persp_names.get(perspective, perspective)})"
        if detail["relation_to_next"]:
            step += f"\n  --[{detail['relation_to_next']}]-->\n"
        steps.append(step)
    return "".join(steps)


def generate_final_report_llm(
    event_description: str,
    matched_communities: Dict[str, Set[int]],
    recalled_nodes: Dict[str, List[dict]],
    top_paths: List[Dict],
    path_finder: ConstrainedPathFinder,
    community_reports: Dict[str, pd.DataFrame],
) -> str:
    """调用LLM生成最终综合分析报告"""
    print("  [最终报告] 调用LLM生成分析报告...")

    # 构建路径文本
    path_text_parts = []
    for i, path_info in enumerate(top_paths[:3], 1):
        details = path_finder.get_path_details(path_info["path"])
        path_str = format_path_text(details, path_finder.node_info)
        path_text_parts.append(f"证据路径 {i}（cost={path_info['cost']:.2f}）:\n{path_str}")

    # 构建社区报告摘要
    report_summary_parts = []
    report_map = {
        "responsibility": ("责任方", community_reports.get("responsibility")),
        "violation": ("违规行为", community_reports.get("violation")),
        "regulatory": ("监管机构", community_reports.get("regulatory")),
    }
    for persp, (pname, df) in report_map.items():
        if df is None:
            continue
        for cid in sorted(matched_communities.get(persp, set())):
            rows = df[df["community"] == cid]
            if not rows.empty:
                row = rows.iloc[0]
                report_summary_parts.append(
                    f"[{pname}社区#{cid}] {row.get('title', '')}\n摘要: {str(row.get('summary', ''))[:200]}"
                )

    system_prompt = """你是一位专业的资本市场合规律师和监管专家。请基于提供的知识图谱证据路径和社区报告，
对资本市场违规事件进行深度、结构化的法律分析。回答须专业严谨，直接给出结论，不使用markdown格式。"""

    prompt = f"""你现在是一位非常专业的资本市场合规律师。

已知事实：
{event_description}

从知识图谱中召回的证据路径（按重要性排序）：
{chr(10).join(path_text_parts) if path_text_parts else "未找到有效证据路径"}

相关社区报告摘要（供参考）：
{chr(10).join(report_summary_parts) if report_summary_parts else "无社区报告"}

请严格按照以下结构回答，不得遗漏任何一项：

1. 初步定性结论（一句话）
2. 核心事实依据（引用图谱证据路径中的具体节点和关系）
3. 涉及的主要法规条款（尽量写出条款号）
4. 可能适用的具体违规情形（列出最匹配的2~4项）
5. 可能的法律后果区间（行政/民事/刑事）
6. 对当事人的合规建议（3~6条可执行建议）"""

    response = call_deepseek_api(prompt, system_prompt=system_prompt, max_tokens=3000)
    return response if response else "LLM报告生成失败，请检查API配置。"


# ==================== 主系统类 ====================


class EnhancedRegulatoryQuerySystem:
    """增强版穿透式查询系统（节点级证据路径）"""

    def __init__(self, data_dir: str = "data"):
        self.data_dir = data_dir
        print("=" * 70)
        print("  增强版监管违规穿透式查询系统 v2.0")
        print("  （支持节点级证据路径查找）")
        print("=" * 70)
        self._load_data()

    def _load_data(self):
        print("\n[初始化] 加载数据...")

        # 社区可视化数据
        def load_json(filename):
            path = os.path.join(self.data_dir, filename)
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)

        self.regulatory_vis = load_json("regulatory_visualization_data.json")
        self.responsibility_vis = load_json("responsibility_visualization_data.json")
        self.violation_vis = load_json("violation_visualization_data.json")

        # 社区层级关系
        self.community_hierarchy = pd.read_excel(
            os.path.join(self.data_dir, "community_hierarchy_v3_fixed.xlsx"),
            sheet_name="社区层级关系表",
        )

        # 社区报告
        self.community_reports: Dict[str, Optional[pd.DataFrame]] = {}
        report_files = {
            "responsibility": "责任方社区报告.xlsx",
            "violation": "违规行为社区报告.xlsx",
            "regulatory": "监管机构社区报告.xlsx",
        }
        for persp, fname in report_files.items():
            fpath = os.path.join(self.data_dir, fname)
            if os.path.exists(fpath):
                self.community_reports[persp] = pd.read_excel(fpath)
                print(f"  ✓ {fname} 已加载")
            else:
                self.community_reports[persp] = None
                print(f"  ⚠ {fname} 不存在，跳过")

        # 知识图谱原始数据
        kg_nodes, kg_edges = [], []
        kg_path = os.path.join(self.data_dir, "merged_regulatory_unified.txt")
        with open(kg_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    data = json.loads(line)
                    if data["type"] == "node":
                        kg_nodes.append(data)
                    elif data["type"] == "relationship":
                        kg_edges.append(data)
                except Exception:
                    pass
        print(f"  ✓ 知识图谱: {len(kg_nodes)} 节点, {len(kg_edges)} 关系")

        # 初始化子模块
        self.community_matcher = CommunityMatcher(
            self.regulatory_vis,
            self.responsibility_vis,
            self.violation_vis,
            self.community_hierarchy,
        )
        self.subgraph_builder = EvidenceSubgraphBuilder(kg_nodes, kg_edges)
        self.visualizer = DualLayerVisualizer()

        print("[初始化] 完成！\n")

    def query(self, event_description: str, output_dir: str = "output") -> Dict:
        """执行完整的增强版穿透查询"""
        os.makedirs(output_dir, exist_ok=True)

        print(f"\n{'='*70}")
        print(f"  事件描述: {event_description[:80]}...")
        print(f"{'='*70}\n")

        # ===== 阶段1：实体提取 + 社区粗匹配 =====
        entities = self.community_matcher.extract_entities_llm(event_description)
        matched_communities = self.community_matcher.match_communities(
            entities, event_description
        )

        # ===== 阶段2：节点细粒度召回 =====
        recalled_nodes = self.community_matcher.recall_nodes_from_communities(
            matched_communities, entities, event_description
        )

        # ===== 阶段3：构建证据子图 =====
        subgraph, node_perspective = self.subgraph_builder.build_subgraph(
            recalled_nodes, self.community_hierarchy
        )

        # ===== 阶段4：约束路径搜索 =====
        path_finder = ConstrainedPathFinder(
            subgraph, node_perspective, self.subgraph_builder.node_info
        )

        # 确定起点（责任方节点）和终点（监管机构节点）
        start_nodes = [
            self.subgraph_builder._get_node_id_by_name(n["name"])
            for n in recalled_nodes.get("responsibility", [])
            if self.subgraph_builder._get_node_id_by_name(n["name"])
        ]
        end_nodes = [
            self.subgraph_builder._get_node_id_by_name(n["name"])
            for n in recalled_nodes.get("regulatory", [])
            if self.subgraph_builder._get_node_id_by_name(n["name"])
        ]

        # 如果没有监管机构节点，用子图中所有regulatory节点
        if not end_nodes:
            end_nodes = [
                nid
                for nid, persp in node_perspective.items()
                if persp == "regulatory" and nid in subgraph
            ]

        top_paths = path_finder.find_paths(start_nodes, end_nodes)

        # ===== 阶段5A：宏观可视化 =====
        macro_file = os.path.join(output_dir, "macro_community_paths.png")
        self.visualizer.visualize_macro_community_paths(
            matched_communities, self.community_hierarchy, macro_file
        )

        # ===== 阶段5B：微观可视化 =====
        micro_file = os.path.join(output_dir, "micro_evidence_paths.png")
        self.visualizer.visualize_micro_evidence_paths(
            subgraph,
            node_perspective,
            self.subgraph_builder.node_info,
            top_paths,
            micro_file,
        )

        # ===== 生成路径文本报告 =====
        path_report_lines = []
        path_report_lines.append("=" * 60)
        path_report_lines.append("宏观责任链（社区级）")
        path_report_lines.append("=" * 60)

        # 构建宏观路径文本
        activated_hierarchy = self.community_hierarchy[
            self.community_hierarchy.apply(
                lambda r: (
                    int(r["source_community_id"]) in matched_communities.get(r["source_perspective"], set())
                    and int(r["target_community_id"]) in matched_communities.get(r["target_perspective"], set())
                ),
                axis=1,
            )
        ]
        if not activated_hierarchy.empty:
            for _, row in activated_hierarchy.iterrows():
                sp = row["source_perspective"]
                sc = int(row["source_community_id"])
                tp = row["target_perspective"]
                tc = int(row["target_community_id"])
                rt = row["relation_type"]
                score = float(row.get("score", 0))
                pnames = {"responsibility": "责任方", "violation": "违规行为", "regulatory": "监管机构"}
                path_report_lines.append(
                    f"  {pnames.get(sp, sp)}社区#{sc} --[{rt}, score={score:.2f}]--> {pnames.get(tp, tp)}社区#{tc}"
                )
        else:
            path_report_lines.append("  未找到激活的社区连接路径")

        path_report_lines.append("")
        path_report_lines.append("=" * 60)
        path_report_lines.append(f"微观证据链（节点级 Top-{len(top_paths)} 路径）")
        path_report_lines.append("=" * 60)

        if top_paths:
            for i, path_info in enumerate(top_paths, 1):
                details = path_finder.get_path_details(path_info["path"])
                path_report_lines.append(f"\n路径{i}（路径长度={path_info['length']}, 综合cost={path_info['cost']:.3f}）:")
                for detail in details:
                    name = detail["node_name"]
                    ntype = detail["node_type"]
                    persp = {"responsibility": "责任方", "violation": "违规", "regulatory": "监管", "other": "其他"}.get(
                        detail["perspective"], detail["perspective"]
                    )
                    if detail["relation_to_next"]:
                        path_report_lines.append(f"  [{persp}] {name}（{ntype}）")
                        path_report_lines.append(f"    ↓ [{detail['relation_to_next']}]")
                    else:
                        path_report_lines.append(f"  [{persp}] {name}（{ntype}）")
        else:
            path_report_lines.append("  未找到有效的节点级证据路径")

        path_report_text = "\n".join(path_report_lines)

        # 保存路径报告
        path_report_file = os.path.join(output_dir, "dual_layer_paths.txt")
        with open(path_report_file, "w", encoding="utf-8") as f:
            f.write(path_report_text)
        print(f"\n  双层路径报告已保存: {path_report_file}")
        print("\n" + path_report_text)

        # ===== LLM最终分析报告 =====
        final_answer = generate_final_report_llm(
            event_description,
            matched_communities,
            recalled_nodes,
            top_paths,
            path_finder,
            self.community_reports,
        )

        final_report_file = os.path.join(output_dir, "final_analysis.txt")
        with open(final_report_file, "w", encoding="utf-8") as f:
            f.write("=" * 70 + "\n")
            f.write("监管违规事件综合分析报告（增强版）\n")
            f.write("=" * 70 + "\n\n")
            f.write(f"【事件描述】\n{event_description}\n\n")
            f.write(f"【双层路径分析】\n{path_report_text}\n\n")
            f.write(f"【综合分析结论】\n{final_answer}\n")
        print(f"\n  最终分析报告已保存: {final_report_file}")

        # ===== 生成汇总Excel =====
        self._save_summary_excel(
            event_description,
            entities,
            matched_communities,
            recalled_nodes,
            top_paths,
            path_finder,
            os.path.join(output_dir, "analysis_summary.xlsx"),
        )

        print(f"\n{'='*70}")
        print("  查询完成！")
        print(f"  输出目录: {output_dir}")
        print(f"  - 宏观社区图: macro_community_paths.png")
        print(f"  - 微观节点图: micro_evidence_paths.png")
        print(f"  - 双层路径报告: dual_layer_paths.txt")
        print(f"  - 综合分析报告: final_analysis.txt")
        print(f"  - 汇总Excel: analysis_summary.xlsx")
        print(f"{'='*70}\n")

        return {
            "entities": entities,
            "matched_communities": matched_communities,
            "recalled_nodes": recalled_nodes,
            "top_paths": top_paths,
            "path_report": path_report_text,
            "final_answer": final_answer,
            "macro_file": macro_file,
            "micro_file": micro_file,
        }

    def _save_summary_excel(
        self,
        event_description: str,
        entities: Dict,
        matched_communities: Dict[str, Set[int]],
        recalled_nodes: Dict[str, List[dict]],
        top_paths: List[Dict],
        path_finder: ConstrainedPathFinder,
        output_file: str,
    ):
        """保存汇总Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill

            wb = Workbook()

            # Sheet1: 事件概述
            ws1 = wb.active
            ws1.title = "事件概述"
            ws1["A1"] = "事件描述"
            ws1["A1"].font = Font(bold=True)
            ws1["B1"] = event_description
            ws1["A2"] = "提取实体"
            ws1["A2"].font = Font(bold=True)
            ws1["B2"] = json.dumps(entities, ensure_ascii=False)
            ws1.column_dimensions["B"].width = 80

            # Sheet2: 激活社区
            ws2 = wb.create_sheet("激活社区")
            ws2.append(["视角", "社区ID"])
            for p, comms in matched_communities.items():
                for c in sorted(comms):
                    ws2.append([p, c])

            # Sheet3: 召回节点
            ws3 = wb.create_sheet("召回节点")
            ws3.append(["视角", "社区ID", "节点名称", "节点类型", "是否核心", "度数"])
            for p, nodes in recalled_nodes.items():
                for n in nodes:
                    ws3.append([
                        p,
                        n.get("community_id", ""),
                        n.get("name", ""),
                        n.get("type", ""),
                        "是" if n.get("isCore") else "否",
                        n.get("degree", 0),
                    ])

            # Sheet4: 证据路径
            ws4 = wb.create_sheet("证据路径")
            ws4.append(["路径序号", "步骤", "节点名称", "节点类型", "视角", "关系到下一节点"])
            for i, path_info in enumerate(top_paths, 1):
                details = path_finder.get_path_details(path_info["path"])
                for detail in details:
                    ws4.append([
                        i,
                        detail["index"] + 1,
                        detail["node_name"],
                        detail["node_type"],
                        detail["perspective"],
                        detail.get("relation_to_next", ""),
                    ])

            for ws in [ws3, ws4]:
                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = 20

            wb.save(output_file)
            print(f"  汇总Excel已保存: {output_file}")
        except Exception as e:
            print(f"  ⚠ 保存Excel失败: {e}")


# ==================== 主函数 ====================


def main():
    """主函数（交互式输入）"""
    system = EnhancedRegulatoryQuerySystem(data_dir="data")

    print("请输入事件描述（输入完成后按Enter，再输入'END'并按Enter结束）：")
    print("-" * 60)

    lines = []
    while True:
        line = input()
        if line.strip() == "END":
            break
        lines.append(line)

    event_description = "\n".join(lines).strip()
    if not event_description:
        print("错误: 事件描述不能为空！")
        return

    results = system.query(event_description, output_dir="enhanced_output")

    print("\n" + "=" * 70)
    print("最终分析结论:")
    print("=" * 70)
    print(results["final_answer"])


if __name__ == "__main__":
    main()
