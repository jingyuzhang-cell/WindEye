#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
步骤5：社区报告生成器
================================================================
读取 step1_gnn_clustering.py 输出的聚类结果，
调用 DeepSeek API 为三种视角每个社区生成结构化报告。

输出：
  community_reports/监管机构社区报告.xlsx
  community_reports/违规行为社区报告.xlsx
  community_reports/责任方社区报告.xlsx

Excel 字段（与原版 generate_community_reports.py 完全一致）：
  id, community, level, title, parent, children,
  summary, key_words, findings, rank,
  rating_explanation, full_content_json
================================================================
"""

import os
import sys
import json
import uuid
import time
import hashlib
import traceback
from typing import Dict, List, Tuple, Optional
from pathlib import Path

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CURRENT_DIR = Path(__file__).resolve().parent
PARENT_DIR = CURRENT_DIR.parent
if str(PARENT_DIR) not in sys.path:
    sys.path.insert(0, str(PARENT_DIR))

from report_settings import (
    CLUSTERING_OUTPUT_DIR as DEFAULT_CLUSTERING_OUTPUT_DIR,
    COMMUNITY_REPORTS_DIR as DEFAULT_COMMUNITY_REPORTS_DIR,
    DEEPSEEK_API_BASE,
    DEEPSEEK_API_KEY,
    DEEPSEEK_MODEL,
    GRAPH_DATA_FILE as DEFAULT_GRAPH_DATA_FILE,
    POLICY_DOC_DIR as DEFAULT_POLICY_DOC_DIR,
    ensure_output_dirs,
)

# ══════════════════════════════════════════════════════════════
# 配置
# ══════════════════════════════════════════════════════════════

API_BASE   = DEEPSEEK_API_BASE
API_KEY    = DEEPSEEK_API_KEY
MODEL_NAME = DEEPSEEK_MODEL

# 输入目录（step1 的输出目录）
CLUSTERING_RESULT_DIR = str(DEFAULT_CLUSTERING_OUTPUT_DIR)
# 政策法规原文目录（可选）
POLICY_DOC_DIR        = str(DEFAULT_POLICY_DOC_DIR)
# 图谱数据文件
GRAPH_DATA_FILE       = str(DEFAULT_GRAPH_DATA_FILE)
# 输出目录
OUTPUT_DIR            = str(DEFAULT_COMMUNITY_REPORTS_DIR)

PERSPECTIVES = {
    "responsibility": "责任方",
    "regulatory":     "监管机构",
    "violation":      "违规行为",
}

MAX_RETRIES      = 3
RETRY_DELAY      = 5
REQUEST_INTERVAL = 1.5
MAX_TOKENS       = 4096
TEMPERATURE      = 1.0

# ══════════════════════════════════════════════════════════════
# 数据加载
# ══════════════════════════════════════════════════════════════

def load_graph_data(filepath: str) -> Tuple[Dict, List]:
    nodes_dict, rels_list = {}, []
    with open(filepath, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                data = json.loads(line)
                if data.get('type') == 'node':
                    nodes_dict[data['id']] = data
                elif data.get('type') == 'relationship':
                    rels_list.append(data)
            except json.JSONDecodeError:
                continue
    print(f"  ✓ 图谱: {len(nodes_dict)} 节点, {len(rels_list)} 关系")
    return nodes_dict, rels_list


def load_policy_documents(policy_dir: str) -> str:
    if not os.path.exists(policy_dir):
        print(f"  ⚠ 政策原文目录不存在: {policy_dir}，仅使用图谱数据")
        return ""
    all_text = []
    for fname in sorted(f for f in os.listdir(policy_dir) if f.endswith('.txt')):
        fpath   = os.path.join(policy_dir, fname)
        content = open(fpath, 'r', encoding='utf-8').read().strip()
        if content:
            all_text.append(f"=== {fname} ===\n{content}")
            print(f"  ✓ 已加载: {fname} ({len(content)} 字)")
    return "\n\n".join(all_text)


def load_clustering_results(
    result_dir: str, perspective_key: str
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, Optional[dict]]:
    px = perspective_key
    df_nodes   = pd.read_csv(os.path.join(result_dir, f"{px}_1_节点聚类结果.csv"),   encoding='utf-8-sig')
    df_rels    = pd.read_csv(os.path.join(result_dir, f"{px}_2_关系聚类结果.csv"),   encoding='utf-8-sig')
    df_summary = pd.read_csv(os.path.join(result_dir, f"{px}_3_社区统计摘要.csv"),   encoding='utf-8-sig')

    vis_path = os.path.join(result_dir, f"{px}_visualization_data.json")
    vis_data = json.load(open(vis_path, 'r', encoding='utf-8')) if os.path.exists(vis_path) else None

    df_summary = df_summary[df_summary['社区ID'] != '全局统计'].copy()
    df_summary['社区ID'] = df_summary['社区ID'].astype(int)

    pname = PERSPECTIVES[perspective_key]
    print(f"  ✓ {pname}: {len(df_nodes)} 节点, {len(df_rels)} 关系, "
          f"{len(df_summary)} 社区")
    return df_nodes, df_rels, df_summary, vis_data


# ══════════════════════════════════════════════════════════════
# 社区数据提取
# ══════════════════════════════════════════════════════════════

def extract_community_data(
    community_id: int,
    df_nodes: pd.DataFrame,
    df_rels:  pd.DataFrame,
    df_summary: pd.DataFrame,
    nodes_dict: Dict,
) -> Dict:
    summary_row  = df_summary[df_summary['社区ID'] == community_id]
    summary_info = summary_row.iloc[0].to_dict() if len(summary_row) > 0 else {}

    comm_nodes   = df_nodes[df_nodes['所属社区'] == community_id].copy()
    comm_node_ids = set(comm_nodes['节点ID'].tolist())

    internal_rels = df_rels[
        df_rels['起始节点ID'].isin(comm_node_ids) &
        df_rels['终止节点ID'].isin(comm_node_ids)
    ].copy()

    cross_rels = df_rels[
        (df_rels['起始节点ID'].isin(comm_node_ids) |
         df_rels['终止节点ID'].isin(comm_node_ids)) &
        (df_rels['是否跨社区'] == '是')
    ].copy()

    nodes_by_type: Dict[str, List] = {}
    for _, row in comm_nodes.iterrows():
        t = row['节点类型']
        if t not in nodes_by_type:
            nodes_by_type[t] = []
        nodes_by_type[t].append({
            'id':             row['节点ID'],
            'name':           row['节点名称'],
            'type':           t,
            'is_core':        row['是否核心类型'] == '是',
            'degree':         int(row['度数']),
            'weighted_degree': float(row['加权度数']),
            'in_degree':      int(row['入度']),
            'out_degree':     int(row['出度']),
        })

    relations = []
    for _, row in internal_rels.iterrows():
        relations.append({
            'id':          row['关系ID'],
            'type':        row['关系类型'],
            'weight':      float(row['计算权重']),
            'source':      row['起始节点名称'],
            'source_id':   row['起始节点ID'],
            'source_type': row['起始节点类型'],
            'target':      row['终止节点名称'],
            'target_id':   row['终止节点ID'],
            'target_type': row['终止节点类型'],
        })
    relations.sort(key=lambda x: x['weight'], reverse=True)

    return {
        'community_id':        community_id,
        'summary_stats':       summary_info,
        'nodes_by_type':       nodes_by_type,
        'total_nodes':         len(comm_nodes),
        'total_internal_rels': len(internal_rels),
        'total_cross_rels':    len(cross_rels),
        'top_relations':       relations[:80],
        'all_node_ids':        list(comm_node_ids),
    }


def find_relevant_policy_text(
    community_data: Dict, nodes_dict: Dict, policy_text: str
) -> str:
    if not policy_text:
        return ""
    terms = set()
    for ntype, nlist in community_data['nodes_by_type'].items():
        if ntype in ('Section', 'Chapter', 'Title', 'Law'):
            for node in nlist:
                if node['name']:
                    terms.add(node['name'])
    if not terms:
        return ""
    matched = []
    for line in policy_text.split('\n'):
        ls = line.strip()
        if not ls or ls.startswith('==='):
            continue
        if any(t in ls for t in terms):
            matched.append(ls)
    result = '\n'.join(matched[:30])
    return result[:3000] + '\n...(已截断)' if len(result) > 3000 else result


# ══════════════════════════════════════════════════════════════
# DeepSeek API 调用
# ══════════════════════════════════════════════════════════════

def call_deepseek_api(prompt: str, system_prompt: str) -> Optional[str]:
    if not API_KEY:
        return None
    headers = {
        'Content-Type':  'application/json',
        'Authorization': f'Bearer {API_KEY}',
    }
    payload = {
        'model':    MODEL_NAME,
        'messages': [
            {'role': 'system', 'content': system_prompt},
            {'role': 'user',   'content': prompt},
        ],
        'max_tokens':  MAX_TOKENS,
        'temperature': TEMPERATURE,
    }
    for attempt in range(MAX_RETRIES):
        try:
            resp = requests.post(f"{API_BASE}/chat/completions",
                                 headers=headers, json=payload, timeout=120)
            if resp.status_code == 200:
                return resp.json()['choices'][0]['message']['content']
            elif resp.status_code == 429:
                wait = RETRY_DELAY * (attempt + 1)
                print(f"    ⚠ 频率限制，等待{wait}s...")
                time.sleep(wait)
            else:
                print(f"    ⚠ API错误 [{resp.status_code}]: {resp.text[:200]}")
                if attempt < MAX_RETRIES - 1:
                    time.sleep(RETRY_DELAY)
        except requests.exceptions.Timeout:
            print(f"    ⚠ 超时 ({attempt+1}/{MAX_RETRIES})")
            if attempt < MAX_RETRIES - 1:
                time.sleep(RETRY_DELAY)
        except Exception as e:
            print(f"    ⚠ 异常: {e}")
            if attempt < MAX_RETRIES - 1:
                time.sleep(RETRY_DELAY)
    return None


# ══════════════════════════════════════════════════════════════
# Prompt 构建
# ══════════════════════════════════════════════════════════════

def build_report_prompt(
    community_data: Dict,
    perspective_name: str,
    relevant_policy: str,
) -> Tuple[str, str]:
    system_prompt = f"""你是一位资本市场法规监管领域的专业分析师。你的任务是基于知识图谱的GNN社区聚类结果和相关政策法规原文，为指定社区生成结构化分析报告。

当前分析视角：{perspective_name}社区（基于GNN多视角嵌入增强的Leiden聚类）

请严格按照JSON格式输出报告，包含以下字段：
- title: 社区报告标题（简洁概括社区核心主题，15-30字）
- summary: 执行摘要（150-300字，概述社区整体结构、主要实体、实体间关系、GNN嵌入揭示的潜在关联）
- rating: 重要性/影响力评分（0.0-10.0的浮点数）
- rating_explanation: 评分解释（一句话说明评分依据）
- key_words: 关键词列表，考虑角度：针对"信披违规、高价超募、抱团压价、重大舆情、操纵市场"中哪类场景，约束哪些责任方，约束哪些行为
- findings: 发现/洞察列表（3-6条），每条包含：
  - summary: 一句话概括（10-20字）
  - explanation: 详细解释（80-200字，引用格式：[Data: Entities (ID1,ID2), Relationships (ID1,ID2,+more)]）

仅输出合法JSON对象，不含任何其他文字、Markdown标记或代码块标记。"""

    # 节点描述
    nodes_desc_parts = []
    for ntype, nlist in community_data['nodes_by_type'].items():
        sorted_nodes = sorted(nlist, key=lambda x: x['weighted_degree'], reverse=True)
        lines = []
        for n in sorted_nodes[:15]:
            core_mark = ' [核心]' if n['is_core'] else ''
            lines.append(
                f"  - {n['name']} (ID:{n['id']}, "
                f"度={n['degree']}, 加权度={n['weighted_degree']:.1f}, "
                f"入={n['in_degree']}, 出={n['out_degree']}){core_mark}"
            )
        remaining = len(nlist) - len(lines)
        if remaining > 0:
            lines.append(f"  ... 另有 {remaining} 个{ntype}节点")
        nodes_desc_parts.append(
            f"\n【{ntype}】({len(nlist)}个):\n" + '\n'.join(lines)
        )
    nodes_desc = '\n'.join(nodes_desc_parts)

    # 关系描述
    rels_lines = []
    for rel in community_data['top_relations'][:40]:
        rels_lines.append(
            f"  - [{rel['type']}] {rel['source']}({rel['source_type']}) → "
            f"{rel['target']}({rel['target_type']}) "
            f"(ID:{rel['id']}, 权重={rel['weight']:.2f})"
        )
    if len(community_data['top_relations']) > 40:
        rels_lines.append(f"  ... 另有 {len(community_data['top_relations'])-40} 条关系")
    rels_desc = '\n'.join(rels_lines)

    # 统计描述
    stats = community_data['summary_stats']
    stats_desc = (
        f"社区编号: {community_data['community_id']}\n"
        f"节点总数: {community_data['total_nodes']}\n"
        f"内部关系数: {community_data['total_internal_rels']}\n"
        f"跨社区关系数: {community_data['total_cross_rels']}\n"
        f"核心节点数: {stats.get('核心节点数','N/A')}\n"
        f"子图密度: {stats.get('子图密度','N/A')}\n"
        f"平均度数: {stats.get('平均度数','N/A')}\n"
        f"平均加权度数: {stats.get('平均加权度数','N/A')}\n"
        f"总权重: {stats.get('总权重','N/A')}\n"
        f"节点类型分布: {stats.get('节点类型分布','N/A')}"
    )

    user_prompt = f"""请为以下{perspective_name}社区（GNN嵌入增强Leiden聚类）生成结构化分析报告。

## 一、社区统计信息
{stats_desc}

## 二、社区节点详情
{nodes_desc}

## 三、社区内部关系（按权重降序，Top 40）
{rels_desc}
"""
    if relevant_policy:
        user_prompt += f"\n## 四、相关政策法规原文\n{relevant_policy}\n"

    user_prompt += """
请基于以上信息生成JSON格式的社区报告。注意：
1. findings中的explanation必须引用具体的实体ID和关系ID
2. key_words要从"信披违规、高价超募、抱团压价、重大舆情、操纵市场"等场景角度考虑
3. rating需综合考虑社区规模、核心节点数、密度和法律影响力
4. 仅输出JSON对象，不要包含```json标记或其他文字"""

    return system_prompt, user_prompt


# ══════════════════════════════════════════════════════════════
# 响应解析 & 回退方案
# ══════════════════════════════════════════════════════════════

def parse_llm_response(text: str, community_id: int) -> Optional[Dict]:
    if not text:
        return None
    text = text.strip()
    # 去除 Markdown 代码块标记
    for marker in ['```json', '```']:
        if text.startswith(marker):
            text = text[len(marker):]
    if text.endswith('```'):
        text = text[:-3]
    text = text.strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError as e:
        print(f"    ⚠ JSON解析失败(社区{community_id}): {e}")
        s, e_idx = text.find('{'), text.rfind('}')
        if s != -1 and e_idx > s:
            try:
                return json.loads(text[s:e_idx + 1])
            except json.JSONDecodeError:
                pass
        print(f"    ⚠ 无法解析响应，使用回退方案")
        return None


def generate_fallback_report(community_data: Dict, perspective_name: str) -> Dict:
    stats     = community_data['summary_stats']
    all_nodes = [n for nlist in community_data['nodes_by_type'].values() for n in nlist]
    core_names = [n['name'] for n in all_nodes if n['is_core']]
    all_names  = [n['name'] for n in all_nodes]
    rel_types  = set(r['type'] for r in community_data['top_relations'])

    top_entities = (core_names or all_names)[:3]
    title = f"{perspective_name}社区{community_data['community_id']}: " + \
            '、'.join(top_entities[:3])
    if len(title) > 50:
        title = title[:47] + '...'

    summary = (
        f"本社区（{perspective_name}视角-社区{community_data['community_id']}，"
        f"基于GNN嵌入增强Leiden聚类）"
        f"包含{community_data['total_nodes']}个节点和"
        f"{community_data['total_internal_rels']}条内部关系。"
        f"核心节点：{'、'.join(core_names[:5]) if core_names else '无'}。"
        f"主要关系类型：{'、'.join(list(rel_types)[:5])}。"
        f"子图密度={stats.get('子图密度','N/A')}，"
        f"平均加权度数={stats.get('平均加权度数','N/A')}。"
    )

    findings = []
    if all_nodes:
        top_node = sorted(all_nodes, key=lambda x: x['weighted_degree'], reverse=True)[0]
        findings.append({
            'summary':     f"{top_node['name']}为社区核心枢纽",
            'explanation': (
                f"{top_node['name']}加权度数最高"
                f"（weightedDegree={top_node['weighted_degree']:.1f}），"
                f"度数={top_node['degree']}，为GNN嵌入空间中的中心节点。"
                f"[Data: Entities ({top_node['id']})]"
            ),
        })
    if community_data['top_relations']:
        tr = community_data['top_relations'][0]
        findings.append({
            'summary':     f"最强关系「{tr['type']}」",
            'explanation': (
                f"社区内权重最高关系：{tr['source']} → {tr['target']}，"
                f"类型=「{tr['type']}」，权重={tr['weight']:.2f}。"
                f"[Data: Relationships ({tr['id']})]"
            ),
        })
    findings.append({
        'summary':     "社区结构特征",
        'explanation': (
            f"含{community_data['total_nodes']}节点，"
            f"{community_data['total_internal_rels']}条内部关系，"
            f"{community_data['total_cross_rels']}条跨社区关系，"
            f"密度={stats.get('子图密度','N/A')}。"
            f"类型分布：{stats.get('节点类型分布','N/A')}。"
        ),
    })

    key_words = list(rel_types)[:5] + [n for n in core_names[:3] if n not in rel_types]

    return {
        'title':              title,
        'summary':            summary,
        'rating':             round(min(
                                  community_data['total_nodes'] * 0.3 +
                                  community_data['total_internal_rels'] * 0.1 +
                                  int(stats.get('核心节点数', 0)) * 0.5,
                                  10.0), 1),
        'rating_explanation': (
            f"基于社区规模({community_data['total_nodes']}节点)和"
            f"核心节点数({stats.get('核心节点数',0)})的自动评分"
        ),
        'key_words': key_words,
        'findings':  findings,
    }


# ══════════════════════════════════════════════════════════════
# 报告生成主流程
# ══════════════════════════════════════════════════════════════

def generate_report_id(perspective: str, community_id: int) -> str:
    raw = f"{perspective}_{community_id}_{uuid.uuid4().hex[:8]}"
    return hashlib.md5(raw.encode()).hexdigest()[:16]


def generate_community_report(
    community_id:    int,
    df_nodes:        pd.DataFrame,
    df_rels:         pd.DataFrame,
    df_summary:      pd.DataFrame,
    nodes_dict:      Dict,
    policy_text:     str,
    perspective_key: str,
    perspective_name: str,
) -> Dict:
    community_data  = extract_community_data(
        community_id, df_nodes, df_rels, df_summary, nodes_dict
    )
    relevant_policy = find_relevant_policy_text(community_data, nodes_dict, policy_text)
    system_p, user_p = build_report_prompt(community_data, perspective_name, relevant_policy)

    print(f"    → 调用API (社区{community_id})...")
    llm_resp       = call_deepseek_api(user_p, system_p)
    report_content = parse_llm_response(llm_resp, community_id)

    if report_content is None:
        print(f"    → 回退方案 (社区{community_id})")
        report_content = generate_fallback_report(community_data, perspective_name)

    # 规范化 findings
    findings = report_content.get('findings', [])
    if isinstance(findings, str):
        try:
            findings = json.loads(findings)
        except json.JSONDecodeError:
            findings = [{'summary': findings, 'explanation': findings}]

    full_content = {
        'title':              report_content.get('title', f"{perspective_name}社区{community_id}"),
        'summary':            report_content.get('summary', ''),
        'rating':             float(report_content.get('rating', 5.0)),
        'rating_explanation': report_content.get('rating_explanation', ''),
        'key_words':          report_content.get('key_words', []),
        'findings':           findings,
    }

    return {
        'id':                 generate_report_id(perspective_key, community_id),
        'community':          community_id,
        'level':              0,
        'title':              full_content['title'],
        'parent':             -1,
        'children':           json.dumps([], ensure_ascii=False),
        'summary':            full_content['summary'],
        'key_words':          json.dumps(full_content['key_words'], ensure_ascii=False),
        'findings':           json.dumps(findings, ensure_ascii=False),
        'rank':               full_content['rating'],
        'rating_explanation': full_content['rating_explanation'],
        'full_content_json':  json.dumps(full_content, ensure_ascii=False),
    }


def generate_reports_for_perspective(
    perspective_key:  str,
    perspective_name: str,
    result_dir:       str,
    nodes_dict:       Dict,
    policy_text:      str,
) -> List[Dict]:
    print(f"\n{'=' * 60}")
    print(f"  生成 {perspective_name} 社区报告")
    print(f"{'=' * 60}")

    df_nodes, df_rels, df_summary, _ = load_clustering_results(result_dir, perspective_key)
    community_ids = sorted(df_summary['社区ID'].unique().tolist())
    print(f"  共 {len(community_ids)} 个社区")

    reports = []
    for idx, cid in enumerate(community_ids):
        print(f"\n  [{idx+1}/{len(community_ids)}] 社区 {cid}:")
        report = generate_community_report(
            community_id=cid,
            df_nodes=df_nodes, df_rels=df_rels, df_summary=df_summary,
            nodes_dict=nodes_dict, policy_text=policy_text,
            perspective_key=perspective_key, perspective_name=perspective_name,
        )
        reports.append(report)
        print(f"    ✓ {report['title'][:45]}...")
        if idx < len(community_ids) - 1:
            time.sleep(REQUEST_INTERVAL)

    return reports


# ══════════════════════════════════════════════════════════════
# Excel 导出
# ══════════════════════════════════════════════════════════════

EXCEL_COLUMNS = [
    ('id',                 18),
    ('community',          12),
    ('level',               8),
    ('title',              40),
    ('parent',             10),
    ('children',           12),
    ('summary',            60),
    ('key_words',          35),
    ('findings',           70),
    ('rank',               10),
    ('rating_explanation', 45),
    ('full_content_json',  80),
]


def save_reports_to_excel(reports: List[Dict], output_path: str, perspective_name: str):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{perspective_name}社区报告"

    header_font  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill('solid', fgColor='2F5496')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border  = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'),
    )

    for col_i, (col_name, col_width) in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_i, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_i)].width = col_width

    data_font  = Font(name='Arial', size=10)
    data_align = Alignment(vertical='top', wrap_text=True)
    alt_fill   = PatternFill('solid', fgColor='F2F7FB')

    for row_i, report in enumerate(reports, 2):
        is_alt = row_i % 2 == 0
        for col_i, (col_name, _) in enumerate(EXCEL_COLUMNS, 1):
            cell = ws.cell(row=row_i, column=col_i, value=report.get(col_name, ''))
            cell.font      = data_font
            cell.alignment = data_align
            cell.border    = thin_border
            if is_alt:
                cell.fill  = alt_fill

    ws.freeze_panes   = 'A2'
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 30
    for row_i in range(2, len(reports) + 2):
        ws.row_dimensions[row_i].height = 80

    wb.save(output_path)
    print(f"  ✓ 已保存: {output_path} ({len(reports)} 条记录)")


# ══════════════════════════════════════════════════════════════
# 主函数
# ══════════════════════════════════════════════════════════════

def main():
    ensure_output_dirs()
    print("=" * 60)
    print("  步骤5：社区报告生成器")
    print("  基于 GNN Leiden 聚类结果 + DeepSeek API")
    print("=" * 60)

    if not API_KEY:
        print("\n⚠ 未设置 DEEPSEEK_API_KEY，将使用回退方案")

    # 检查聚类结果
    if not os.path.exists(CLUSTERING_RESULT_DIR):
        print(f"\n未找到聚类结果目录: {CLUSTERING_RESULT_DIR}")
        print("请先运行 step1_gnn_clustering.py")
        sys.exit(1)

    # 加载图谱数据
    print("\n[1/4] 加载知识图谱数据...")
    nodes_dict, _ = load_graph_data(GRAPH_DATA_FILE)

    # 加载政策原文
    print("\n[2/4] 加载政策法规原文...")
    policy_text = load_policy_documents(POLICY_DOC_DIR)

    # 生成报告
    print("\n[3/4] 为各视角生成社区报告...")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    all_reports = {}

    for pk, pname in PERSPECTIVES.items():
        check = os.path.join(CLUSTERING_RESULT_DIR, f"{pk}_1_节点聚类结果.csv")
        if not os.path.exists(check):
            print(f"\n  ⚠ 跳过 {pname}：聚类文件不存在")
            continue
        all_reports[pk] = generate_reports_for_perspective(
            perspective_key=pk, perspective_name=pname,
            result_dir=CLUSTERING_RESULT_DIR,
            nodes_dict=nodes_dict, policy_text=policy_text,
        )

    # 导出 Excel
    print(f"\n[4/4] 导出 Excel 报告...")
    for pk, reports in all_reports.items():
        pname = PERSPECTIVES[pk]
        out_path = os.path.join(OUTPUT_DIR, f"{pname}社区报告.xlsx")
        save_reports_to_excel(reports, out_path, pname)

    # 完成汇总
    print(f"\n{'=' * 60}")
    print("  步骤5 完成！")
    print(f"  输出目录: {OUTPUT_DIR}")
    for pk, reports in all_reports.items():
        print(f"    {PERSPECTIVES[pk]}社区报告.xlsx  ({len(reports)} 个社区)")
    print(f"{'=' * 60}")
    return all_reports


if __name__ == '__main__':
    main()
